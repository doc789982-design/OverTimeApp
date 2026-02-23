# -*- coding: utf-8 -*-
"""
OvertimeTab (один файл)
Python + PySide6 + SQLite
Экспорт в Excel: openpyxl

pip install PySide6 openpyxl
python overtime_app_all.py
"""

from __future__ import annotations

import json
import os
import shutil
import sqlite3
import sys
import uuid
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Callable, Optional

from PySide6.QtCore import Qt, QDate, QByteArray, QTimer, QPoint, QEvent
from PySide6.QtGui import QColor, QBrush, QTextCharFormat, QPalette, QAction
from PySide6.QtWidgets import (
    QApplication,
    QAbstractSpinBox,
    QCalendarWidget,
    QCheckBox,
    QComboBox,
    QDateEdit,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMenu,
    QMenuBar,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QSpinBox,
    QSplitter,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QTableView,
)

APP_NAME = "OvertimeTab"
SCHEMA_VERSION = 2  # opening_minutes/opening_days


# -----------------------------
# Utils
# -----------------------------
def dt_parse(s: str) -> datetime:
    return datetime.fromisoformat(s)


def dt_iso(dt: datetime) -> str:
    return dt.isoformat(timespec="minutes")


def d_parse(s: str) -> date:
    return date.fromisoformat(s)


def d_iso(d: date) -> str:
    return d.isoformat()


def fmt_date_iso(iso: str) -> str:
    try:
        return d_parse(iso).strftime("%d.%m.%Y")
    except Exception:
        return iso or ""


def fmt_dt_iso(iso: str) -> str:
    try:
        return dt_parse(iso).strftime("%d.%m.%Y %H:%M")
    except Exception:
        return iso or ""


def parse_month(s: str) -> tuple[int, int]:
    y, m = s.split("-")
    return int(y), int(m)


def next_month(year: int, month: int) -> tuple[int, int]:
    return (year + 1, 1) if month == 12 else (year, month + 1)


def month_bounds_dt(year: int, month: int) -> tuple[datetime, datetime]:
    start = datetime(year, month, 1, 0, 0)
    ny, nm = next_month(year, month)
    end = datetime(ny, nm, 1, 0, 0)
    return start, end


def year_bounds_dt(year: int) -> tuple[datetime, datetime]:
    return datetime(year, 1, 1, 0, 0), datetime(year + 1, 1, 1, 0, 0)


def minutes_to_hhmm(m: int) -> str:
    sign = "-" if m < 0 else ""
    m = abs(m)
    h = m // 60
    mm = m % 60
    return f"{sign}{h}:{mm:02d}"


def qdate_to_date(qd: QDate) -> date:
    return date(qd.year(), qd.month(), qd.day())


def intersect(a0: datetime, a1: datetime, b0: datetime, b1: datetime) -> Optional[tuple[datetime, datetime]]:
    s = max(a0, b0)
    e = min(a1, b1)
    return (s, e) if s < e else None


def merge_intervals(intervals: list[tuple[datetime, datetime]]) -> list[tuple[datetime, datetime]]:
    if not intervals:
        return []
    intervals.sort(key=lambda x: x[0])
    merged = [intervals[0]]
    for s, e in intervals[1:]:
        ps, pe = merged[-1]
        if s <= pe:
            merged[-1] = (ps, max(pe, e))
        else:
            merged.append((s, e))
    return merged


def yyyymm_from_end_date(end_date_iso: str) -> str:
    return (end_date_iso or "")[:7]


def is_emp_active_in_month(emp: sqlite3.Row, year: int, month: int) -> bool:
    m = f"{year:04d}-{month:02d}"
    if emp["start_month"] > m:
        return False
    if emp["end_date"] is None:
        return True
    return yyyymm_from_end_date(emp["end_date"]) >= m


def safe_int(x: Any, default: int) -> int:
    try:
        return int(x)
    except Exception:
        return default


# -----------------------------
# Theme
# -----------------------------
def make_dark_palette() -> QPalette:
    p = QPalette()
    p.setColor(QPalette.Window, QColor(30, 30, 30))
    p.setColor(QPalette.WindowText, QColor(230, 230, 230))
    p.setColor(QPalette.Base, QColor(22, 22, 22))
    p.setColor(QPalette.AlternateBase, QColor(35, 35, 35))
    p.setColor(QPalette.ToolTipBase, QColor(230, 230, 230))
    p.setColor(QPalette.ToolTipText, QColor(30, 30, 30))
    p.setColor(QPalette.Text, QColor(230, 230, 230))
    p.setColor(QPalette.Button, QColor(45, 45, 45))
    p.setColor(QPalette.ButtonText, QColor(230, 230, 230))
    p.setColor(QPalette.Link, QColor(80, 170, 255))
    p.setColor(QPalette.Highlight, QColor(60, 120, 200))
    p.setColor(QPalette.HighlightedText, QColor(255, 255, 255))
    return p


def apply_theme(app: QApplication, theme: str) -> None:
    theme = (theme or "system").lower()
    if theme == "dark":
        app.setPalette(make_dark_palette())
    elif theme == "light":
        app.setPalette(QPalette())
    else:
        app.setPalette(QPalette())


# -----------------------------
# UI: visual spinners / pickers
# -----------------------------
def configure_spinbox(widget: QAbstractSpinBox) -> None:
    widget.setButtonSymbols(QAbstractSpinBox.UpDownArrows)
    widget.setAccelerated(True)
    widget.setKeyboardTracking(False)
    le = widget.lineEdit()
    if le is not None:
        le.setReadOnly(True)
        le.setCursor(Qt.ArrowCursor)


class TwoDigitSpinBox(QSpinBox):
    def textFromValue(self, v: int) -> str:  # type: ignore[override]
        return f"{v:02d}"

    def valueFromText(self, text: str) -> int:  # type: ignore[override]
        try:
            return int(text)
        except Exception:
            return 0


class DurationPicker(QWidget):
    """Длительность: часы (0..999) + минуты (00..59)"""

    def __init__(self, max_hours: int = 999, parent: QWidget | None = None):
        super().__init__(parent)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)

        self.sp_h = QSpinBox()
        self.sp_h.setRange(0, max_hours)
        configure_spinbox(self.sp_h)

        self.sp_m = TwoDigitSpinBox()
        self.sp_m.setRange(0, 59)
        configure_spinbox(self.sp_m)

        lay.addWidget(self.sp_h)
        lay.addWidget(QLabel(":"))
        lay.addWidget(self.sp_m)
        lay.addStretch(1)

    def set_minutes_total(self, minutes: int) -> None:
        minutes = max(0, int(minutes))
        self.sp_h.setValue(minutes // 60)
        self.sp_m.setValue(minutes % 60)

    def minutes_total(self) -> int:
        return int(self.sp_h.value()) * 60 + int(self.sp_m.value())


class TimeOfDayPicker(QWidget):
    """Время суток: 00..23 : 00..59"""

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)

        self.sp_h = TwoDigitSpinBox()
        self.sp_h.setRange(0, 23)
        configure_spinbox(self.sp_h)

        self.sp_m = TwoDigitSpinBox()
        self.sp_m.setRange(0, 59)
        configure_spinbox(self.sp_m)

        lay.addWidget(self.sp_h)
        lay.addWidget(QLabel(":"))
        lay.addWidget(self.sp_m)
        lay.addStretch(1)

    def set_time(self, t: time) -> None:
        self.sp_h.setValue(t.hour)
        self.sp_m.setValue(t.minute)

    def get_time(self) -> time:
        return time(int(self.sp_h.value()), int(self.sp_m.value()))


class DateTimePicker(QWidget):
    """Дата (через popup) + визуальные крутилки времени"""

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)

        self.de = QDateEdit()
        self.de.setCalendarPopup(True)
        self.de.setDate(QDate.currentDate())
        configure_spinbox(self.de)

        self.tp = TimeOfDayPicker()

        lay.addWidget(self.de)
        lay.addWidget(self.tp)
        lay.addStretch(1)

    def set_datetime(self, dt: datetime) -> None:
        self.de.setDate(QDate(dt.year, dt.month, dt.day))
        self.tp.set_time(dt.time().replace(second=0, microsecond=0))

    def get_datetime(self) -> datetime:
        d0 = qdate_to_date(self.de.date())
        t0 = self.tp.get_time()
        return datetime(d0.year, d0.month, d0.day, t0.hour, t0.minute)


class NoWheelCalendarWidget(QCalendarWidget):
    """Запрет листания месяцев колесиком мыши."""

    def wheelEvent(self, e):  # type: ignore[override]
        e.ignore()


# -----------------------------
# Snackbar / Undo
# -----------------------------
class Snackbar(QFrame):
    def __init__(self, parent: QWidget):
        super().__init__(parent)
        self.setObjectName("Snackbar")
        self.setFrameShape(QFrame.StyledPanel)
        self.setStyleSheet(
            """
            QFrame#Snackbar {
                background: rgba(40, 40, 40, 230);
                border-radius: 10px;
            }
            QFrame#Snackbar QLabel { color: white; }
            QFrame#Snackbar QPushButton {
                color: #9ecbff;
                background: transparent;
                border: none;
                font-weight: 600;
            }
        """
        )
        self.hide()

        self._action: Optional[Callable[[], None]] = None
        self._timer = QTimer(self)
        self._timer.setSingleShot(True)
        self._timer.timeout.connect(self.hide)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(14, 10, 14, 10)
        self.lbl = QLabel("")
        self.btn = QPushButton("")
        self.btn.hide()
        self.btn.clicked.connect(self._on_action)
        lay.addWidget(self.lbl, 1)
        lay.addWidget(self.btn, 0)

    def _on_action(self) -> None:
        self.hide()
        if self._action:
            self._action()

    def show_message(
        self,
        text: str,
        duration_ms: int = 4000,
        action_text: str | None = None,
        action: Callable[[], None] | None = None,
    ) -> None:
        self.lbl.setText(text)
        self._action = action
        if action_text and action:
            self.btn.setText(action_text)
            self.btn.show()
        else:
            self.btn.hide()
        self.adjustSize()
        self.show()
        self.raise_()
        self._timer.start(duration_ms)


# -----------------------------
# App config
# -----------------------------
def app_data_dir() -> Path:
    base = Path.home() / f".{APP_NAME.lower()}"
    base.mkdir(parents=True, exist_ok=True)
    return base


def app_db_store_dir() -> Path:
    d = app_data_dir() / "databases"
    d.mkdir(parents=True, exist_ok=True)
    return d


def config_path() -> Path:
    return app_data_dir() / "config.json"


@dataclass
class AppConfig:
    db_paths: list[str]
    last_db_path: Optional[str] = None
    ui: dict[str, Any] = None  # type: ignore

    @staticmethod
    def load() -> "AppConfig":
        p = config_path()
        if not p.exists():
            return AppConfig(db_paths=[], last_db_path=None, ui={})
        data = json.loads(p.read_text(encoding="utf-8"))
        return AppConfig(
            db_paths=data.get("db_paths", []),
            last_db_path=data.get("last_db_path"),
            ui=data.get("ui", {}) or {},
        )

    def save(self) -> None:
        config_path().write_text(
            json.dumps(
                {"db_paths": self.db_paths, "last_db_path": self.last_db_path, "ui": self.ui},
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )


# -----------------------------
# DB
# -----------------------------
class DBError(Exception):
    pass


class CalendarMissingError(DBError):
    pass


class DB:
    def __init__(self, path: str):
        self.path = path
        self.conn = sqlite3.connect(path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys = ON;")
        self._init_or_migrate()

    def close(self) -> None:
        self.conn.close()

    def _has_table(self, name: str) -> bool:
        r = self.conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (name,)).fetchone()
        return r is not None

    def _init_or_migrate(self) -> None:
        if not self._has_table("meta"):
            self._init_schema()
            return
        ver = self.get_meta_int("schema_version", 0)
        if ver == SCHEMA_VERSION:
            return
        if ver == 1 and SCHEMA_VERSION == 2:
            self._migrate_1_to_2()
            return
        raise DBError(f"Схема базы версии {ver}, ожидается {SCHEMA_VERSION}. Миграции не реализованы.")

    def _migrate_1_to_2(self) -> None:
        self.conn.execute("BEGIN;")
        try:
            self.conn.execute("ALTER TABLE employee ADD COLUMN opening_minutes INTEGER NOT NULL DEFAULT 0;")
            self.conn.execute("ALTER TABLE employee ADD COLUMN opening_days INTEGER NOT NULL DEFAULT 0;")
            self.set_meta("schema_version", "2")
            self.conn.execute("COMMIT;")
        except Exception:
            self.conn.execute("ROLLBACK;")
            raise

    def _init_schema(self) -> None:
        c = self.conn
        c.execute("BEGIN;")
        try:
            c.execute("CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT NOT NULL);")
            c.execute(
                """
                CREATE TABLE department_settings (
                    id INTEGER PRIMARY KEY CHECK (id=1),
                    department_name TEXT NOT NULL,
                    resp_position TEXT,
                    resp_rank TEXT,
                    resp_last_name TEXT,
                    resp_first_name TEXT,
                    resp_middle_name TEXT
                );
                """
            )
            c.execute(
                """
                CREATE TABLE employee (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    last_name TEXT NOT NULL,
                    first_name TEXT NOT NULL,
                    middle_name TEXT,
                    rank TEXT,
                    position TEXT,
                    start_month TEXT NOT NULL,
                    end_date TEXT,
                    end_reason TEXT,
                    opening_minutes INTEGER NOT NULL DEFAULT 0,
                    opening_days INTEGER NOT NULL DEFAULT 0
                );
                """
            )
            c.execute(
                """
                CREATE TABLE duty (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER NOT NULL REFERENCES employee(id) ON DELETE CASCADE,
                    start_dt TEXT NOT NULL,
                    end_dt TEXT NOT NULL,
                    comment TEXT
                );
                """
            )
            c.execute(
                """
                CREATE TABLE compensation (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER NOT NULL REFERENCES employee(id) ON DELETE CASCADE,
                    unit TEXT NOT NULL,
                    method TEXT NOT NULL,
                    event_date TEXT,
                    amount_minutes INTEGER,
                    amount_days INTEGER,
                    order_no TEXT,
                    order_date TEXT,
                    comment TEXT
                );
                """
            )
            c.execute(
                """
                CREATE TABLE comp_day_off_date (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    compensation_id INTEGER NOT NULL REFERENCES compensation(id) ON DELETE CASCADE,
                    employee_id INTEGER NOT NULL REFERENCES employee(id) ON DELETE CASCADE,
                    day_off_date TEXT NOT NULL,
                    UNIQUE(employee_id, day_off_date)
                );
                """
            )
            c.execute(
                """
                CREATE TABLE calendar_day (
                    date TEXT PRIMARY KEY,
                    is_working INTEGER NOT NULL
                );
                """
            )

            self.set_meta("schema_version", str(SCHEMA_VERSION))
            self.set_meta("db_uuid", str(uuid.uuid4()))
            self.set_meta("created_at", datetime.now().isoformat(timespec="seconds"))
            c.execute("INSERT INTO department_settings (id, department_name) VALUES (1, ?)", ("Отдел",))
            c.execute("COMMIT;")
        except Exception:
            c.execute("ROLLBACK;")
            raise

    def get_meta(self, key: str, default: Optional[str] = None) -> Optional[str]:
        r = self.conn.execute("SELECT value FROM meta WHERE key=?", (key,)).fetchone()
        return r["value"] if r else default

    def get_meta_int(self, key: str, default: int = 0) -> int:
        v = self.get_meta(key)
        return int(v) if v is not None else default

    def set_meta(self, key: str, value: str) -> None:
        self.conn.execute(
            "INSERT INTO meta(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, value),
        )

    def get_department_name(self) -> str:
        r = self.conn.execute("SELECT department_name FROM department_settings WHERE id=1").fetchone()
        return r["department_name"] if r else "Отдел"

    def get_department_settings(self) -> sqlite3.Row:
        return self.conn.execute("SELECT * FROM department_settings WHERE id=1").fetchone()

    def update_department_settings(self, **fields: Any) -> None:
        if not fields:
            return
        cols = ", ".join([f"{k}=?" for k in fields.keys()])
        vals = list(fields.values())
        self.conn.execute(f"UPDATE department_settings SET {cols} WHERE id=1", vals)

    # calendar
    def calendar_year_complete(self, year: int) -> bool:
        y0 = f"{year:04d}-01-01"
        y1 = f"{year:04d}-12-31"
        r = self.conn.execute("SELECT COUNT(*) AS c FROM calendar_day WHERE date BETWEEN ? AND ?", (y0, y1)).fetchone()
        days = 366 if date(year, 12, 31).timetuple().tm_yday == 366 else 365
        return int(r["c"]) == days

    def create_calendar_year_default(self, year: int) -> None:
        y_start = date(year, 1, 1)
        y_end = date(year, 12, 31)
        self.conn.execute("DELETE FROM calendar_day WHERE date BETWEEN ? AND ?", (d_iso(y_start), d_iso(y_end)))
        rows = []
        cur = y_start
        while cur <= y_end:
            is_working = 1 if cur.weekday() < 5 else 0
            rows.append((d_iso(cur), is_working))
            cur += timedelta(days=1)
        self.conn.executemany("INSERT INTO calendar_day(date,is_working) VALUES (?,?)", rows)

    def is_working_day(self, d0: date) -> bool:
        r = self.conn.execute("SELECT is_working FROM calendar_day WHERE date=?", (d_iso(d0),)).fetchone()
        if not r:
            raise CalendarMissingError(f"Календарь не создан/неполный для даты {d_iso(d0)}")
        return bool(r["is_working"])

    def toggle_calendar_day(self, d0: date) -> None:
        r = self.conn.execute("SELECT is_working FROM calendar_day WHERE date=?", (d_iso(d0),)).fetchone()
        if not r:
            raise CalendarMissingError("Календарь года не создан.")
        new_val = 0 if r["is_working"] else 1
        self.conn.execute("UPDATE calendar_day SET is_working=? WHERE date=?", (new_val, d_iso(d0)))

    # employees
    def list_employees_for_month(self, year: int, month: int, active_only: bool, search: str) -> list[sqlite3.Row]:
        m = f"{year:04d}-{month:02d}"
        search = (search or "").strip().lower()
        params: list[Any] = []
        where = []
        if active_only:
            where.append("start_month <= ?")
            where.append("(end_date IS NULL OR substr(end_date,1,7) >= ?)")
            params += [m, m]
        if search:
            where.append(
                "(lower(last_name) LIKE ? OR lower(first_name) LIKE ? OR lower(middle_name) LIKE ? OR lower(rank) LIKE ? OR lower(position) LIKE ?)"
            )
            s = f"%{search}%"
            params += [s, s, s, s, s]
        sql = "SELECT * FROM employee"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY last_name, first_name, middle_name"
        return self.conn.execute(sql, params).fetchall()

    def get_employee(self, employee_id: int) -> sqlite3.Row:
        r = self.conn.execute("SELECT * FROM employee WHERE id=?", (employee_id,)).fetchone()
        if not r:
            raise DBError("Сотрудник не найден.")
        return r

    def add_employee(
        self,
        last: str,
        first: str,
        middle: str,
        rank: str,
        position: str,
        start_month: str,
        opening_minutes: int,
        opening_days: int,
    ) -> int:
        cur = self.conn.execute(
            """
            INSERT INTO employee(last_name,first_name,middle_name,rank,position,start_month,opening_minutes,opening_days)
            VALUES (?,?,?,?,?,?,?,?)
            """,
            (last, first, middle or None, rank or None, position or None, start_month, opening_minutes, opening_days),
        )
        return int(cur.lastrowid)

    def update_employee(self, employee_id: int, **fields: Any) -> None:
        if not fields:
            return
        cols = ", ".join([f"{k}=?" for k in fields.keys()])
        vals = list(fields.values()) + [employee_id]
        self.conn.execute(f"UPDATE employee SET {cols} WHERE id=?", vals)

    def delete_employee(self, employee_id: int) -> None:
        self.conn.execute("DELETE FROM employee WHERE id=?", (employee_id,))

    def employee_data_years(self, employee_id: int) -> set[int]:
        years: set[int] = set()
        rows = self.conn.execute(
            """
            SELECT DISTINCT substr(start_dt,1,4) AS y FROM duty WHERE employee_id=?
            UNION
            SELECT DISTINCT substr(event_date,1,4) AS y FROM compensation WHERE employee_id=? AND event_date IS NOT NULL
            UNION
            SELECT DISTINCT substr(day_off_date,1,4) AS y FROM comp_day_off_date WHERE employee_id=?
            """,
            (employee_id, employee_id, employee_id),
        ).fetchall()
        for r in rows:
            if r["y"]:
                years.add(int(r["y"]))
        return years

    def has_records_after_date(self, employee_id: int, end_date: date) -> bool:
        r1 = self.conn.execute(
            "SELECT 1 FROM duty WHERE employee_id=? AND substr(start_dt,1,10) > ? LIMIT 1",
            (employee_id, d_iso(end_date)),
        ).fetchone()
        if r1:
            return True
        r2 = self.conn.execute(
            "SELECT 1 FROM compensation WHERE employee_id=? AND event_date IS NOT NULL AND event_date > ? LIMIT 1",
            (employee_id, d_iso(end_date)),
        ).fetchone()
        if r2:
            return True
        r3 = self.conn.execute(
            "SELECT 1 FROM comp_day_off_date WHERE employee_id=? AND day_off_date > ? LIMIT 1",
            (employee_id, d_iso(end_date)),
        ).fetchone()
        return bool(r3)

    # duties
    def list_duties_for_month(self, employee_id: int, year: int, month: int) -> list[sqlite3.Row]:
        start, end = month_bounds_dt(year, month)
        return self.conn.execute(
            """
            SELECT * FROM duty
            WHERE employee_id=?
              AND end_dt > ?
              AND start_dt < ?
            ORDER BY start_dt
            """,
            (employee_id, dt_iso(start), dt_iso(end)),
        ).fetchall()

    def list_duties_for_period(self, employee_id: int, start: datetime, end: datetime) -> list[sqlite3.Row]:
        return self.conn.execute(
            """
            SELECT * FROM duty
            WHERE employee_id=?
              AND end_dt > ?
              AND start_dt < ?
            ORDER BY start_dt
            """,
            (employee_id, dt_iso(start), dt_iso(end)),
        ).fetchall()

    def get_duty(self, duty_id: int) -> sqlite3.Row:
        r = self.conn.execute("SELECT * FROM duty WHERE id=?", (duty_id,)).fetchone()
        if not r:
            raise DBError("Дежурство не найдено.")
        return r

    def add_duty(self, employee_id: int, start: datetime, end: datetime, comment: str) -> int:
        cur = self.conn.execute(
            "INSERT INTO duty(employee_id,start_dt,end_dt,comment) VALUES (?,?,?,?)",
            (employee_id, dt_iso(start), dt_iso(end), comment or None),
        )
        return int(cur.lastrowid)

    def update_duty(self, duty_id: int, **fields: Any) -> None:
        if not fields:
            return
        cols = ", ".join([f"{k}=?" for k in fields.keys()])
        vals = list(fields.values()) + [duty_id]
        self.conn.execute(f"UPDATE duty SET {cols} WHERE id=?", vals)

    def delete_duty(self, duty_id: int) -> None:
        self.conn.execute("DELETE FROM duty WHERE id=?", (duty_id,))

    def find_overlapping_duties(
        self,
        employee_id: int,
        start: datetime,
        end: datetime,
        exclude_duty_id: Optional[int] = None,
    ) -> list[tuple[int, datetime, datetime]]:
        params: list[Any] = [employee_id, dt_iso(start), dt_iso(end)]
        sql = """
            SELECT id, start_dt, end_dt FROM duty
            WHERE employee_id=?
              AND end_dt > ?
              AND start_dt < ?
        """
        if exclude_duty_id is not None:
            sql += " AND id <> ?"
            params.append(exclude_duty_id)
        rows = self.conn.execute(sql, params).fetchall()
        return [(int(r["id"]), dt_parse(r["start_dt"]), dt_parse(r["end_dt"])) for r in rows]

    def snapshot_duty(self, duty_id: int) -> dict[str, Any]:
        r = self.get_duty(duty_id)
        return {
            "employee_id": int(r["employee_id"]),
            "start_dt": r["start_dt"],
            "end_dt": r["end_dt"],
            "comment": r["comment"] or "",
        }

    def restore_duty(self, snap: dict[str, Any]) -> int:
        return self.add_duty(
            snap["employee_id"],
            dt_parse(snap["start_dt"]),
            dt_parse(snap["end_dt"]),
            snap.get("comment", ""),
        )

    # compensations
    def list_compensations_for_month(self, employee_id: int, year: int, month: int) -> list[sqlite3.Row]:
        start = date(year, month, 1)
        ny, nm = next_month(year, month)
        end = date(ny, nm, 1)
        return self.conn.execute(
            """
            SELECT * FROM compensation
            WHERE employee_id=?
              AND (
                (event_date IS NOT NULL AND event_date >= ? AND event_date < ?)
                OR
                (unit='days' AND method='day_off' AND id IN (
                    SELECT DISTINCT compensation_id FROM comp_day_off_date
                    WHERE employee_id=? AND day_off_date >= ? AND day_off_date < ?
                ))
              )
            ORDER BY COALESCE(event_date,'9999-12-31'), id
            """,
            (employee_id, d_iso(start), d_iso(end), employee_id, d_iso(start), d_iso(end)),
        ).fetchall()

    def get_compensation(self, comp_id: int) -> sqlite3.Row:
        r = self.conn.execute("SELECT * FROM compensation WHERE id=?", (comp_id,)).fetchone()
        if not r:
            raise DBError("Компенсация не найдена.")
        return r

    def get_comp_dates(self, compensation_id: int) -> list[str]:
        rows = self.conn.execute(
            "SELECT day_off_date FROM comp_day_off_date WHERE compensation_id=? ORDER BY day_off_date",
            (compensation_id,),
        ).fetchall()
        return [r["day_off_date"] for r in rows]

    def add_compensation_hours_dayoff(self, employee_id: int, event_date: date, minutes: int, comment: str) -> int:
        cur = self.conn.execute(
            """
            INSERT INTO compensation(employee_id,unit,method,event_date,amount_minutes,comment)
            VALUES (?,?,?,?,?,?)
            """,
            (employee_id, "hours", "day_off", d_iso(event_date), minutes, comment or None),
        )
        return int(cur.lastrowid)

    def add_compensation_days_dayoff(self, employee_id: int, dates: list[date], comment: str) -> int:
        cur = self.conn.execute(
            "INSERT INTO compensation(employee_id,unit,method,amount_days,comment) VALUES (?,?,?,?,?)",
            (employee_id, "days", "day_off", len(dates), comment or None),
        )
        comp_id = int(cur.lastrowid)
        for d0 in dates:
            self.conn.execute(
                "INSERT INTO comp_day_off_date(compensation_id,employee_id,day_off_date) VALUES (?,?,?)",
                (comp_id, employee_id, d_iso(d0)),
            )
        return comp_id

    def add_compensation_money(
        self,
        employee_id: int,
        unit: str,
        amount_minutes: int | None,
        amount_days: int | None,
        order_no: str,
        order_date: date,
        comment: str,
    ) -> int:
        cur = self.conn.execute(
            """
            INSERT INTO compensation(employee_id,unit,method,event_date,amount_minutes,amount_days,order_no,order_date,comment)
            VALUES (?,?,?,?,?,?,?,?,?)
            """,
            (
                employee_id,
                unit,
                "money",
                d_iso(order_date),
                amount_minutes,
                amount_days,
                order_no,
                d_iso(order_date),
                comment or None,
            ),
        )
        return int(cur.lastrowid)

    def update_compensation(self, comp_id: int, **fields: Any) -> None:
        if not fields:
            return
        cols = ", ".join([f"{k}=?" for k in fields.keys()])
        vals = list(fields.values()) + [comp_id]
        self.conn.execute(f"UPDATE compensation SET {cols} WHERE id=?", vals)

    def replace_comp_dayoff_dates(self, comp_id: int, employee_id: int, dates: list[date]) -> None:
        self.conn.execute("DELETE FROM comp_day_off_date WHERE compensation_id=?", (comp_id,))
        for d0 in dates:
            self.conn.execute(
                "INSERT INTO comp_day_off_date(compensation_id,employee_id,day_off_date) VALUES (?,?,?)",
                (comp_id, employee_id, d_iso(d0)),
            )
        self.update_compensation(comp_id, amount_days=len(dates))

    def delete_compensation(self, comp_id: int) -> None:
        self.conn.execute("DELETE FROM compensation WHERE id=?", (comp_id,))

    def snapshot_compensation(self, comp_id: int) -> dict[str, Any]:
        c = self.get_compensation(comp_id)
        snap = {k: c[k] for k in c.keys()}
        snap["id"] = int(c["id"])
        snap["employee_id"] = int(c["employee_id"])
        if c["unit"] == "days" and c["method"] == "day_off":
            snap["day_off_dates"] = self.get_comp_dates(int(c["id"]))
        else:
            snap["day_off_dates"] = []
        return snap

    def restore_compensation(self, snap: dict[str, Any]) -> int:
        employee_id = int(snap["employee_id"])
        unit = snap["unit"]
        method = snap["method"]
        comment = snap.get("comment") or ""
        if method == "money":
            order_no = snap.get("order_no") or ""
            order_date = d_parse(snap.get("order_date") or snap.get("event_date"))
            if unit == "hours":
                return self.add_compensation_money(
                    employee_id,
                    "hours",
                    int(snap.get("amount_minutes") or 0),
                    None,
                    order_no,
                    order_date,
                    comment,
                )
            else:
                return self.add_compensation_money(
                    employee_id,
                    "days",
                    None,
                    int(snap.get("amount_days") or 0),
                    order_no,
                    order_date,
                    comment,
                )
        else:
            if unit == "hours":
                event_date = d_parse(snap.get("event_date"))
                return self.add_compensation_hours_dayoff(
                    employee_id,
                    event_date,
                    int(snap.get("amount_minutes") or 0),
                    comment,
                )
            else:
                dates = [d_parse(x) for x in (snap.get("day_off_dates") or [])]
                return self.add_compensation_days_dayoff(employee_id, dates, comment)

    def list_comp_events_for_year(self, employee_id: int, year: int) -> list[dict[str, Any]]:
        y0 = date(year, 1, 1)
        y1 = date(year + 1, 1, 1)
        events: list[dict[str, Any]] = []

        rows = self.conn.execute(
            """
            SELECT unit, method, event_date, amount_minutes, amount_days, order_no
            FROM compensation
            WHERE employee_id=?
              AND event_date IS NOT NULL
              AND event_date >= ? AND event_date < ?
            """,
            (employee_id, d_iso(y0), d_iso(y1)),
        ).fetchall()
        for r in rows:
            d0 = d_parse(r["event_date"])
            if r["unit"] == "hours":
                events.append(
                    {
                        "date": d0,
                        "hours": int(r["amount_minutes"] or 0),
                        "days": 0,
                        "desc": f"компенсация часов ({'приказ '+(r['order_no'] or '') if r['method']=='money' else 'выходной'}) {d_iso(d0)}",
                    }
                )
            else:
                events.append(
                    {
                        "date": d0,
                        "hours": 0,
                        "days": int(r["amount_days"] or 0),
                        "desc": f"компенсация дней (приказ {(r['order_no'] or '')}) {d_iso(d0)}",
                    }
                )

        rows2 = self.conn.execute(
            """
            SELECT day_off_date
            FROM comp_day_off_date
            WHERE employee_id=?
              AND day_off_date >= ? AND day_off_date < ?
            """,
            (employee_id, d_iso(y0), d_iso(y1)),
        ).fetchall()
        for r in rows2:
            d0 = d_parse(r["day_off_date"])
            events.append({"date": d0, "hours": 0, "days": 1, "desc": f"выходной (день) {d_iso(d0)}"})

        events.sort(key=lambda e: (e["date"], e["desc"]))
        return events


# -----------------------------
# Business logic
# -----------------------------
def employee_period_bounds(emp: sqlite3.Row) -> tuple[date, Optional[date]]:
    y, m = parse_month(emp["start_month"])
    start = date(y, m, 1)
    end = d_parse(emp["end_date"]) if emp["end_date"] else None
    return start, end


def validate_event_in_employee_period(emp: sqlite3.Row, event_date: date) -> Optional[str]:
    start, end = employee_period_bounds(emp)
    if event_date < start:
        return "Дата раньше начала работы сотрудника в этом отделе."
    if end is not None and event_date > end:
        return "Дата позже перевода/увольнения сотрудника."
    return None


def opening_effective(emp: sqlite3.Row, year: int, until_exclusive_dt: datetime) -> tuple[int, int]:
    start_y, start_m = parse_month(emp["start_month"])
    start_dt = datetime(start_y, start_m, 1, 0, 0)
    if year != start_y:
        return 0, 0
    if until_exclusive_dt <= start_dt:
        return 0, 0
    return int(emp["opening_minutes"] or 0), int(emp["opening_days"] or 0)


def extract_night_intervals(s: datetime, e: datetime) -> list[tuple[datetime, datetime]]:
    out: list[tuple[datetime, datetime]] = []
    day = s.date() - timedelta(days=1)
    last_day = e.date()
    while day <= last_day:
        w0 = datetime.combine(day, time(22, 0))
        w1 = datetime.combine(day + timedelta(days=1), time(6, 0))
        inter = intersect(s, e, w0, w1)
        if inter:
            out.append(inter)
        day += timedelta(days=1)
    return out


def compute_night_minutes(db: DB, employee_id: int, start: datetime, end: datetime) -> int:
    duties = db.list_duties_for_period(employee_id, start, end)
    intervals: list[tuple[datetime, datetime]] = []
    for d in duties:
        s = max(dt_parse(d["start_dt"]), start)
        e = min(dt_parse(d["end_dt"]), end)
        if s >= e:
            continue
        intervals.extend(extract_night_intervals(s, e))
    merged = merge_intervals(intervals)
    return sum(int((e - s).total_seconds() // 60) for s, e in merged)


def compute_nonworking_days(db: DB, employee_id: int, start: datetime, end: datetime) -> int:
    duties = db.list_duties_for_period(employee_id, start, end)
    hit_dates: set[date] = set()
    for row in duties:
        ds = max(dt_parse(row["start_dt"]), start)
        de = min(dt_parse(row["end_dt"]), end)
        if ds >= de:
            continue
        cur = ds.date()
        last = de.date()
        while cur <= last:
            w0 = datetime.combine(cur, time(6, 0))
            w1 = datetime.combine(cur, time(22, 0))
            if intersect(ds, de, w0, w1) is not None:
                if not db.is_working_day(cur):
                    hit_dates.add(cur)
            cur += timedelta(days=1)
    return len(hit_dates)


def compute_accrual_hours_up_to(db: DB, employee_id: int, year: int, inclusive_date: date) -> int:
    emp = db.get_employee(employee_id)
    y_start, _ = year_bounds_dt(year)
    end_dt = datetime.combine(inclusive_date + timedelta(days=1), time(0, 0))
    open_m, _ = opening_effective(emp, year, end_dt)
    return open_m + compute_night_minutes(db, employee_id, y_start, end_dt)


def compute_accrual_days_up_to(db: DB, employee_id: int, year: int, inclusive_date: date) -> int:
    emp = db.get_employee(employee_id)
    y_start, _ = year_bounds_dt(year)
    end_dt = datetime.combine(inclusive_date + timedelta(days=1), time(0, 0))
    _, open_d = opening_effective(emp, year, end_dt)
    return open_d + compute_nonworking_days(db, employee_id, y_start, end_dt)


def validate_non_negative_over_year(db: DB, employee_id: int, year: int) -> tuple[bool, str]:
    events = db.list_comp_events_for_year(employee_id, year)
    spent_minutes = 0
    spent_days = 0
    need_check_days = any(e["days"] > 0 for e in events)

    for ev in events:
        spent_minutes += ev["hours"]
        spent_days += ev["days"]

        acc_minutes = compute_accrual_hours_up_to(db, employee_id, year, ev["date"])
        if spent_minutes > acc_minutes:
            need = spent_minutes - acc_minutes
            return False, f"Изменение не применено: после '{ev['desc']}' не хватает {minutes_to_hhmm(need)} часов."

        if need_check_days:
            try:
                acc_days = compute_accrual_days_up_to(db, employee_id, year, ev["date"])
            except CalendarMissingError:
                return False, "Календарь года не создан/неполный. Невозможно проверить баланс по дням."
            if spent_days > acc_days:
                need = spent_days - acc_days
                return False, f"Изменение не применено: после '{ev['desc']}' не хватает {need} дней."

    return True, ""


def compute_month_summary(db: DB, employee_id: int, year: int, month: int) -> dict[str, Any]:
    emp = db.get_employee(employee_id)
    m_start, m_end = month_bounds_dt(year, month)
    y_start, _ = year_bounds_dt(year)

    month_minutes = compute_night_minutes(db, employee_id, m_start, m_end)
    open_m, open_d = opening_effective(emp, year, m_end)
    ytd_minutes = open_m + compute_night_minutes(db, employee_id, y_start, m_end)

    days_ok = True
    days_err = ""
    month_days = 0
    ytd_days = 0
    try:
        month_days = compute_nonworking_days(db, employee_id, m_start, m_end)
        ytd_days = open_d + compute_nonworking_days(db, employee_id, y_start, m_end)
    except CalendarMissingError:
        days_ok = False
        days_err = "Календарь года не создан/неполный: дни за нерабочие не рассчитываются."

    m0 = m_start.date()
    m1 = m_end.date()
    y0 = date(year, 1, 1)

    r = db.conn.execute(
        """
        SELECT COALESCE(SUM(amount_minutes),0) AS m
        FROM compensation
        WHERE employee_id=?
          AND unit='hours'
          AND event_date IS NOT NULL
          AND event_date >= ? AND event_date < ?
        """,
        (employee_id, d_iso(m0), d_iso(m1)),
    ).fetchone()
    comp_month_minutes = int(r["m"])

    r = db.conn.execute(
        """
        SELECT COALESCE(SUM(amount_minutes),0) AS m
        FROM compensation
        WHERE employee_id=?
          AND unit='hours'
          AND event_date IS NOT NULL
          AND event_date >= ? AND event_date < ?
        """,
        (employee_id, d_iso(y0), d_iso(m1)),
    ).fetchone()
    comp_ytd_minutes = int(r["m"])

    r = db.conn.execute(
        """
        SELECT COALESCE(SUM(amount_days),0) AS d
        FROM compensation
        WHERE employee_id=?
          AND unit='days' AND method='money'
          AND event_date IS NOT NULL
          AND event_date >= ? AND event_date < ?
        """,
        (employee_id, d_iso(m0), d_iso(m1)),
    ).fetchone()
    comp_month_days_money = int(r["d"])

    r = db.conn.execute(
        """
        SELECT COALESCE(SUM(amount_days),0) AS d
        FROM compensation
        WHERE employee_id=?
          AND unit='days' AND method='money'
          AND event_date IS NOT NULL
          AND event_date >= ? AND event_date < ?
        """,
        (employee_id, d_iso(y0), d_iso(m1)),
    ).fetchone()
    comp_ytd_days_money = int(r["d"])

    r = db.conn.execute(
        "SELECT COUNT(*) AS c FROM comp_day_off_date WHERE employee_id=? AND day_off_date >= ? AND day_off_date < ?",
        (employee_id, d_iso(m0), d_iso(m1)),
    ).fetchone()
    comp_month_days_dayoff = int(r["c"])

    r = db.conn.execute(
        "SELECT COUNT(*) AS c FROM comp_day_off_date WHERE employee_id=? AND day_off_date >= ? AND day_off_date < ?",
        (employee_id, d_iso(y0), d_iso(m1)),
    ).fetchone()
    comp_ytd_days_dayoff = int(r["c"])

    comp_month_days = comp_month_days_money + comp_month_days_dayoff
    comp_ytd_days = comp_ytd_days_money + comp_ytd_days_dayoff

    bal_minutes = ytd_minutes - comp_ytd_minutes
    bal_days = ytd_days - comp_ytd_days

    return {
        "days_ok": days_ok,
        "days_err": days_err,
        "opening_minutes": int(emp["opening_minutes"] or 0),
        "opening_days": int(emp["opening_days"] or 0),
        "month_minutes": month_minutes,
        "month_days": month_days,
        "comp_month_minutes": comp_month_minutes,
        "comp_month_days": comp_month_days,
        "bal_minutes": bal_minutes,
        "bal_days": bal_days,
    }


# -----------------------------
# Export (Excel)
# -----------------------------
def ensure_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore

        return openpyxl
    except Exception as e:
        raise DBError("Для экспорта нужен пакет openpyxl. Установите: pip install openpyxl") from e


def export_simple_xlsx(db: DB, employee_id: int, year: int, month: int, out_path: str) -> None:
    openpyxl = ensure_openpyxl()
    from openpyxl.styles import Font

    emp = db.get_employee(employee_id)
    summ = compute_month_summary(db, employee_id, year, month)
    dept = db.get_department_settings()

    duties = db.list_duties_for_month(employee_id, year, month)
    comps = db.list_compensations_for_month(employee_id, year, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Итоги"

    fio = f"{emp['last_name']} {emp['first_name']} {emp['middle_name'] or ''}".strip()
    resp_fio = " ".join([x for x in [dept["resp_last_name"], dept["resp_first_name"], dept["resp_middle_name"]] if x]).strip()

    ws["A1"] = "Отдел:"
    ws["B1"] = dept["department_name"]
    ws["A2"] = "Сотрудник:"
    ws["B2"] = fio
    ws["A3"] = "Звание:"
    ws["B3"] = emp["rank"] or ""
    ws["A4"] = "Должность:"
    ws["B4"] = emp["position"] or ""
    ws["A5"] = "Период:"
    ws["B5"] = f"{year:04d}-{month:02d}"

    ws["A7"] = "Начислено за месяц (часы):"
    ws["B7"] = minutes_to_hhmm(summ["month_minutes"])
    ws["A8"] = "Начислено за месяц (дни):"
    ws["B8"] = summ["month_days"] if summ["days_ok"] else ""
    ws["A9"] = "Списано за месяц (часы):"
    ws["B9"] = minutes_to_hhmm(summ["comp_month_minutes"])
    ws["A10"] = "Списано за месяц (дни):"
    ws["B10"] = summ["comp_month_days"] if summ["days_ok"] else ""
    ws["A11"] = "Остаток на конец месяца (часы):"
    ws["B11"] = minutes_to_hhmm(summ["bal_minutes"])
    ws["A12"] = "Остаток на конец месяца (дни):"
    ws["B12"] = summ["bal_days"] if summ["days_ok"] else ""

    ws["A14"] = "Ответственный:"
    ws["B14"] = f"{dept['resp_position'] or ''} {dept['resp_rank'] or ''} {resp_fio}".strip()

    for cell in ["A1", "A2", "A3", "A4", "A5", "A7", "A8", "A9", "A10", "A11", "A12", "A14"]:
        ws[cell].font = Font(bold=True)

    ws2 = wb.create_sheet("Дежурства")
    ws2.append(["Начало", "Конец", "Комментарий"])
    for r in duties:
        ws2.append([fmt_dt_iso(r["start_dt"]), fmt_dt_iso(r["end_dt"]), r["comment"] or ""])

    ws3 = wb.create_sheet("Компенсации")
    ws3.append(["Тип", "Способ", "Дата/Приказ", "Количество", "Комментарий"])
    for r in comps:
        unit = r["unit"]
        method = r["method"]
        typ = "Часы" if unit == "hours" else "Дни"
        mth = "Выходной" if method == "day_off" else "Деньги"
        when = ""
        amount = ""
        if method == "money":
            when = f"№{r['order_no']} от {fmt_date_iso(r['order_date'])}"
            amount = minutes_to_hhmm(int(r["amount_minutes"] or 0)) if unit == "hours" else str(int(r["amount_days"] or 0))
        else:
            if unit == "hours":
                when = fmt_date_iso(r["event_date"] or "")
                amount = minutes_to_hhmm(int(r["amount_minutes"] or 0))
            else:
                dates = db.get_comp_dates(int(r["id"]))
                when = ", ".join([fmt_date_iso(x) for x in dates])
                amount = str(len(dates))
        ws3.append([typ, mth, when, amount, r["comment"] or ""])

    wb.save(out_path)


# -----------------------------
# Dialogs helpers
# -----------------------------
def simple_text_input(parent: QWidget, title: str, prompt: str) -> tuple[str, bool]:
    dlg = QDialog(parent)
    dlg.setWindowTitle(title)
    lay = QVBoxLayout(dlg)
    lay.addWidget(QLabel(prompt))
    edit = QLineEdit()
    lay.addWidget(edit)
    bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
    lay.addWidget(bb)
    bb.accepted.connect(dlg.accept)
    bb.rejected.connect(dlg.reject)
    ok = dlg.exec() == QDialog.Accepted
    return edit.text(), ok


class BaseSelectorDialog(QDialog):
    def __init__(self, cfg: AppConfig, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle("Выбор базы (отдела)")
        self.cfg = cfg
        self.selected_path: Optional[str] = None

        layout = QVBoxLayout(self)
        self.list = QListWidget()
        layout.addWidget(self.list)

        btns = QHBoxLayout()
        self.btn_create = QPushButton("Создать базу отдела…")
        self.btn_attach = QPushButton("Подключить существующую…")
        self.btn_import_copy = QPushButton("Импортировать (копировать)…")
        self.btn_export_copy = QPushButton("Экспорт копии базы…")
        self.btn_open = QPushButton("Открыть")
        self.btn_remove = QPushButton("Убрать из списка")
        btns.addWidget(self.btn_create)
        btns.addWidget(self.btn_attach)
        btns.addWidget(self.btn_import_copy)
        btns.addWidget(self.btn_export_copy)
        btns.addStretch(1)
        btns.addWidget(self.btn_remove)
        btns.addWidget(self.btn_open)
        layout.addLayout(btns)

        self.btn_create.clicked.connect(self.on_create)
        self.btn_attach.clicked.connect(self.on_attach)
        self.btn_import_copy.clicked.connect(self.on_import_copy)
        self.btn_export_copy.clicked.connect(self.on_export_copy)
        self.btn_open.clicked.connect(self.on_open)
        self.btn_remove.clicked.connect(self.on_remove)

        self.refresh()

    def refresh(self) -> None:
        self.list.clear()
        self.cfg.db_paths = [p for p in self.cfg.db_paths if os.path.exists(p)]
        for p in self.cfg.db_paths:
            name = p
            try:
                db = DB(p)
                name = db.get_department_name()
                db.close()
            except Exception:
                name = f"(неизвестная база) {Path(p).name}"
            it = QListWidgetItem(f"{name}\n{p}")
            it.setData(Qt.UserRole, p)
            self.list.addItem(it)

        if self.cfg.last_db_path:
            for i in range(self.list.count()):
                if self.list.item(i).data(Qt.UserRole) == self.cfg.last_db_path:
                    self.list.setCurrentRow(i)
                    break

    def current_path(self) -> Optional[str]:
        it = self.list.currentItem()
        return it.data(Qt.UserRole) if it else None

    def on_create(self) -> None:
        name, ok = simple_text_input(self, "Название отдела", "Введите название отдела:")
        if not ok or not name.strip():
            return
        default_dir = str(app_db_store_dir())
        path, _ = QFileDialog.getSaveFileName(
            self, "Создать базу отдела", os.path.join(default_dir, f"{name}.sqlite"), "SQLite (*.sqlite)"
        )
        if not path:
            return
        if not path.endswith(".sqlite"):
            path += ".sqlite"
        if os.path.exists(path):
            QMessageBox.warning(self, "Ошибка", "Файл уже существует.")
            return

        db = DB(path)
        db.update_department_settings(department_name=name.strip())
        db.conn.commit()
        db.close()

        if path not in self.cfg.db_paths:
            self.cfg.db_paths.append(path)
        self.cfg.last_db_path = path
        self.cfg.save()
        self.refresh()

    def on_attach(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Подключить базу", "", "SQLite (*.sqlite *.db);;All files (*.*)")
        if not path:
            return
        try:
            db = DB(path)
            _ = db.get_department_name()
            db.close()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть базу:\n{e}")
            return

        if path not in self.cfg.db_paths:
            self.cfg.db_paths.append(path)
        self.cfg.last_db_path = path
        self.cfg.save()
        self.refresh()

    def on_import_copy(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Импортировать базу (копировать)", "", "SQLite (*.sqlite *.db);;All files (*.*)"
        )
        if not path:
            return
        try:
            db = DB(path)
            name = db.get_department_name()
            db.close()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть базу:\n{e}")
            return

        dest = app_db_store_dir() / f"{name}_{uuid.uuid4().hex[:8]}.sqlite"
        shutil.copy2(path, dest)
        dest_path = str(dest)

        if dest_path not in self.cfg.db_paths:
            self.cfg.db_paths.append(dest_path)
        self.cfg.last_db_path = dest_path
        self.cfg.save()
        self.refresh()

    def on_export_copy(self) -> None:
        src = self.current_path()
        if not src:
            QMessageBox.information(self, "Экспорт", "Выберите базу из списка.")
            return
        name = Path(src).stem
        try:
            db = DB(src)
            name = db.get_department_name()
            db.close()
        except Exception:
            pass

        out, _ = QFileDialog.getSaveFileName(self, "Экспорт копии базы", f"{name}.sqlite", "SQLite (*.sqlite)")
        if not out:
            return
        if not out.endswith(".sqlite"):
            out += ".sqlite"
        shutil.copy2(src, out)
        QMessageBox.information(self, "Экспорт", "Копия базы сохранена.")

    def on_remove(self) -> None:
        p = self.current_path()
        if not p:
            return
        self.cfg.db_paths = [x for x in self.cfg.db_paths if x != p]
        if self.cfg.last_db_path == p:
            self.cfg.last_db_path = None
        self.cfg.save()
        self.refresh()

    def on_open(self) -> None:
        p = self.current_path()
        if not p:
            QMessageBox.information(self, "Выбор базы", "Выберите базу из списка.")
            return
        self.selected_path = p
        self.cfg.last_db_path = p
        self.cfg.save()
        self.accept()


class DepartmentSettingsDialog(QDialog):
    def __init__(self, db: DB, parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Настройки отдела")
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        s = db.get_department_settings()
        self.ed_name = QLineEdit(s["department_name"])
        form.addRow("Название отдела:", self.ed_name)

        form.addRow(QLabel("<b>Ответственный за ведение табеля</b>"))
        self.ed_pos = QLineEdit(s["resp_position"] or "")
        self.ed_rank = QLineEdit(s["resp_rank"] or "")
        self.ed_last = QLineEdit(s["resp_last_name"] or "")
        self.ed_first = QLineEdit(s["resp_first_name"] or "")
        self.ed_mid = QLineEdit(s["resp_middle_name"] or "")
        form.addRow("Должность:", self.ed_pos)
        form.addRow("Звание:", self.ed_rank)
        form.addRow("Фамилия:", self.ed_last)
        form.addRow("Имя:", self.ed_first)
        form.addRow("Отчество:", self.ed_mid)

        bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        lay.addWidget(bb)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)

    def accept(self) -> None:
        self.db.conn.execute("BEGIN;")
        try:
            self.db.update_department_settings(
                department_name=self.ed_name.text().strip() or "Отдел",
                resp_position=self.ed_pos.text().strip() or None,
                resp_rank=self.ed_rank.text().strip() or None,
                resp_last_name=self.ed_last.text().strip() or None,
                resp_first_name=self.ed_first.text().strip() or None,
                resp_middle_name=self.ed_mid.text().strip() or None,
            )
            self.db.conn.execute("COMMIT;")
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))
            return
        super().accept()


class EmployeeDialog(QDialog):
    def __init__(self, title: str, default_year: int, default_month: int, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle(title)
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.ed_last = QLineEdit()
        self.ed_first = QLineEdit()
        self.ed_mid = QLineEdit()
        self.ed_rank = QLineEdit()
        self.ed_pos = QLineEdit()

        self.cb_year = QComboBox()
        self.cb_month = QComboBox()
        for y in range(date.today().year - 1, date.today().year + 6):
            self.cb_year.addItem(str(y), y)
        for m in range(1, 13):
            self.cb_month.addItem(f"{m:02d}", m)

        yi = self.cb_year.findData(default_year)
        if yi >= 0:
            self.cb_year.setCurrentIndex(yi)
        self.cb_month.setCurrentIndex(max(0, default_month - 1))

        form.addRow("Фамилия:", self.ed_last)
        form.addRow("Имя:", self.ed_first)
        form.addRow("Отчество:", self.ed_mid)
        form.addRow("Звание:", self.ed_rank)
        form.addRow("Должность:", self.ed_pos)

        row = QHBoxLayout()
        row.addWidget(self.cb_year)
        row.addWidget(self.cb_month)
        w = QWidget()
        w.setLayout(row)
        form.addRow("Месяц приема (ГГГГ-ММ):", w)

        self.opening_time = DurationPicker(max_hours=999)
        self.opening_time.set_minutes_total(0)

        self.opening_days = QSpinBox()
        self.opening_days.setRange(0, 366)
        configure_spinbox(self.opening_days)
        self.opening_days.setValue(0)

        form.addRow("Начальные часы (перенос):", self.opening_time)
        form.addRow("Начальные дни (перенос):", self.opening_days)

        self.bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        lay.addWidget(self.bb)
        self.bb.accepted.connect(self.accept)
        self.bb.rejected.connect(self.reject)

    def set_from_employee(self, emp: sqlite3.Row) -> None:
        self.ed_last.setText(emp["last_name"])
        self.ed_first.setText(emp["first_name"])
        self.ed_mid.setText(emp["middle_name"] or "")
        self.ed_rank.setText(emp["rank"] or "")
        self.ed_pos.setText(emp["position"] or "")

        y, m = parse_month(emp["start_month"])
        yi = self.cb_year.findData(y)
        if yi >= 0:
            self.cb_year.setCurrentIndex(yi)
        self.cb_month.setCurrentIndex(m - 1)

        self.opening_time.set_minutes_total(int(emp["opening_minutes"] or 0))
        self.opening_days.setValue(int(emp["opening_days"] or 0))

    def get_values(self) -> dict[str, Any]:
        y = int(self.cb_year.currentData())
        m = int(self.cb_month.currentData())
        start_month = f"{y:04d}-{m:02d}"
        return {
            "last_name": self.ed_last.text().strip(),
            "first_name": self.ed_first.text().strip(),
            "middle_name": self.ed_mid.text().strip(),
            "rank": self.ed_rank.text().strip(),
            "position": self.ed_pos.text().strip(),
            "start_month": start_month,
            "opening_minutes": self.opening_time.minutes_total(),
            "opening_days": int(self.opening_days.value()),
        }

    def accept(self) -> None:
        v = self.get_values()
        if not v["last_name"] or not v["first_name"]:
            QMessageBox.warning(self, "Проверка", "Фамилия и имя обязательны.")
            return
        super().accept()


class EndDateDialog(QDialog):
    def __init__(self, title: str, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle(title)
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.de = QDateEdit()
        self.de.setCalendarPopup(True)
        self.de.setDate(QDate.currentDate())
        configure_spinbox(self.de)

        form.addRow("Дата:", self.de)
        self.bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lay.addWidget(self.bb)
        self.bb.accepted.connect(self.accept)
        self.bb.rejected.connect(self.reject)

    def get_date(self) -> date:
        return qdate_to_date(self.de.date())


class DutyDialog(QDialog):
    def __init__(self, db: DB, employee_id: int, exclude_duty_id: Optional[int] = None, parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.employee_id = employee_id
        self.exclude_duty_id = exclude_duty_id

        self.setWindowTitle("Дежурство")
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.p_start = DateTimePicker()
        self.p_end = DateTimePicker()
        now = datetime.now().replace(second=0, microsecond=0)
        self.p_start.set_datetime(now)
        self.p_end.set_datetime(now + timedelta(hours=1))

        self.ed_comment = QLineEdit()

        form.addRow("Начало:", self.p_start)
        form.addRow("Конец:", self.p_end)
        form.addRow("Комментарий:", self.ed_comment)

        self.lbl_warn = QLabel("")
        self.lbl_warn.setStyleSheet("color:#b35a00;")
        lay.addWidget(self.lbl_warn)

        bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        lay.addWidget(bb)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)

        for w in [self.p_start.de, self.p_end.de]:
            w.dateChanged.connect(lambda _=None: self.update_warning())
        for w in [self.p_start.tp.sp_h, self.p_start.tp.sp_m, self.p_end.tp.sp_h, self.p_end.tp.sp_m]:
            w.valueChanged.connect(lambda _=None: self.update_warning())

        self.update_warning()

    def set_values(self, start: datetime, end: datetime, comment: str) -> None:
        self.p_start.set_datetime(start)
        self.p_end.set_datetime(end)
        self.ed_comment.setText(comment or "")
        self.update_warning()

    def get_values(self) -> tuple[datetime, datetime, str]:
        return self.p_start.get_datetime(), self.p_end.get_datetime(), self.ed_comment.text().strip()

    def update_warning(self) -> None:
        s, e, _ = self.get_values()
        if s >= e:
            self.lbl_warn.setText("")
            return
        overlaps = self.db.find_overlapping_duties(self.employee_id, s, e, exclude_duty_id=self.exclude_duty_id)
        if not overlaps:
            self.lbl_warn.setText("")
            return
        intervals = []
        for _, os_, oe in overlaps:
            inter = intersect(s, e, os_, oe)
            if inter:
                intervals.append(inter)
        merged = merge_intervals(intervals)
        minutes = sum(int((b - a).total_seconds() // 60) for a, b in merged)
        self.lbl_warn.setText(
            f"Есть пересечение с {len(overlaps)} дежурствами. Перекрывающееся время ({minutes_to_hhmm(minutes)}) не будет учтено в балансе."
        )

    def accept(self) -> None:
        s, e, _ = self.get_values()
        if s >= e:
            QMessageBox.warning(self, "Проверка", "Конец должен быть позже начала.")
            return
        super().accept()


class FillPeriodDialog(QDialog):
    def __init__(self, parent: QWidget | None = None, default_only_working: bool = True):
        super().__init__(parent)
        self.setWindowTitle("Заполнить периодом")
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.d1 = QDateEdit()
        self.d1.setCalendarPopup(True)
        self.d1.setDate(QDate.currentDate())
        configure_spinbox(self.d1)

        self.d2 = QDateEdit()
        self.d2.setCalendarPopup(True)
        self.d2.setDate(QDate.currentDate())
        configure_spinbox(self.d2)

        self.rb_only_working = QRadioButton("Только рабочие дни")
        self.rb_only_working.setChecked(default_only_working)

        form.addRow("Начало:", self.d1)
        form.addRow("Конец:", self.d2)
        lay.addWidget(self.rb_only_working)

        self.lbl_warn = QLabel("Внимание: выходные/праздничные дни обычно нельзя указывать как компенсационный выходной.")
        self.lbl_warn.setStyleSheet("color:#b00020;")
        lay.addWidget(self.lbl_warn)
        self.lbl_warn.setVisible(not self.rb_only_working.isChecked())
        self.rb_only_working.toggled.connect(lambda v: self.lbl_warn.setVisible(not v))

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lay.addWidget(bb)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)

    def values(self) -> tuple[date, date, bool]:
        return qdate_to_date(self.d1.date()), qdate_to_date(self.d2.date()), self.rb_only_working.isChecked()


class FillCountDialog(QDialog):
    def __init__(self, parent: QWidget | None = None, default_only_working: bool = True):
        super().__init__(parent)
        self.setWindowTitle("Заполнить по количеству")
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.d1 = QDateEdit()
        self.d1.setCalendarPopup(True)
        self.d1.setDate(QDate.currentDate())
        configure_spinbox(self.d1)

        self.sp = QSpinBox()
        self.sp.setRange(1, 366)
        self.sp.setValue(1)
        configure_spinbox(self.sp)

        self.rb_only_working = QRadioButton("Только рабочие дни")
        self.rb_only_working.setChecked(default_only_working)

        form.addRow("Старт:", self.d1)
        form.addRow("Количество N:", self.sp)
        lay.addWidget(self.rb_only_working)

        self.lbl_warn = QLabel("Внимание: выходные/праздничные дни обычно нельзя указывать как компенсационный выходной.")
        self.lbl_warn.setStyleSheet("color:#b00020;")
        lay.addWidget(self.lbl_warn)
        self.lbl_warn.setVisible(not self.rb_only_working.isChecked())
        self.rb_only_working.toggled.connect(lambda v: self.lbl_warn.setVisible(not v))

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lay.addWidget(bb)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)

    def values(self) -> tuple[date, int, bool]:
        return qdate_to_date(self.d1.date()), int(self.sp.value()), self.rb_only_working.isChecked()


class MultiDateCalendarDialog(QDialog):
    """Клик по дате = выделить/снять. Можно выбирать даты разных месяцев."""

    def __init__(self, db: DB, year: int, parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.year = year
        self.selected: set[date] = set()

        self.setWindowTitle("Выбор дат выходных")
        lay = QVBoxLayout(self)

        top = QHBoxLayout()
        self.chk_only_working = QCheckBox("Только рабочие дни")
        self.chk_only_working.setChecked(True)
        top.addWidget(self.chk_only_working)
        top.addStretch(1)
        self.lbl_counter = QLabel("Выбрано: 0")
        top.addWidget(self.lbl_counter)
        lay.addLayout(top)

        self.lbl_warn = QLabel("Внимание: выходные/праздничные дни обычно нельзя указывать как компенсационный выходной.")
        self.lbl_warn.setStyleSheet("color:#b00020;")
        lay.addWidget(self.lbl_warn)
        self.lbl_warn.setVisible(False)

        self.cal = QCalendarWidget()
        self.cal.setMinimumDate(QDate(year, 1, 1))
        self.cal.setMaximumDate(QDate(year, 12, 31))
        lay.addWidget(self.cal)

        btns = QHBoxLayout()
        self.btn_clear = QPushButton("Очистить")
        btns.addWidget(self.btn_clear)
        btns.addStretch(1)
        self.bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.addWidget(self.bb)
        lay.addLayout(btns)

        self.bb.accepted.connect(self.accept)
        self.bb.rejected.connect(self.reject)
        self.btn_clear.clicked.connect(self.clear_all)
        self.cal.clicked.connect(self.on_date_clicked)
        self.chk_only_working.toggled.connect(self.on_only_working_toggled)

        self.fmt_selected = QTextCharFormat()
        self.fmt_selected.setBackground(QBrush(QColor("#9ecbff")))

    def on_only_working_toggled(self, v: bool) -> None:
        self.lbl_warn.setVisible(not v)

    def refresh_counter(self) -> None:
        self.lbl_counter.setText(f"Выбрано: {len(self.selected)}")

    def set_selected_dates(self, dates: list[date]) -> None:
        self.selected = set(dates)
        self.repaint_formats()
        self.refresh_counter()

    def repaint_formats(self) -> None:
        d0 = date(self.year, 1, 1)
        d1 = date(self.year, 12, 31)
        cur = d0
        while cur <= d1:
            qd = QDate(cur.year, cur.month, cur.day)
            self.cal.setDateTextFormat(qd, self.fmt_selected if cur in self.selected else QTextCharFormat())
            cur += timedelta(days=1)

    def clear_all(self) -> None:
        self.selected.clear()
        self.repaint_formats()
        self.refresh_counter()

    def on_date_clicked(self, qd: QDate) -> None:
        d0 = qdate_to_date(qd)

        if self.chk_only_working.isChecked():
            try:
                if not self.db.is_working_day(d0):
                    QMessageBox.information(
                        self,
                        "Выбор дат",
                        "Это нерабочий день. Снимите галочку 'Только рабочие дни', если хотите выбрать его.",
                    )
                    return
            except CalendarMissingError as e:
                QMessageBox.warning(self, "Календарь", str(e))
                return

        if d0 in self.selected:
            self.selected.remove(d0)
            self.cal.setDateTextFormat(qd, QTextCharFormat())
        else:
            self.selected.add(d0)
            self.cal.setDateTextFormat(qd, self.fmt_selected)

        self.refresh_counter()

    def get_dates(self) -> list[date]:
        return sorted(self.selected)


# -----------------------------
# Compensation dialog
# -----------------------------
class CompensationDialog(QDialog):
    def __init__(self, db: DB, employee: sqlite3.Row, year: int, comp_id: Optional[int] = None, parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.employee = employee
        self.year = year
        self.comp_id = comp_id
        self.edit_mode = comp_id is not None

        self.setWindowTitle("Компенсация" + (" (редактирование)" if self.edit_mode else ""))
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.cb_unit = QComboBox()
        self.cb_unit.addItem("Часы", "hours")
        self.cb_unit.addItem("Дни", "days")
        self.cb_method = QComboBox()
        self.cb_method.addItem("Выходной", "day_off")
        self.cb_method.addItem("Деньги (приказ)", "money")

        form.addRow("Списать:", self.cb_unit)
        form.addRow("Способ:", self.cb_method)

        # hours day_off
        self.de_event = QDateEdit()
        self.de_event.setCalendarPopup(True)
        self.de_event.setDate(QDate.currentDate())
        configure_spinbox(self.de_event)

        self.amount_hours = DurationPicker(max_hours=999)
        self.amount_hours.set_minutes_total(60)

        # days day_off: only list of dates, count derived
        self.list_dates = QListWidget()
        self.lbl_days_count = QLabel("Количество дней: 0")
        self.btn_fill_period = QPushButton("Заполнить периодом")
        self.btn_fill_count = QPushButton("Заполнить по количеству")
        self.btn_pick_calendar = QPushButton("Выбрать в календаре…")
        btnrow = QHBoxLayout()
        btnrow.addWidget(self.btn_fill_period)
        btnrow.addWidget(self.btn_fill_count)
        btnrow.addWidget(self.btn_pick_calendar)
        w_btn = QWidget()
        w_btn.setLayout(btnrow)

        # money
        self.ed_order_no = QLineEdit()
        self.de_order_date = QDateEdit()
        self.de_order_date.setCalendarPopup(True)
        self.de_order_date.setDate(QDate.currentDate())
        configure_spinbox(self.de_order_date)

        self.sp_money_days = QSpinBox()
        self.sp_money_days.setRange(1, 366)
        self.sp_money_days.setValue(1)
        configure_spinbox(self.sp_money_days)

        self.money_hours = DurationPicker(max_hours=999)
        self.money_hours.set_minutes_total(60)

        self.ed_comment = QLineEdit()
        form.addRow("Комментарий:", self.ed_comment)

        self.grp_hours_dayoff = QGroupBox("Выходной (часы)")
        f1 = QFormLayout()
        self.grp_hours_dayoff.setLayout(f1)
        f1.addRow("Дата:", self.de_event)
        f1.addRow("Количество (чч:мм):", self.amount_hours)

        self.grp_days_dayoff = QGroupBox("Выходной (дни)")
        v2 = QVBoxLayout()
        self.grp_days_dayoff.setLayout(v2)
        v2.addWidget(self.lbl_days_count)
        v2.addWidget(self.list_dates)
        v2.addWidget(w_btn)

        self.grp_money = QGroupBox("Деньги (приказ)")
        f3 = QFormLayout()
        self.grp_money.setLayout(f3)
        f3.addRow("№ приказа:", self.ed_order_no)
        f3.addRow("Дата приказа:", self.de_order_date)
        self.lbl_money_amount = QLabel("")
        f3.addRow("Количество:", self.lbl_money_amount)
        f3.addRow("Дни:", self.sp_money_days)
        f3.addRow("Часы (чч:мм):", self.money_hours)

        lay.addWidget(self.grp_hours_dayoff)
        lay.addWidget(self.grp_days_dayoff)
        lay.addWidget(self.grp_money)

        self.bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        lay.addWidget(self.bb)
        self.bb.accepted.connect(self.accept)
        self.bb.rejected.connect(self.reject)

        self.cb_unit.currentIndexChanged.connect(self.update_visibility)
        self.cb_method.currentIndexChanged.connect(self.update_visibility)
        self.btn_fill_period.clicked.connect(self.on_fill_period)
        self.btn_fill_count.clicked.connect(self.on_fill_count)
        self.btn_pick_calendar.clicked.connect(self.on_pick_calendar)

        if self.edit_mode:
            self.load_existing()

        self.update_visibility()

        if self.edit_mode:
            self.cb_unit.setEnabled(False)
            self.cb_method.setEnabled(False)

    def update_days_count(self) -> None:
        self.lbl_days_count.setText(f"Количество дней: {self.list_dates.count()}")

    def load_existing(self) -> None:
        c = self.db.get_compensation(self.comp_id)
        unit = c["unit"]
        method = c["method"]
        self.cb_unit.setCurrentIndex(0 if unit == "hours" else 1)
        self.cb_method.setCurrentIndex(0 if method == "day_off" else 1)
        self.ed_comment.setText(c["comment"] or "")

        if method == "money":
            self.ed_order_no.setText(c["order_no"] or "")
            if c["order_date"]:
                d0 = d_parse(c["order_date"])
                self.de_order_date.setDate(QDate(d0.year, d0.month, d0.day))
            if unit == "days":
                self.sp_money_days.setValue(int(c["amount_days"] or 1))
            else:
                self.money_hours.set_minutes_total(int(c["amount_minutes"] or 60))
        else:
            if unit == "hours":
                if c["event_date"]:
                    d0 = d_parse(c["event_date"])
                    self.de_event.setDate(QDate(d0.year, d0.month, d0.day))
                self.amount_hours.set_minutes_total(int(c["amount_minutes"] or 60))
            else:
                dates = self.db.get_comp_dates(int(c["id"]))
                self.list_dates.clear()
                for x in dates:
                    self.list_dates.addItem(x)
                self.update_days_count()

    def update_visibility(self) -> None:
        unit = self.cb_unit.currentData()
        method = self.cb_method.currentData()

        self.grp_hours_dayoff.setVisible(unit == "hours" and method == "day_off")
        self.grp_days_dayoff.setVisible(unit == "days" and method == "day_off")
        self.grp_money.setVisible(method == "money")

        if method == "money":
            if unit == "days":
                self.sp_money_days.setVisible(True)
                self.money_hours.setVisible(False)
                self.lbl_money_amount.setText("Дни")
            else:
                self.sp_money_days.setVisible(False)
                self.money_hours.setVisible(True)
                self.lbl_money_amount.setText("Часы")

        self.update_days_count()

    def on_fill_period(self) -> None:
        if not self.db.calendar_year_complete(self.year):
            QMessageBox.warning(self, "Календарь", "Календарь года не создан/неполный. Создайте календарь.")
            return
        dlg = FillPeriodDialog(self, default_only_working=True)
        if dlg.exec() != QDialog.Accepted:
            return
        d1, d2, only_working = dlg.values()
        if d2 < d1:
            QMessageBox.warning(self, "Проверка", "Конец периода раньше начала.")
            return

        dates: list[date] = []
        cur = d1
        while cur <= d2:
            if only_working:
                if self.db.is_working_day(cur):
                    dates.append(cur)
            else:
                dates.append(cur)
            cur += timedelta(days=1)

        self.list_dates.clear()
        for d0 in dates:
            self.list_dates.addItem(d_iso(d0))
        self.update_days_count()

    def on_fill_count(self) -> None:
        if not self.db.calendar_year_complete(self.year):
            QMessageBox.warning(self, "Календарь", "Календарь года не создан/неполный. Создайте календарь.")
            return
        dlg = FillCountDialog(self, default_only_working=True)
        if dlg.exec() != QDialog.Accepted:
            return
        start, n, only_working = dlg.values()
        dates: list[date] = []
        cur = start
        while len(dates) < n:
            if only_working:
                if self.db.is_working_day(cur):
                    dates.append(cur)
            else:
                dates.append(cur)
            cur += timedelta(days=1)

        self.list_dates.clear()
        for d0 in dates:
            self.list_dates.addItem(d_iso(d0))
        self.update_days_count()

    def on_pick_calendar(self) -> None:
        if not self.db.calendar_year_complete(self.year):
            QMessageBox.warning(self, "Календарь", "Календарь года не создан/неполный. Создайте календарь.")
            return

        current_dates: list[date] = []
        for i in range(self.list_dates.count()):
            try:
                current_dates.append(d_parse(self.list_dates.item(i).text()))
            except Exception:
                pass

        dlg = MultiDateCalendarDialog(self.db, self.year, parent=self)
        dlg.set_selected_dates(current_dates)
        if dlg.exec() != QDialog.Accepted:
            return

        dates = dlg.get_dates()
        self.list_dates.clear()
        for d0 in dates:
            self.list_dates.addItem(d_iso(d0))
        self.update_days_count()

    def _validate_employee_all_years(self, employee_id: int) -> tuple[bool, str]:
        years = self.db.employee_data_years(employee_id)
        if not years:
            years = {self.year}
        for y in sorted(years):
            ok, msg = validate_non_negative_over_year(self.db, employee_id, y)
            if not ok:
                return False, msg
        return True, ""

    def accept(self) -> None:
        unit = self.cb_unit.currentData()
        method = self.cb_method.currentData()
        comment = self.ed_comment.text().strip()

        # money
        if method == "money":
            order_no = self.ed_order_no.text().strip()
            if not order_no:
                QMessageBox.warning(self, "Проверка", "Укажите № приказа.")
                return
            od = qdate_to_date(self.de_order_date.date())
            msg = validate_event_in_employee_period(self.employee, od)
            if msg:
                QMessageBox.warning(self, "Проверка", msg)
                return
            if unit == "days" and not self.db.calendar_year_complete(self.year):
                QMessageBox.warning(self, "Календарь", "Календарь года не создан/неполный. Нельзя списывать дни.")
                return

            self.db.conn.execute("BEGIN;")
            try:
                if unit == "days":
                    days = int(self.sp_money_days.value())
                    if self.edit_mode:
                        self.db.update_compensation(
                            self.comp_id,
                            event_date=d_iso(od),
                            amount_days=days,
                            amount_minutes=None,
                            order_no=order_no,
                            order_date=d_iso(od),
                            comment=comment or None,
                        )
                    else:
                        self.db.add_compensation_money(self.employee["id"], "days", None, days, order_no, od, comment)
                else:
                    minutes = self.money_hours.minutes_total()
                    if minutes <= 0:
                        raise DBError("Количество часов должно быть больше 0.")
                    if self.edit_mode:
                        self.db.update_compensation(
                            self.comp_id,
                            event_date=d_iso(od),
                            amount_minutes=minutes,
                            amount_days=None,
                            order_no=order_no,
                            order_date=d_iso(od),
                            comment=comment or None,
                        )
                    else:
                        self.db.add_compensation_money(self.employee["id"], "hours", minutes, None, order_no, od, comment)

                ok, msg2 = self._validate_employee_all_years(self.employee["id"])
                if not ok:
                    self.db.conn.execute("ROLLBACK;")
                    QMessageBox.warning(self, "Недостаточно", msg2)
                    return
                self.db.conn.execute("COMMIT;")
                super().accept()
            except Exception as e:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.critical(self, "Ошибка", str(e))
            return

        # day_off hours
        if unit == "hours":
            d0 = qdate_to_date(self.de_event.date())
            msg = validate_event_in_employee_period(self.employee, d0)
            if msg:
                QMessageBox.warning(self, "Проверка", msg)
                return
            minutes = self.amount_hours.minutes_total()
            if minutes <= 0:
                QMessageBox.warning(self, "Проверка", "Количество часов должно быть больше 0.")
                return

            self.db.conn.execute("BEGIN;")
            try:
                if self.edit_mode:
                    self.db.update_compensation(
                        self.comp_id,
                        event_date=d_iso(d0),
                        amount_minutes=minutes,
                        comment=comment or None,
                    )
                else:
                    self.db.add_compensation_hours_dayoff(self.employee["id"], d0, minutes, comment)

                ok, msg2 = self._validate_employee_all_years(self.employee["id"])
                if not ok:
                    self.db.conn.execute("ROLLBACK;")
                    QMessageBox.warning(self, "Недостаточно", msg2)
                    return
                self.db.conn.execute("COMMIT;")
                super().accept()
            except Exception as e:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.critical(self, "Ошибка", str(e))
            return

        # day_off days
        if not self.db.calendar_year_complete(self.year):
            QMessageBox.warning(self, "Календарь", "Календарь года не создан/неполный. Нельзя сохранять выходные днями.")
            return

        if self.list_dates.count() == 0:
            QMessageBox.warning(self, "Проверка", "Выберите хотя бы одну дату выходного.")
            return

        dates: list[date] = []
        for i in range(self.list_dates.count()):
            d0 = d_parse(self.list_dates.item(i).text())
            msg = validate_event_in_employee_period(self.employee, d0)
            if msg:
                QMessageBox.warning(self, "Проверка", f"Дата {d_iso(d0)}: {msg}")
                return
            dates.append(d0)

        self.db.conn.execute("BEGIN;")
        try:
            if self.edit_mode:
                self.db.update_compensation(self.comp_id, comment=comment or None)
                self.db.replace_comp_dayoff_dates(self.comp_id, self.employee["id"], dates)
            else:
                self.db.add_compensation_days_dayoff(self.employee["id"], dates, comment)

            ok, msg2 = self._validate_employee_all_years(self.employee["id"])
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")
            super().accept()
        except sqlite3.IntegrityError:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.warning(self, "Проверка", "Нельзя поставить два выходных дня на одну и ту же дату.")
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))


class CalendarDialog(QDialog):
    def __init__(self, db: DB, main: "MainWindow", parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.main = main
        self.setWindowTitle("Календарь (рабочие/нерабочие)")
        lay = QVBoxLayout(self)

        top = QHBoxLayout()
        self.cb_year = QComboBox()
        for y in range(date.today().year - 1, date.today().year + 6):
            self.cb_year.addItem(str(y), y)
        self.cb_month = QComboBox()
        for m in range(1, 13):
            self.cb_month.addItem(f"{m:02d}", m)
        self.btn_create_year = QPushButton("Создать календарь на год")
        top.addWidget(QLabel("Год:"))
        top.addWidget(self.cb_year)
        top.addWidget(QLabel("Месяц:"))
        top.addWidget(self.cb_month)
        top.addStretch(1)
        top.addWidget(self.btn_create_year)
        lay.addLayout(top)

        self.grid = QTableWidget(6, 7)
        self.grid.setHorizontalHeaderLabels(["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"])
        self.grid.verticalHeader().setVisible(False)
        self.grid.setEditTriggers(QTableWidget.NoEditTriggers)
        self.grid.setSelectionMode(QTableWidget.NoSelection)
        self.grid.cellClicked.connect(self.on_cell_clicked)
        lay.addWidget(self.grid)

        self.btn_create_year.clicked.connect(self.on_create_year)
        self.cb_year.currentIndexChanged.connect(self.refresh)
        self.cb_month.currentIndexChanged.connect(self.refresh)

        self.cb_year.setCurrentIndex(max(0, self.cb_year.findData(date.today().year)))
        self.cb_month.setCurrentIndex(date.today().month - 1)
        self.refresh()

    def on_create_year(self) -> None:
        y = int(self.cb_year.currentData())
        self.db.conn.execute("BEGIN;")
        try:
            self.db.create_calendar_year_default(y)
            ok_all = self.main.validate_all_employees_year(y)
            if not ok_all[0]:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", ok_all[1])
                return
            self.db.conn.execute("COMMIT;")
            self.refresh()
            self.main.refresh_all()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    def refresh(self) -> None:
        y = int(self.cb_year.currentData())
        m = int(self.cb_month.currentData())
        first = date(y, m, 1)
        start_col = first.weekday()
        ny, nm = next_month(y, m)
        last_day = (date(ny, nm, 1) - timedelta(days=1)).day

        for r in range(6):
            for c in range(7):
                self.grid.setItem(r, c, QTableWidgetItem(""))

        day_num = 1
        row = 0
        col = start_col
        while day_num <= last_day:
            d0 = date(y, m, day_num)
            it = QTableWidgetItem(str(day_num))
            it.setData(Qt.UserRole, d_iso(d0))
            try:
                if not self.db.is_working_day(d0):
                    it.setBackground(QBrush(QColor("#ffe0e0")))
            except CalendarMissingError:
                it.setBackground(QBrush(QColor("#e0e0e0")))
            self.grid.setItem(row, col, it)
            day_num += 1
            col += 1
            if col >= 7:
                col = 0
                row += 1

    def on_cell_clicked(self, r: int, c: int) -> None:
        it = self.grid.item(r, c)
        if not it or not it.text().strip():
            return
        d0 = d_parse(it.data(Qt.UserRole))

        self.db.conn.execute("BEGIN;")
        try:
            self.db.toggle_calendar_day(d0)
            ok_all = self.main.validate_all_employees_year(d0.year)
            if not ok_all[0]:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", ok_all[1])
                return
            self.db.conn.execute("COMMIT;")
            self.refresh()
            self.main.refresh_all()
        except CalendarMissingError as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.warning(self, "Календарь", str(e))
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))


# -----------------------------
# Duty calendar tab (per employee)
# -----------------------------
class DutyCalendarTab(QWidget):
    def __init__(self, db: DB, card: "EmployeeCardWidget"):
        super().__init__()
        self.db = db
        self.card = card
        self.employee: Optional[sqlite3.Row] = None

        lay = QVBoxLayout(self)

        defaults_box = QGroupBox("Стандарты для быстрых действий")
        defaults_lay = QFormLayout()
        defaults_box.setLayout(defaults_lay)

        self.std_start = TimeOfDayPicker()
        self.std_duration = DurationPicker(max_hours=48)
        self.std_comp_hours = DurationPicker(max_hours=24)

        self._load_defaults()

        defaults_lay.addRow("Дежурство (двойной ЛКМ) старт:", self.std_start)
        defaults_lay.addRow("Дежурство (двойной ЛКМ) длительность:", self.std_duration)
        defaults_lay.addRow("Выходной (двойной ПКМ) часы:", self.std_comp_hours)

        lay.addWidget(defaults_box)

        top = QHBoxLayout()
        self.btn_add = QPushButton("")
        top.addWidget(self.btn_add)
        top.addStretch(1)
        lay.addLayout(top)

        self.cal = NoWheelCalendarWidget()
        self.cal.setGridVisible(True)
        self.cal.setNavigationBarVisible(False)
        lay.addWidget(self.cal)

        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["Начало", "Конец", "Комментарий"])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setWordWrap(True)
        self.table.setAlternatingRowColors(True)
        lay.addWidget(self.table, 1)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.Stretch)

        self.cal.selectionChanged.connect(self.refresh_day_list)

        # ЛКМ double-click: activated (Qt обычно вызывает это на double-click/enter)
        self.cal.activated.connect(self.on_left_double_click)

        # ПКМ double-click через eventFilter на внутренней таблице
        self._cal_view: Optional[QTableView] = self.cal.findChild(QTableView)
        if self._cal_view:
            self._cal_view.viewport().installEventFilter(self)

        self.btn_add.clicked.connect(self.add_duty_for_selected_date)
        self.table.cellDoubleClicked.connect(lambda r, c: self.edit_selected())

        for w in [
            self.std_start.sp_h,
            self.std_start.sp_m,
            self.std_duration.sp_h,
            self.std_duration.sp_m,
            self.std_comp_hours.sp_h,
            self.std_comp_hours.sp_m,
        ]:
            w.valueChanged.connect(self._defaults_changed)

        self._update_add_button_text()

    def _load_defaults(self) -> None:
        ui = (self.card.main.cfg.ui or {})
        start_str = ui.get("duty_std_start", "08:00")
        try:
            hh, mm = start_str.split(":")
            self.std_start.sp_h.setValue(int(hh))
            self.std_start.sp_m.setValue(int(mm))
        except Exception:
            self.std_start.sp_h.setValue(8)
            self.std_start.sp_m.setValue(0)

        dur_min = int(ui.get("duty_std_duration_minutes", 24 * 60))
        self.std_duration.set_minutes_total(dur_min)

        comp_min = int(ui.get("comp_std_minutes", 8 * 60))
        self.std_comp_hours.set_minutes_total(comp_min)

    def _save_defaults(self) -> None:
        ui = (self.card.main.cfg.ui or {})
        ui["duty_std_start"] = f"{int(self.std_start.sp_h.value()):02d}:{int(self.std_start.sp_m.value()):02d}"
        ui["duty_std_duration_minutes"] = int(self.std_duration.minutes_total())
        ui["comp_std_minutes"] = int(self.std_comp_hours.minutes_total())
        self.card.main.cfg.ui = ui
        self.card.main.cfg.save()

    def _defaults_changed(self) -> None:
        self._save_defaults()
        self._update_add_button_text()

    def _update_add_button_text(self) -> None:
        start_t = self.std_start.get_time()
        dur_min = self.std_duration.minutes_total()
        d0 = date(2000, 1, 1)
        start_dt = datetime.combine(d0, start_t)
        end_dt = start_dt + timedelta(minutes=dur_min)
        end_str = end_dt.strftime("%H:%M")
        plus = ""
        if end_dt.date() != start_dt.date():
            plus = f"(+{(end_dt.date() - start_dt.date()).days})"
        self.btn_add.setText(f"Добавить дежурство (диалог) {start_dt.strftime('%H:%M')}–{end_str}{plus}")

    def eventFilter(self, obj, event):  # type: ignore[override]
        if self._cal_view and obj is self._cal_view.viewport():
            if event.type() == QEvent.MouseButtonDblClick and event.button() == Qt.RightButton:
                idx = self._cal_view.indexAt(event.pos())
                if idx.isValid():
                    d0 = self._date_for_cell(idx.row(), idx.column())
                    if d0.year == self.card.year and d0.month == self.card.month:
                        self.quick_add_hour_compensation(d0)
                    else:
                        self.card.main.toast("Дата вне выбранного месяца", 2500)
                return True
        return super().eventFilter(obj, event)

    def _date_for_cell(self, row: int, col: int) -> date:
        y = self.cal.yearShown()
        m = self.cal.monthShown()
        first = date(y, m, 1)
        fd = int(self.cal.firstDayOfWeek()) - 1  # 0..6
        fw = first.weekday()  # 0..6
        offset = (fw - fd) % 7
        day_offset = row * 7 + col - offset
        return first + timedelta(days=day_offset)

    def set_employee(self, emp: Optional[sqlite3.Row]) -> None:
        self.employee = emp
        self.refresh()

    def refresh(self) -> None:
        if not self.employee:
            self.table.setRowCount(0)
            return

        y = self.card.year
        m = self.card.month
        self.cal.setCurrentPage(y, m)
        self.cal.setSelectedDate(QDate(y, m, 1))

        self._paint_month_marks()
        self.refresh_day_list()

    def _paint_month_marks(self) -> None:
        if not self.employee:
            return
        y = self.card.year
        m = self.card.month
        ny, nm = next_month(y, m)
        first = date(y, m, 1)
        last = date(ny, nm, 1) - timedelta(days=1)

        fmt = QTextCharFormat()
        fmt.setBackground(QBrush(QColor("#c3f7c3")))

        cur = first
        while cur <= last:
            self.cal.setDateTextFormat(QDate(cur.year, cur.month, cur.day), QTextCharFormat())
            cur += timedelta(days=1)

        duties = self.db.list_duties_for_month(self.employee["id"], y, m)
        marked: set[date] = set()
        for r in duties:
            s = dt_parse(r["start_dt"])
            e = dt_parse(r["end_dt"])
            curd = s.date()
            endd = e.date()
            while curd <= endd:
                d0 = datetime.combine(curd, time(0, 0))
                d1 = d0 + timedelta(days=1)
                if intersect(s, e, d0, d1):
                    marked.add(curd)
                curd += timedelta(days=1)

        for d0 in marked:
            if d0.month == m and d0.year == y:
                self.cal.setDateTextFormat(QDate(d0.year, d0.month, d0.day), fmt)

    def selected_date(self) -> date:
        return qdate_to_date(self.cal.selectedDate())

    def refresh_day_list(self) -> None:
        self.table.setRowCount(0)
        if not self.employee:
            return
        d = self.selected_date()
        s_day = datetime.combine(d, time(0, 0))
        e_day = s_day + timedelta(days=1)

        duties = self.db.list_duties_for_period(self.employee["id"], s_day, e_day)
        self.table.setRowCount(len(duties))
        for i, r in enumerate(duties):
            self.table.setItem(i, 0, QTableWidgetItem(fmt_dt_iso(r["start_dt"])))
            self.table.setItem(i, 1, QTableWidgetItem(fmt_dt_iso(r["end_dt"])))
            self.table.setItem(i, 2, QTableWidgetItem(r["comment"] or ""))
            self.table.item(i, 0).setData(Qt.UserRole, int(r["id"]))

    def selected_duty_id(self) -> Optional[int]:
        sel = self.table.selectedItems()
        if not sel:
            return None
        return int(sel[0].data(Qt.UserRole))

    # --- LMB double click: add immediately (NO dialog) ---
    def on_left_double_click(self, qd: QDate) -> None:
        d0 = qdate_to_date(qd)
        if not (d0.year == self.card.year and d0.month == self.card.month):
            self.card.main.toast("Дата вне выбранного месяца", 2500)
            return
        self.cal.setSelectedDate(qd)
        self.quick_add_default_duty(d0)

    def quick_add_default_duty(self, d: date) -> None:
        if not self.employee:
            return

        start_t = self.std_start.get_time()
        dur_min = self.std_duration.minutes_total()
        start = datetime.combine(d, start_t)
        end = start + timedelta(minutes=dur_min)

        msg = validate_event_in_employee_period(self.employee, start.date())
        if msg:
            QMessageBox.warning(self, "Проверка", msg)
            return

        self.db.conn.execute("BEGIN;")
        try:
            duty_id = self.db.add_duty(self.employee["id"], start, end, "")

            ok, msg2 = self.card.validate_employee_all_years(self.employee["id"])
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return

            self.db.conn.execute("COMMIT;")

            start_str = start.strftime("%H:%M")
            end_str = end.strftime("%H:%M")
            plus = ""
            if end.date() != start.date():
                plus = f" (+{(end.date() - start.date()).days})"

            def undo() -> None:
                self.db.conn.execute("BEGIN;")
                try:
                    self.db.delete_duty(duty_id)
                    ok2, msg3 = self.card.validate_employee_all_years(self.employee["id"])
                    if not ok2:
                        self.db.conn.execute("ROLLBACK;")
                        self.card.main.toast(f"Отмена не выполнена: {msg3}", 6000)
                        return
                    self.db.conn.execute("COMMIT;")
                    self.card.main.toast("Отменено")
                    self.card.refresh()
                except Exception as e:
                    self.db.conn.execute("ROLLBACK;")
                    self.card.main.toast(f"Отмена не выполнена: {e}", 6000)

            self.card.main.toast_undo(
                f"Дежурство добавлено: {d.strftime('%d.%m.%Y')} {start_str}–{end_str}{plus}",
                undo_action=undo,
                duration_ms=10000,
            )
            self.card.refresh()

        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    # button: open dialog with defaults (as before)
    def add_duty_for_selected_date(self) -> None:
        if not self.employee:
            return

        d = self.selected_date()
        start_t = self.std_start.get_time()
        dur_min = self.std_duration.minutes_total()
        start = datetime.combine(d, start_t)
        end = start + timedelta(minutes=dur_min)

        dlg = DutyDialog(self.db, self.employee["id"], parent=self)
        dlg.set_values(start, end, "")
        if dlg.exec() != QDialog.Accepted:
            return
        s, e, comment = dlg.get_values()

        msg = validate_event_in_employee_period(self.employee, s.date())
        if msg:
            QMessageBox.warning(self, "Проверка", msg)
            return

        self.db.conn.execute("BEGIN;")
        try:
            self.db.add_duty(self.employee["id"], s, e, comment)
            ok, msg2 = self.card.validate_employee_all_years(self.employee["id"])
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")
            self.card.main.toast("Дежурство добавлено")
            self.card.refresh()
        except Exception as ex:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(ex))

    def edit_selected(self) -> None:
        if not self.employee:
            return
        did = self.selected_duty_id()
        if did is None:
            return
        row = self.db.get_duty(did)
        dlg = DutyDialog(self.db, self.employee["id"], exclude_duty_id=did, parent=self)
        dlg.set_values(dt_parse(row["start_dt"]), dt_parse(row["end_dt"]), row["comment"] or "")
        if dlg.exec() != QDialog.Accepted:
            return
        s, e, comment = dlg.get_values()

        msg = validate_event_in_employee_period(self.employee, s.date())
        if msg:
            QMessageBox.warning(self, "Проверка", msg)
            return

        self.db.conn.execute("BEGIN;")
        try:
            self.db.update_duty(did, start_dt=dt_iso(s), end_dt=dt_iso(e), comment=comment or None)
            ok, msg2 = self.card.validate_employee_all_years(self.employee["id"])
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")
            self.card.main.toast("Сохранено")
            self.card.refresh()
        except Exception as ex:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(ex))

    # --- RMB double click: quick add hour compensation (day_off) ---
    def quick_add_hour_compensation(self, d0: date) -> None:
        if not self.employee:
            return

        msg = validate_event_in_employee_period(self.employee, d0)
        if msg:
            QMessageBox.warning(self, "Проверка", msg)
            return

        minutes = self.std_comp_hours.minutes_total()
        if minutes <= 0:
            QMessageBox.warning(self, "Проверка", "Стандарт компенсации должен быть больше 0.")
            return

        self.db.conn.execute("BEGIN;")
        try:
            comp_id = self.db.add_compensation_hours_dayoff(self.employee["id"], d0, minutes, "")
            ok, msg2 = self.card.validate_employee_all_years(self.employee["id"])
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")

            def undo():
                self.db.conn.execute("BEGIN;")
                try:
                    self.db.delete_compensation(comp_id)
                    self.db.conn.execute("COMMIT;")
                    self.card.main.toast("Отменено")
                    self.card.refresh()
                except Exception as e:
                    self.db.conn.execute("ROLLBACK;")
                    self.card.main.toast(f"Не удалось отменить: {e}", 6000)

            self.card.main.toast_undo(
                f"Выходной: списано {minutes_to_hhmm(minutes)} на {d0.strftime('%d.%m.%Y')}",
                undo_action=undo,
                duration_ms=10000,
            )
            self.card.refresh()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))


# -----------------------------
# Tabs: duties & compensations (tables)
# -----------------------------
class DutyTab(QWidget):
    def __init__(self, db: DB, card: "EmployeeCardWidget"):
        super().__init__()
        self.db = db
        self.card = card
        self.employee: Optional[sqlite3.Row] = None

        lay = QVBoxLayout(self)
        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["Начало", "Конец", "Комментарий"])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setWordWrap(True)
        self.table.setAlternatingRowColors(True)
        self.table.cellDoubleClicked.connect(lambda r, c: self.edit())
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.on_context_menu)
        lay.addWidget(self.table, 1)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.Stretch)

        btns = QHBoxLayout()
        self.btn_add = QPushButton("Добавить")
        self.btn_edit = QPushButton("Редактировать")
        self.btn_del = QPushButton("Удалить")
        btns.addWidget(self.btn_add)
        btns.addWidget(self.btn_edit)
        btns.addWidget(self.btn_del)
        btns.addStretch(1)
        lay.addLayout(btns)

        self.btn_add.clicked.connect(self.add)
        self.btn_edit.clicked.connect(self.edit)
        self.btn_del.clicked.connect(self.delete)

    def get_col_widths(self) -> list[int]:
        return [self.table.columnWidth(i) for i in range(self.table.columnCount())]

    def set_col_widths(self, widths: Any) -> None:
        if not isinstance(widths, list):
            return
        for i, w in enumerate(widths):
            if i < self.table.columnCount():
                self.table.setColumnWidth(i, safe_int(w, self.table.columnWidth(i)))

    def on_context_menu(self, pos: QPoint) -> None:
        menu = QMenu(self)
        menu.addAction("Добавить", self.add)
        did = self.selected_duty_id()
        if did is not None:
            menu.addAction("Редактировать", self.edit)
            menu.addAction("Удалить", self.delete)
        menu.exec(self.table.mapToGlobal(pos))

    def set_employee(self, emp: Optional[sqlite3.Row]) -> None:
        self.employee = emp
        self.refresh()

    def refresh(self) -> None:
        self.table.setRowCount(0)
        if not self.employee:
            return
        rows = self.db.list_duties_for_month(self.employee["id"], self.card.year, self.card.month)

        duties = []
        for r in rows:
            duties.append((int(r["id"]), dt_parse(r["start_dt"]), dt_parse(r["end_dt"]), r["comment"] or ""))

        overlap_ids: set[int] = set()
        for i in range(len(duties)):
            id1, s1, e1, _ = duties[i]
            for j in range(i + 1, len(duties)):
                id2, s2, e2, _ = duties[j]
                if intersect(s1, e1, s2, e2):
                    overlap_ids.add(id1)
                    overlap_ids.add(id2)

        self.table.setRowCount(len(duties))
        for i, (did, s, e, comment) in enumerate(duties):
            it0 = QTableWidgetItem(s.strftime("%d.%m.%Y %H:%M"))
            it1 = QTableWidgetItem(e.strftime("%d.%m.%Y %H:%M"))
            it2 = QTableWidgetItem(comment)
            it0.setData(Qt.UserRole, did)
            self.table.setItem(i, 0, it0)
            self.table.setItem(i, 1, it1)
            self.table.setItem(i, 2, it2)

            if did in overlap_ids:
                for c in range(3):
                    self.table.item(i, c).setBackground(QBrush(QColor("#fff2cc")))

    def selected_duty_id(self) -> Optional[int]:
        sel = self.table.selectedItems()
        if not sel:
            return None
        return int(sel[0].data(Qt.UserRole))

    def add(self) -> None:
        if not self.employee:
            return
        dlg = DutyDialog(self.db, self.employee["id"], parent=self)
        if dlg.exec() != QDialog.Accepted:
            return
        s, e, comment = dlg.get_values()
        msg = validate_event_in_employee_period(self.employee, s.date())
        if msg:
            QMessageBox.warning(self, "Проверка", msg)
            return

        self.db.conn.execute("BEGIN;")
        try:
            self.db.add_duty(self.employee["id"], s, e, comment)
            ok, msg2 = self.card.validate_employee_all_years(self.employee["id"])
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")
            self.card.main.toast("Дежурство добавлено")
            self.card.refresh()
        except Exception as ex:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(ex))

    def edit(self) -> None:
        if not self.employee:
            return
        did = self.selected_duty_id()
        if did is None:
            return
        row = self.db.get_duty(did)
        dlg = DutyDialog(self.db, self.employee["id"], exclude_duty_id=did, parent=self)
        dlg.set_values(dt_parse(row["start_dt"]), dt_parse(row["end_dt"]), row["comment"] or "")
        if dlg.exec() != QDialog.Accepted:
            return
        s, e, comment = dlg.get_values()
        msg = validate_event_in_employee_period(self.employee, s.date())
        if msg:
            QMessageBox.warning(self, "Проверка", msg)
            return

        self.db.conn.execute("BEGIN;")
        try:
            self.db.update_duty(did, start_dt=dt_iso(s), end_dt=dt_iso(e), comment=comment or None)
            ok, msg2 = self.card.validate_employee_all_years(self.employee["id"])
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")
            self.card.main.toast("Сохранено")
            self.card.refresh()
        except Exception as ex:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(ex))

    def delete(self) -> None:
        if not self.employee:
            return
        did = self.selected_duty_id()
        if did is None:
            return
        if QMessageBox.question(self, "Удаление", "Удалить дежурство?") != QMessageBox.Yes:
            return

        snap = self.db.snapshot_duty(did)

        self.db.conn.execute("BEGIN;")
        try:
            self.db.delete_duty(did)
            ok, msg2 = self.card.validate_employee_all_years(self.employee["id"])
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")

            def undo() -> None:
                self.db.conn.execute("BEGIN;")
                try:
                    self.db.restore_duty(snap)
                    ok2, msg3 = self.card.validate_employee_all_years(self.employee["id"])
                    if not ok2:
                        self.db.conn.execute("ROLLBACK;")
                        self.card.main.toast(f"Не удалось восстановить: {msg3}", 6000)
                        return
                    self.db.conn.execute("COMMIT;")
                    self.card.main.toast("Восстановлено")
                    self.card.refresh()
                except Exception as e:
                    self.db.conn.execute("ROLLBACK;")
                    self.card.main.toast(f"Не удалось восстановить: {e}", 6000)

            self.card.main.toast_undo("Дежурство удалено", undo_action=undo, duration_ms=10000)
            self.card.refresh()
        except Exception as ex:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(ex))


class CompTab(QWidget):
    def __init__(self, db: DB, card: "EmployeeCardWidget"):
        super().__init__()
        self.db = db
        self.card = card
        self.employee: Optional[sqlite3.Row] = None

        lay = QVBoxLayout(self)
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["Тип", "Способ", "Дата/Приказ", "Количество", "Комментарий"])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setWordWrap(True)
        self.table.setAlternatingRowColors(True)
        self.table.cellDoubleClicked.connect(lambda r, c: self.edit())
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.on_context_menu)
        lay.addWidget(self.table, 1)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.Stretch)

        btns = QHBoxLayout()
        self.btn_add = QPushButton("Добавить")
        self.btn_edit = QPushButton("Редактировать")
        self.btn_del = QPushButton("Удалить")
        btns.addWidget(self.btn_add)
        btns.addWidget(self.btn_edit)
        btns.addWidget(self.btn_del)
        btns.addStretch(1)
        lay.addLayout(btns)

        self.btn_add.clicked.connect(self.add)
        self.btn_edit.clicked.connect(self.edit)
        self.btn_del.clicked.connect(self.delete)

    def get_col_widths(self) -> list[int]:
        return [self.table.columnWidth(i) for i in range(self.table.columnCount())]

    def set_col_widths(self, widths: Any) -> None:
        if not isinstance(widths, list):
            return
        for i, w in enumerate(widths):
            if i < self.table.columnCount():
                self.table.setColumnWidth(i, safe_int(w, self.table.columnWidth(i)))

    def on_context_menu(self, pos: QPoint) -> None:
        menu = QMenu(self)
        menu.addAction("Добавить", self.add)
        cid = self.selected_comp_id()
        if cid is not None:
            menu.addAction("Редактировать", self.edit)
            menu.addAction("Удалить", self.delete)
        menu.exec(self.table.mapToGlobal(pos))

    def set_employee(self, emp: Optional[sqlite3.Row]) -> None:
        self.employee = emp
        self.refresh()

    def refresh(self) -> None:
        self.table.setRowCount(0)
        if not self.employee:
            return
        rows = self.db.list_compensations_for_month(self.employee["id"], self.card.year, self.card.month)
        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            unit = r["unit"]
            method = r["method"]
            typ = "Часы" if unit == "hours" else "Дни"
            mth = "Выходной" if method == "day_off" else "Деньги"
            when = ""
            amount = ""
            if method == "money":
                when = f"№{r['order_no']} от {fmt_date_iso(r['order_date'])}"
                amount = minutes_to_hhmm(int(r["amount_minutes"] or 0)) if unit == "hours" else str(int(r["amount_days"] or 0))
            else:
                if unit == "hours":
                    when = fmt_date_iso(r["event_date"] or "")
                    amount = minutes_to_hhmm(int(r["amount_minutes"] or 0))
                else:
                    dates = self.db.get_comp_dates(int(r["id"]))
                    when = ", ".join([fmt_date_iso(x) for x in dates[:3]]) + ("…" if len(dates) > 3 else "")
                    amount = str(len(dates))

            it0 = QTableWidgetItem(typ)
            it0.setData(Qt.UserRole, int(r["id"]))
            self.table.setItem(i, 0, it0)
            self.table.setItem(i, 1, QTableWidgetItem(mth))
            self.table.setItem(i, 2, QTableWidgetItem(when))
            self.table.setItem(i, 3, QTableWidgetItem(amount))
            self.table.setItem(i, 4, QTableWidgetItem(r["comment"] or ""))

    def selected_comp_id(self) -> Optional[int]:
        sel = self.table.selectedItems()
        if not sel:
            return None
        return int(sel[0].data(Qt.UserRole))

    def add(self) -> None:
        if not self.employee:
            return
        dlg = CompensationDialog(self.db, self.employee, self.card.year, comp_id=None, parent=self)
        if dlg.exec() == QDialog.Accepted:
            self.card.main.toast("Сохранено")
            self.card.refresh()

    def edit(self) -> None:
        if not self.employee:
            return
        cid = self.selected_comp_id()
        if cid is None:
            return
        dlg = CompensationDialog(self.db, self.employee, self.card.year, comp_id=cid, parent=self)
        if dlg.exec() == QDialog.Accepted:
            self.card.main.toast("Сохранено")
            self.card.refresh()

    def delete(self) -> None:
        if not self.employee:
            return
        cid = self.selected_comp_id()
        if cid is None:
            return
        if QMessageBox.question(self, "Удаление", "Удалить компенсацию?") != QMessageBox.Yes:
            return

        snap = self.db.snapshot_compensation(cid)

        self.db.conn.execute("BEGIN;")
        try:
            self.db.delete_compensation(cid)
            ok, msg2 = self.card.validate_employee_all_years(self.employee["id"])
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")

            def undo() -> None:
                self.db.conn.execute("BEGIN;")
                try:
                    self.db.restore_compensation(snap)
                    ok2, msg3 = self.card.validate_employee_all_years(self.employee["id"])
                    if not ok2:
                        self.db.conn.execute("ROLLBACK;")
                        self.card.main.toast(f"Не удалось восстановить: {msg3}", 6000)
                        return
                    self.db.conn.execute("COMMIT;")
                    self.card.main.toast("Восстановлено")
                    self.card.refresh()
                except sqlite3.IntegrityError:
                    self.db.conn.execute("ROLLBACK;")
                    self.card.main.toast("Не удалось восстановить (конфликт дат выходных).", 6000)
                except Exception as e:
                    self.db.conn.execute("ROLLBACK;")
                    self.card.main.toast(f"Не удалось восстановить: {e}", 6000)

            self.card.main.toast_undo("Компенсация удалена", undo_action=undo, duration_ms=10000)
            self.card.refresh()
        except Exception as ex:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(ex))


# -----------------------------
# Employee card
# -----------------------------
class EmployeeCardWidget(QWidget):
    def __init__(self, db: DB, main: "MainWindow"):
        super().__init__()
        self.db = db
        self.main = main
        self.employee_id: Optional[int] = None
        self.year = date.today().year
        self.month = date.today().month

        lay = QVBoxLayout(self)
        self.lbl_title = QLabel("Сотрудник не выбран")
        self.lbl_title.setStyleSheet("font-size:16px; font-weight:600;")
        lay.addWidget(self.lbl_title)

        self.lbl_summary = QLabel("")
        lay.addWidget(self.lbl_summary)

        self.tabs = QTabWidget()
        lay.addWidget(self.tabs, 1)

        # Calendar is primary
        self.tab_cal = DutyCalendarTab(self.db, self)
        self.tab_duties = DutyTab(self.db, self)
        self.tab_comps = CompTab(self.db, self)

        self.tabs.addTab(self.tab_cal, "Календарь")
        self.tabs.addTab(self.tab_duties, "Дежурства")
        self.tabs.addTab(self.tab_comps, "Компенсации")
        self.tabs.setCurrentIndex(0)

    def set_period(self, year: int, month: int) -> None:
        self.year = year
        self.month = month
        self.refresh()

    def set_employee_id(self, employee_id: Optional[int]) -> None:
        self.employee_id = employee_id
        self.refresh()

    def validate_employee_all_years(self, employee_id: int) -> tuple[bool, str]:
        return self.main.validate_employee_all_years(employee_id)

    def refresh(self) -> None:
        if self.employee_id is None:
            self.lbl_title.setText("Сотрудник не выбран")
            self.lbl_summary.setText("")
            self.tab_cal.set_employee(None)
            self.tab_duties.set_employee(None)
            self.tab_comps.set_employee(None)
            return

        emp = self.db.get_employee(self.employee_id)
        fio = f"{emp['last_name']} {emp['first_name']} {emp['middle_name'] or ''}".strip()
        rank = (emp["rank"] or "").strip()
        pos = (emp["position"] or "").strip()
        sub = " — ".join([x for x in [rank, pos] if x])
        self.lbl_title.setText(f"{fio}  ({sub})")

        try:
            summ = compute_month_summary(self.db, self.employee_id, self.year, self.month)
            parts = [f"<b>Период:</b> {self.year:04d}-{self.month:02d}"]
            if summ["opening_minutes"] or summ["opening_days"]:
                parts.append(f"<b>Перенос:</b> часы {minutes_to_hhmm(summ['opening_minutes'])}, дни {summ['opening_days']}")
            parts.append(
                f"<b>Начислено за месяц:</b> часы {minutes_to_hhmm(summ['month_minutes'])}, дни {summ['month_days'] if summ['days_ok'] else '—'}"
            )
            parts.append(
                f"<b>Списано за месяц:</b> часы {minutes_to_hhmm(summ['comp_month_minutes'])}, дни {summ['comp_month_days'] if summ['days_ok'] else '—'}"
            )
            parts.append(
                f"<b>Остаток на конец месяца:</b> часы {minutes_to_hhmm(summ['bal_minutes'])}, дни {summ['bal_days'] if summ['days_ok'] else '—'}"
            )
            if not summ["days_ok"]:
                parts.append(f"<span style='color:#b00020'>{summ['days_err']}</span>")
            self.lbl_summary.setText("<br>".join(parts))
        except Exception as e:
            self.lbl_summary.setText(f"<span style='color:#b00020'>Ошибка расчета: {e}</span>")

        self.tab_cal.set_employee(emp)
        self.tab_duties.set_employee(emp)
        self.tab_comps.set_employee(emp)


# -----------------------------
# Main window
# -----------------------------
class MainWindow(QMainWindow):
    def __init__(self, db: DB, cfg: AppConfig):
        super().__init__()
        self.db = db
        self.cfg = cfg
        self._declined_calendar_years: set[int] = set()

        self.setWindowTitle(f"{db.get_department_name()} — учет времени")
        self.resize(1200, 700)

        self.snackbar = Snackbar(self)

        menubar = QMenuBar(self)
        self.setMenuBar(menubar)

        m_file = menubar.addMenu("Файл")
        m_file.addAction("Экспорт (Excel, простой)…").triggered.connect(self.export_simple)

        m_view = menubar.addMenu("Вид")
        m_theme = m_view.addMenu("Тема")
        self.act_theme_system = QAction("Системная", self, checkable=True)
        self.act_theme_light = QAction("Светлая", self, checkable=True)
        self.act_theme_dark = QAction("Тёмная", self, checkable=True)
        self.theme_group = [self.act_theme_system, self.act_theme_light, self.act_theme_dark]
        for a in self.theme_group:
            m_theme.addAction(a)
            a.triggered.connect(self.on_theme_changed)

        m_service = menubar.addMenu("Сервис")
        m_service.addAction("Календарь").triggered.connect(self.open_calendar)
        m_service.addAction("Настройки отдела").triggered.connect(self.open_settings)

        splitter = QSplitter()
        self.splitter = splitter
        self.setCentralWidget(splitter)

        # left
        left = QWidget()
        left_lay = QVBoxLayout(left)

        top = QHBoxLayout()
        self.cb_year = QComboBox()
        for y in range(date.today().year - 1, date.today().year + 6):
            self.cb_year.addItem(str(y), y)
        self.cb_month = QComboBox()
        for m in range(1, 13):
            self.cb_month.addItem(f"{m:02d}", m)
        self.cb_active = QComboBox()
        self.cb_active.addItem("Активные", True)
        self.cb_active.addItem("Все", False)

        top.addWidget(QLabel("Год:"))
        top.addWidget(self.cb_year)
        top.addWidget(QLabel("Месяц:"))
        top.addWidget(self.cb_month)
        top.addStretch(1)
        top.addWidget(self.cb_active)
        left_lay.addLayout(top)

        self.ed_search = QLineEdit()
        self.ed_search.setPlaceholderText("Поиск (ФИО/звание/должность)…")
        left_lay.addWidget(self.ed_search)

        self.list_emp = QListWidget()
        self.list_emp.setContextMenuPolicy(Qt.CustomContextMenu)
        self.list_emp.customContextMenuRequested.connect(self.on_emp_context_menu)
        left_lay.addWidget(self.list_emp)

        btn_row = QHBoxLayout()
        self.btn_add = QPushButton("Добавить")
        self.btn_edit = QPushButton("Редактировать")
        self.btn_del = QPushButton("Удалить")
        self.btn_transfer = QPushButton("Переведен…")
        self.btn_dismiss = QPushButton("Уволен…")
        btn_row.addWidget(self.btn_add)
        btn_row.addWidget(self.btn_edit)
        btn_row.addWidget(self.btn_del)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_transfer)
        btn_row.addWidget(self.btn_dismiss)
        left_lay.addLayout(btn_row)

        splitter.addWidget(left)

        # right
        self.card = EmployeeCardWidget(self.db, self)
        splitter.addWidget(self.card)
        splitter.setStretchFactor(1, 3)

        # signals
        self.cb_year.currentIndexChanged.connect(self.on_period_changed)
        self.cb_month.currentIndexChanged.connect(self.on_period_changed)
        self.cb_active.currentIndexChanged.connect(self.refresh_employees)
        self.ed_search.textChanged.connect(self.refresh_employees)
        self.list_emp.currentItemChanged.connect(self.on_emp_selected)

        self.btn_add.clicked.connect(self.add_employee)
        self.btn_edit.clicked.connect(self.edit_employee)
        self.btn_del.clicked.connect(self.delete_employee)
        self.btn_transfer.clicked.connect(lambda: self.set_end_reason("transfer"))
        self.btn_dismiss.clicked.connect(lambda: self.set_end_reason("dismissal"))

        self.restore_ui_state()
        self.on_period_changed()

    # ---- snackbar helpers ----
    def toast(self, text: str, duration_ms: int = 4000) -> None:
        self.snackbar.show_message(text, duration_ms=duration_ms)
        self._reposition_snackbar()

    def toast_undo(self, text: str, undo_action: Callable[[], None], duration_ms: int = 10000) -> None:
        self.snackbar.show_message(text, duration_ms=duration_ms, action_text="Отменить", action=undo_action)
        self._reposition_snackbar()

    def resizeEvent(self, event) -> None:  # type: ignore[override]
        super().resizeEvent(event)
        self._reposition_snackbar()

    def _reposition_snackbar(self) -> None:
        if not self.snackbar.isVisible():
            return
        w = min(720, self.width() - 40)
        self.snackbar.resize(w, self.snackbar.height())
        self.snackbar.move((self.width() - w) // 2, self.height() - self.snackbar.height() - 20)

    # ---- UI state ----
    def restore_ui_state(self) -> None:
        ui = self.cfg.ui or {}

        theme = (ui.get("theme") or "system").lower()
        self.set_theme_checked(theme)
        apply_theme(QApplication.instance(), theme)

        g = ui.get("main_geometry")
        if isinstance(g, str) and g:
            try:
                ba = QByteArray.fromBase64(g.encode("ascii"))
                self.restoreGeometry(ba)
            except Exception:
                pass

        sizes = ui.get("splitter_sizes")
        if isinstance(sizes, list) and len(sizes) >= 2:
            try:
                self.splitter.setSizes([int(sizes[0]), int(sizes[1])])
            except Exception:
                pass

        y = safe_int(ui.get("year"), date.today().year)
        m = safe_int(ui.get("month"), date.today().month)
        yi = self.cb_year.findData(y)
        if yi >= 0:
            self.cb_year.setCurrentIndex(yi)
        self.cb_month.setCurrentIndex(max(0, min(11, m - 1)))

        active_only = ui.get("active_only", True)
        self.cb_active.setCurrentIndex(0 if active_only else 1)
        self.ed_search.setText(ui.get("search", "") or "")

    def save_ui_state(self) -> None:
        ui = self.cfg.ui or {}
        ui["theme"] = self.current_theme()

        ui["main_geometry"] = bytes(self.saveGeometry().toBase64()).decode("ascii")
        ui["splitter_sizes"] = self.splitter.sizes()

        y, m = self.current_year_month()
        ui["year"] = y
        ui["month"] = m
        ui["active_only"] = bool(self.cb_active.currentData())
        ui["search"] = self.ed_search.text()

        eid = self.selected_employee_id()
        ui["selected_employee_id"] = eid

        ui["duty_col_widths"] = self.card.tab_duties.get_col_widths()
        ui["comp_col_widths"] = self.card.tab_comps.get_col_widths()

        self.cfg.ui = ui
        self.cfg.save()

    def closeEvent(self, event) -> None:  # type: ignore[override]
        self.save_ui_state()
        super().closeEvent(event)

    # ---- theme menu ----
    def set_theme_checked(self, theme: str) -> None:
        theme = (theme or "system").lower()
        for a in self.theme_group:
            a.setChecked(False)
        if theme == "dark":
            self.act_theme_dark.setChecked(True)
        elif theme == "light":
            self.act_theme_light.setChecked(True)
        else:
            self.act_theme_system.setChecked(True)

    def current_theme(self) -> str:
        if self.act_theme_dark.isChecked():
            return "dark"
        if self.act_theme_light.isChecked():
            return "light"
        return "system"

    def on_theme_changed(self) -> None:
        sender = self.sender()
        if sender in self.theme_group:
            for a in self.theme_group:
                if a is not sender:
                    a.setChecked(False)
            if not sender.isChecked():
                sender.setChecked(True)

        theme = self.current_theme()
        apply_theme(QApplication.instance(), theme)
        self.cfg.ui = self.cfg.ui or {}
        self.cfg.ui["theme"] = theme
        self.cfg.save()

    # ---- period / selection ----
    def current_year_month(self) -> tuple[int, int]:
        return int(self.cb_year.currentData()), int(self.cb_month.currentData())

    def selected_employee_id(self) -> Optional[int]:
        it = self.list_emp.currentItem()
        return int(it.data(Qt.UserRole)) if it else None

    def on_period_changed(self) -> None:
        y, _ = self.current_year_month()
        if (not self.db.calendar_year_complete(y)) and (y not in self._declined_calendar_years):
            r = QMessageBox.question(
                self,
                "Календарь",
                f"Календарь на {y} год не создан (или неполный).\n"
                f"Создать календарь по умолчанию (Пн–Пт рабочие, Сб–Вс выходные)?\n\n"
                f"Если выбрать 'Нет', операции, зависящие от календаря (дни/выходные), будут недоступны.",
                QMessageBox.Yes | QMessageBox.No,
            )
            if r == QMessageBox.No:
                self._declined_calendar_years.add(y)
            else:
                self.db.conn.execute("BEGIN;")
                try:
                    self.db.create_calendar_year_default(y)
                    ok_all = self.validate_all_employees_year(y)
                    if not ok_all[0]:
                        self.db.conn.execute("ROLLBACK;")
                        QMessageBox.warning(self, "Недостаточно", ok_all[1])
                    else:
                        self.db.conn.execute("COMMIT;")
                        self.toast("Календарь создан. Проверьте праздники/переносы в 'Сервис → Календарь'.", 6000)
                except Exception as e:
                    self.db.conn.execute("ROLLBACK;")
                    QMessageBox.critical(self, "Ошибка", str(e))

        self.refresh_employees()

    def refresh_all(self) -> None:
        self.setWindowTitle(f"{self.db.get_department_name()} — учет времени")
        self.refresh_employees()
        self.card.refresh()

    def refresh_employees(self) -> None:
        y, m = self.current_year_month()
        active_only = bool(self.cb_active.currentData())
        search = self.ed_search.text()
        emps = self.db.list_employees_for_month(y, m, active_only=active_only, search=search)
        current_id = self.selected_employee_id()

        self.list_emp.blockSignals(True)
        self.list_emp.clear()
        for e in emps:
            fio = f"{e['last_name']} {e['first_name']} {e['middle_name'] or ''}".strip()
            rank = (e["rank"] or "").strip()
            pos = (e["position"] or "").strip()
            sub = " — ".join([x for x in [rank, pos] if x])

            active_in_month = is_emp_active_in_month(e, y, m)
            status = ""
            if (not active_in_month) and e["end_date"]:
                if e["end_reason"] == "transfer":
                    status = f"(переведен {fmt_date_iso(e['end_date'])})"
                elif e["end_reason"] == "dismissal":
                    status = f"(уволен {fmt_date_iso(e['end_date'])})"
                else:
                    status = f"(неактивен с {fmt_date_iso(e['end_date'])})"

            sub2 = (sub + " " + status).strip() if status else sub
            it = QListWidgetItem(f"{fio}\n{sub2}")
            it.setData(Qt.UserRole, int(e["id"]))
            if (not active_only) and (not active_in_month):
                it.setForeground(QBrush(QColor("#808080")))
            self.list_emp.addItem(it)
        self.list_emp.blockSignals(False)

        preferred = current_id
        if preferred is None:
            preferred = (self.cfg.ui or {}).get("selected_employee_id")

        if preferred is not None:
            for i in range(self.list_emp.count()):
                if self.list_emp.item(i).data(Qt.UserRole) == preferred:
                    self.list_emp.setCurrentRow(i)
                    break

        self.card.set_period(y, m)

        if self.list_emp.currentItem() is None and self.list_emp.count() > 0:
            self.list_emp.setCurrentRow(0)
        else:
            self.on_emp_selected(self.list_emp.currentItem(), None)

    def on_emp_selected(self, cur: QListWidgetItem, prev: QListWidgetItem) -> None:
        self.card.set_employee_id(self.selected_employee_id())
        ui = self.cfg.ui or {}
        if "duty_col_widths" in ui:
            self.card.tab_duties.set_col_widths(ui["duty_col_widths"])
        if "comp_col_widths" in ui:
            self.card.tab_comps.set_col_widths(ui["comp_col_widths"])

    # ---- context menu employees ----
    def on_emp_context_menu(self, pos: QPoint) -> None:
        menu = QMenu(self)
        menu.addAction("Добавить", self.add_employee)
        eid = self.selected_employee_id()
        if eid is not None:
            menu.addAction("Редактировать", self.edit_employee)
            menu.addSeparator()
            menu.addAction("Переведен…", lambda: self.set_end_reason("transfer"))
            menu.addAction("Уволен…", lambda: self.set_end_reason("dismissal"))
            menu.addSeparator()
            menu.addAction("Удалить", self.delete_employee)
        menu.exec(self.list_emp.mapToGlobal(pos))

    # ---- service actions ----
    def open_calendar(self) -> None:
        CalendarDialog(self.db, self, self).exec()

    def open_settings(self) -> None:
        if DepartmentSettingsDialog(self.db, self).exec() == QDialog.Accepted:
            self.refresh_all()

    # ---- validation ----
    def validate_employee_all_years(self, employee_id: int) -> tuple[bool, str]:
        years = self.db.employee_data_years(employee_id)
        if not years:
            y, _ = self.current_year_month()
            years = {y}
        for yr in sorted(years):
            ok, msg = validate_non_negative_over_year(self.db, employee_id, yr)
            if not ok:
                return False, msg
        return True, ""

    def validate_all_employees_year(self, year: int) -> tuple[bool, str]:
        rows = self.db.conn.execute("SELECT id FROM employee").fetchall()
        for r in rows:
            ok, msg = validate_non_negative_over_year(self.db, int(r["id"]), year)
            if not ok:
                return False, msg
        return True, ""

    # ---- employee actions ----
    def add_employee(self) -> None:
        y, m = self.current_year_month()
        dlg = EmployeeDialog("Добавить сотрудника", default_year=y, default_month=m, parent=self)
        if dlg.exec() != QDialog.Accepted:
            return
        v = dlg.get_values()
        self.db.conn.execute("BEGIN;")
        try:
            self.db.add_employee(
                v["last_name"],
                v["first_name"],
                v["middle_name"],
                v["rank"],
                v["position"],
                v["start_month"],
                v["opening_minutes"],
                v["opening_days"],
            )
            self.db.conn.execute("COMMIT;")
            self.toast("Сотрудник добавлен")
            self.refresh_employees()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    def edit_employee(self) -> None:
        eid = self.selected_employee_id()
        if eid is None:
            return
        emp = self.db.get_employee(eid)
        y, m = self.current_year_month()
        dlg = EmployeeDialog("Редактировать сотрудника", default_year=y, default_month=m, parent=self)
        dlg.set_from_employee(emp)
        if dlg.exec() != QDialog.Accepted:
            return
        v = dlg.get_values()

        self.db.conn.execute("BEGIN;")
        try:
            self.db.update_employee(
                eid,
                last_name=v["last_name"],
                first_name=v["first_name"],
                middle_name=v["middle_name"] or None,
                rank=v["rank"] or None,
                position=v["position"] or None,
                start_month=v["start_month"],
                opening_minutes=v["opening_minutes"],
                opening_days=v["opening_days"],
            )
            ok, msg = self.validate_employee_all_years(eid)
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg)
                return
            self.db.conn.execute("COMMIT;")
            self.toast("Сохранено")
            self.refresh_all()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    def delete_employee(self) -> None:
        eid = self.selected_employee_id()
        if eid is None:
            return
        emp = self.db.get_employee(eid)
        fio = f"{emp['last_name']} {emp['first_name']} {emp['middle_name'] or ''}".strip()
        if (
            QMessageBox.question(self, "Удаление", f"Удалить сотрудника '{fio}'? Будут удалены дежурства и компенсации.")
            != QMessageBox.Yes
        ):
            return
        self.db.conn.execute("BEGIN;")
        try:
            self.db.delete_employee(eid)
            self.db.conn.execute("COMMIT;")
            self.toast("Сотрудник удален")
            self.refresh_all()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    def set_end_reason(self, reason: str) -> None:
        eid = self.selected_employee_id()
        if eid is None:
            return
        title = "Переведен" if reason == "transfer" else "Уволен"
        dlg = EndDateDialog(title, self)
        if dlg.exec() != QDialog.Accepted:
            return
        d0 = dlg.get_date()
        emp = self.db.get_employee(eid)
        start, _ = employee_period_bounds(emp)
        if d0 < start:
            QMessageBox.warning(self, "Проверка", "Дата раньше начала работы в отделе.")
            return
        if self.db.has_records_after_date(eid, d0):
            QMessageBox.warning(
                self,
                "Проверка",
                "Нельзя установить дату перевода/увольнения: у сотрудника есть дежурства или компенсации после этой даты.\n"
                "Сначала исправьте записи после указанной даты.",
            )
            return

        self.db.conn.execute("BEGIN;")
        try:
            self.db.update_employee(eid, end_date=d_iso(d0), end_reason=reason)
            ok, msg = self.validate_employee_all_years(eid)
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg)
                return
            self.db.conn.execute("COMMIT;")
            self.toast("Сохранено")
            self.refresh_all()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    # ---- export ----
    def export_simple(self) -> None:
        eid = self.selected_employee_id()
        if eid is None:
            QMessageBox.information(self, "Экспорт", "Выберите сотрудника.")
            return
        y, m = self.current_year_month()
        out, _ = QFileDialog.getSaveFileName(self, "Экспорт (Excel)", f"export_{y:04d}-{m:02d}.xlsx", "Excel (*.xlsx)")
        if not out:
            return
        if not out.endswith(".xlsx"):
            out += ".xlsx"
        try:
            export_simple_xlsx(self.db, eid, y, m, out)
            self.toast("Экспорт готов")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))


# -----------------------------
# App entry
# -----------------------------
def main() -> None:
    app = QApplication(sys.argv)

    cfg = AppConfig.load()
    theme = (cfg.ui or {}).get("theme", "system")
    apply_theme(app, theme)

    dlg = BaseSelectorDialog(cfg)
    if dlg.exec() != QDialog.Accepted or not dlg.selected_path:
        return

    db = DB(dlg.selected_path)
    win = MainWindow(db, cfg)
    win.show()
    rc = app.exec()
    win.save_ui_state()
    db.close()
    sys.exit(rc)


if __name__ == "__main__":
    main()