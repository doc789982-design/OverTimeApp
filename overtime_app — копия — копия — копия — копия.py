# -*- coding: utf-8 -*-
"""
OvertimeTab (один файл)
Python + PySide6 + SQLite
Экспорт в Excel: openpyxl

pip install PySide6 openpyxl
python overtime_app_all_NEW2.py
"""

from __future__ import annotations

import json
import os
import shutil
import sqlite3
import sys
import uuid
import subprocess
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Callable, Optional

from PySide6.QtCore import Qt, QDate, QByteArray, QTimer, QPoint, QEvent, Signal, QRect
from PySide6.QtPrintSupport import QPrintDialog, QPrinter
from PySide6.QtPdf import QPdfDocument
from PySide6.QtPdfWidgets import QPdfView
from PySide6.QtGui import QColor, QBrush, QTextCharFormat, QPalette, QAction, QPainter, QPen, QPixmap, QKeySequence, QIcon
from PySide6.QtWidgets import (
    QApplication,
    QAbstractSpinBox,
    QCalendarWidget,
    QCheckBox,
    QComboBox,
    QDateEdit,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMenu,
    QMenuBar,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QSpinBox,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QTableView,
    QWidgetAction,
    QGridLayout,
)

from PySide6.QtWidgets import QStackedWidget, QGraphicsOpacityEffect
from PySide6.QtCore import QPropertyAnimation, QEasingCurve

APP_NAME = "OvertimeTab"
SCHEMA_VERSION = 3  # + duty_break


# -----------------------------
# Utils
# -----------------------------
def dt_parse(s: str) -> datetime:
    return datetime.fromisoformat(s)


def dt_iso(dt: datetime) -> str:
    return dt.isoformat(timespec="minutes")


def d_parse(s: str) -> date:
    return date.fromisoformat(s)


def d_iso(d: date) -> str:
    return d.isoformat()


def fmt_date_iso(iso: str) -> str:
    try:
        return d_parse(iso).strftime("%d.%m.%Y")
    except Exception:
        return iso or ""


def fmt_dt_iso(iso: str) -> str:
    try:
        return dt_parse(iso).strftime("%d.%m.%Y %H:%M")
    except Exception:
        return iso or ""


def parse_month(s: str) -> tuple[int, int]:
    y, m = s.split("-")
    return int(y), int(m)


def next_month(year: int, month: int) -> tuple[int, int]:
    return (year + 1, 1) if month == 12 else (year, month + 1)


def month_bounds_dt(year: int, month: int) -> tuple[datetime, datetime]:
    start = datetime(year, month, 1, 0, 0)
    ny, nm = next_month(year, month)
    end = datetime(ny, nm, 1, 0, 0)
    return start, end


def year_bounds_dt(year: int) -> tuple[datetime, datetime]:
    return datetime(year, 1, 1, 0, 0), datetime(year + 1, 1, 1, 0, 0)


def minutes_to_hhmm(m: int) -> str:
    sign = "-" if m < 0 else ""
    m = abs(m)
    h = m // 60
    mm = m % 60
    return f"{sign}{h}:{mm:02d}"

def fio_initials_last(last_name: str | None, first_name: str | None, middle_name: str | None) -> str:
    last = (last_name or "").strip()
    first = (first_name or "").strip()
    middle = (middle_name or "").strip()

    initials: list[str] = []
    if first:
        initials.append(first[0].upper() + ".")
    if middle:
        initials.append(middle[0].upper() + ".")

    init = "".join(initials).strip()

    if init and last:
        return f"{init} {last}".strip()
    if last:
        return last
    return init

def parse_hhmm(s: str) -> time:
    try:
        s = (s or "").strip()
        hh, mm = s.split(":")
        return time(int(hh), int(mm))
    except Exception:
        return time(0, 0)


def fmt_hhmm(t: time) -> str:
    return f"{int(t.hour):02d}:{int(t.minute):02d}"

def ru_plural(n: int, one: str, few: str, many: str) -> str:
    n = abs(int(n))
    if 11 <= (n % 100) <= 14:
        return many
    last = n % 10
    if last == 1:
        return one
    if 2 <= last <= 4:
        return few
    return many


def fmt_minutes_ru_words(minutes: int) -> str:
    sign = "-" if minutes < 0 else ""
    m = abs(int(minutes))
    h = m // 60
    mm = m % 60

    parts: list[str] = []

    # показываем часы всегда, даже если 0 (чтобы было "0 часов")
    parts.append(f"{h} {ru_plural(h, 'час', 'часа', 'часов')}")

    # минуты добавляем только если они есть
    if mm > 0:
        parts.append(f"{mm} {ru_plural(mm, 'минута', 'минуты', 'минут')}")

    return sign + " ".join(parts)


def fmt_days_ru_words(days: int) -> str:
    sign = "-" if days < 0 else ""
    d = abs(int(days))
    return sign + f"{d} {ru_plural(d, 'день', 'дня', 'дней')}"

def qdate_to_date(qd: QDate) -> date:
    return date(qd.year(), qd.month(), qd.day())


def intersect(a0: datetime, a1: datetime, b0: datetime, b1: datetime) -> Optional[tuple[datetime, datetime]]:
    s = max(a0, b0)
    e = min(a1, b1)
    return (s, e) if s < e else None


def merge_intervals(intervals: list[tuple[datetime, datetime]]) -> list[tuple[datetime, datetime]]:
    if not intervals:
        return []
    intervals.sort(key=lambda x: x[0])
    merged = [intervals[0]]
    for s, e in intervals[1:]:
        ps, pe = merged[-1]
        if s <= pe:
            merged[-1] = (ps, max(pe, e))
        else:
            merged.append((s, e))
    return merged

def subtract_intervals(
    base: tuple[datetime, datetime],
    cuts: list[tuple[datetime, datetime]],
) -> list[tuple[datetime, datetime]]:
    """base minus cuts -> список непересекающихся интервалов."""
    bs, be = base
    if bs >= be:
        return []

    if not cuts:
        return [(bs, be)]

    clipped: list[tuple[datetime, datetime]] = []
    for cs, ce in cuts:
        inter = intersect(bs, be, cs, ce)
        if inter:
            clipped.append(inter)

    if not clipped:
        return [(bs, be)]

    clipped = merge_intervals(clipped)

    out: list[tuple[datetime, datetime]] = []
    cur = bs
    for cs, ce in clipped:
        if cs > cur:
            out.append((cur, min(cs, be)))
        cur = max(cur, ce)
        if cur >= be:
            break

    if cur < be:
        out.append((cur, be))

    return [(s, e) for s, e in out if s < e]

def yyyymm_from_end_date(end_date_iso: str) -> str:
    return (end_date_iso or "")[:7]


def is_emp_active_in_month(emp: sqlite3.Row, year: int, month: int) -> bool:
    m = f"{year:04d}-{month:02d}"
    if emp["start_month"] > m:
        return False
    if emp["end_date"] is None:
        return True
    return yyyymm_from_end_date(emp["end_date"]) >= m


def safe_int(x: Any, default: int) -> int:
    try:
        return int(x)
    except Exception:
        return default


# -----------------------------
# Theme
# -----------------------------
# -----------------------------
# Theme & macOS Elegance
# -----------------------------
def get_macos_stylesheet() -> str:
    """Глобальный стиль в духе Apple HIG (Dark Mode)"""
    return """
    /* ========================================= */
    /* 1. ОБЩИЙ ШРИФТ И ФОН                      */
    /* ========================================= */
    QWidget {
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, Helvetica, Arial, sans-serif;
        font-size: 13px;
        color: #EBEBF5;
    }
    QMainWindow, QDialog {
        background-color: #1E1E1E;
    }

    /* ========================================= */
    /* 2. КАРТОЧКИ (ВМЕСТО СТАРЫХ РАМОК)         */
    /* ========================================= */
    QGroupBox {
        background-color: #2C2C2E;
        border-radius: 10px;
        border: none;
        margin-top: 20px; /* Отступ под заголовок */
        padding-top: 15px;
    }
    QGroupBox::title {
        subcontrol-origin: margin;
        subcontrol-position: top left;
        padding: 0px 5px;
        color: #8E8E93;
        font-weight: 600;
        font-size: 12px;
        text-transform: uppercase;
    }

    /* ========================================= */
    /* 3. ПОЛЯ ВВОДА (СВЕЧЕНИЕ ПРИ КЛИКЕ)        */
    /* ========================================= */
    QLineEdit, QSpinBox, QDateEdit, QComboBox {
        background-color: #3A3A3C;
        border: 2px solid transparent; /* Подготовка для свечения */
        border-radius: 6px;
        padding: 5px 8px;
        color: #FFFFFF;
        selection-background-color: #0A84FF;
    }
    /* Тот самый Apple-эффект фокуса */
    QLineEdit:focus, QSpinBox:focus, QDateEdit:focus, QComboBox:focus {
        border: 2px solid #0A84FF;
        background-color: #48484A;
    }
    QComboBox::drop-down {
        border: none;
        width: 20px;
    }

    /* ========================================= */
    /* 4. КНОПКИ                                 */
    /* ========================================= */
    QPushButton {
        background-color: #3A3A3C;
        border: none;
        border-radius: 6px;
        padding: 7px 14px;
        color: #FFFFFF;
        font-weight: 500;
    }
    QPushButton:hover {
        background-color: #48484A;
    }
    QPushButton:pressed {
        background-color: #2C2C2E;
    }
    QPushButton#AccentButton {
        background-color: #0A84FF;
        color: white;
        font-weight: 600;
    }
    QPushButton#AccentButton:hover {
        background-color: #0070E0;
    }

    /* ========================================= */
    /* 5. МЕНЮ ПО ПРАВОМУ КЛИКУ (DROP-SHADOW)    */
    /* ========================================= */
    QMenu {
        background-color: #2C2C2E;
        border-radius: 8px;
        padding: 5px;
        border: 1px solid rgba(255, 255, 255, 0.1);
    }
    QMenu::item {
        padding: 6px 25px 6px 15px;
        border-radius: 5px;
        color: #FFFFFF;
        font-size: 13px;
    }
    QMenu::item:selected {
        background-color: #0A84FF;
        color: white;
    }
    QMenu::separator {
        height: 1px;
        background: rgba(255, 255, 255, 0.1);
        margin: 5px 10px;
    }

    /* ========================================= */
    /* 6. СПИСОК СОТРУДНИКОВ (С АВАТАРКАМИ)      */
    /* ========================================= */
    QListWidget {
        icon-size: 32px;
    }

    /* ========================================= */
    /* 7. ТАБЛИЦЫ (СОВРЕМЕННЫЙ DATA GRID)        */
    /* ========================================= */
    QTableWidget, QTableView {
        background-color: transparent;
        border: none;
        gridline-color: transparent; /* Убираем жесткую сетку */
        outline: none; 
    }
    QTableWidget::item {
        padding: 5px;
        border-bottom: 1px solid rgba(255, 255, 255, 0.05); /* Только тонкая линия снизу */
    }
    QTableWidget::item:selected {
        background-color: #3A3A3C;
        border-radius: 6px;
        color: #FFFFFF;
    }
    /* Заголовки колонок */
    QHeaderView::section {
        background-color: transparent;
        color: #8E8E93;
        font-weight: 600;
        border: none;
        border-bottom: 1px solid rgba(255, 255, 255, 0.1);
        padding: 6px;
        text-align: left;
    }

    /* ========================================= */
    /* 8. СКРОЛЛБАРЫ (КАК В macOS)               */
    /* ========================================= */
    QScrollBar:vertical {
        border: none;
        background: transparent;
        width: 12px;
        margin: 0px;
    }
    QScrollBar::handle:vertical {
        background: rgba(255, 255, 255, 0.2);
        min-height: 30px;
        border-radius: 6px;
        margin: 2px;
    }
    QScrollBar::handle:vertical:hover {
        background: rgba(255, 255, 255, 0.35);
    }
    QScrollBar::sub-line:vertical, QScrollBar::add-line:vertical {
        height: 0px; /* Прячем стрелочки вверх/вниз */
    }
    QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
        background: none;
    }
    
    QScrollBar:horizontal {
        border: none;
        background: transparent;
        height: 12px;
        margin: 0px;
    }
    QScrollBar::handle:horizontal {
        background: rgba(255, 255, 255, 0.2);
        min-width: 30px;
        border-radius: 6px;
        margin: 2px;
    }
    QScrollBar::handle:horizontal:hover {
        background: rgba(255, 255, 255, 0.35);
    }
    QScrollBar::sub-line:horizontal, QScrollBar::add-line:horizontal {
        width: 0px;
    }
    QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
        background: none;
    }
    """

def apply_theme(app: QApplication, theme: str) -> None:
    # Применяем QSS вместо старых палитр
    app.setStyleSheet(get_macos_stylesheet())


# -----------------------------
# UI: visual spinners / pickers
# -----------------------------
def configure_spinbox(widget: QAbstractSpinBox) -> None:
    widget.setButtonSymbols(QAbstractSpinBox.UpDownArrows)
    widget.setAccelerated(True)
    widget.setKeyboardTracking(False)
    le = widget.lineEdit()
    if le is not None:
        le.setReadOnly(True)
        le.setCursor(Qt.ArrowCursor)


class TwoDigitSpinBox(QSpinBox):
    def textFromValue(self, v: int) -> str:  # type: ignore[override]
        return f"{v:02d}"

    def valueFromText(self, text: str) -> int:  # type: ignore[override]
        try:
            return int(text)
        except Exception:
            return 0


class DurationPicker(QWidget):
    """Длительность: часы (0..999) + минуты (00..59)"""

    def __init__(self, max_hours: int = 999, parent: QWidget | None = None):
        super().__init__(parent)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)

        self.sp_h = QSpinBox()
        self.sp_h.setRange(0, max_hours)
        configure_spinbox(self.sp_h)

        self.sp_m = TwoDigitSpinBox()
        self.sp_m.setRange(0, 59)
        configure_spinbox(self.sp_m)

        lay.addWidget(self.sp_h)
        lay.addWidget(QLabel(":"))
        lay.addWidget(self.sp_m)
        lay.addStretch(1)

    def set_minutes_total(self, minutes: int) -> None:
        minutes = max(0, int(minutes))
        self.sp_h.setValue(minutes // 60)
        self.sp_m.setValue(minutes % 60)

    def minutes_total(self) -> int:
        return int(self.sp_h.value()) * 60 + int(self.sp_m.value())


class TimeOfDayPicker(QWidget):
    """Время суток: 00..23 : 00..59"""

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)

        self.sp_h = TwoDigitSpinBox()
        self.sp_h.setRange(0, 23)
        configure_spinbox(self.sp_h)

        self.sp_m = TwoDigitSpinBox()
        self.sp_m.setRange(0, 59)
        configure_spinbox(self.sp_m)

        lay.addWidget(self.sp_h)
        lay.addWidget(QLabel(":"))
        lay.addWidget(self.sp_m)
        lay.addStretch(1)

    def set_time(self, t: time) -> None:
        self.sp_h.setValue(t.hour)
        self.sp_m.setValue(t.minute)

    def get_time(self) -> time:
        return time(int(self.sp_h.value()), int(self.sp_m.value()))


class DateTimePicker(QWidget):
    """Дата (через popup) + визуальные крутилки времени"""

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)

        self.de = QDateEdit()
        self.de.setCalendarPopup(True)
        self.de.setDate(QDate.currentDate())
        configure_spinbox(self.de)

        self.tp = TimeOfDayPicker()

        lay.addWidget(self.de)
        lay.addWidget(self.tp)
        lay.addStretch(1)

    def set_datetime(self, dt: datetime) -> None:
        self.de.setDate(QDate(dt.year, dt.month, dt.day))
        self.tp.set_time(dt.time().replace(second=0, microsecond=0))

    def get_datetime(self) -> datetime:
        d0 = qdate_to_date(self.de.date())
        t0 = self.tp.get_time()
        return datetime(d0.year, d0.month, d0.day, t0.hour, t0.minute)

def create_avatar_icon(name: str, size: int = 32) -> QIcon:
    pixmap = QPixmap(size, size)
    pixmap.fill(Qt.transparent)
    
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.Antialiasing)
    
    # Рисуем темно-серый кружок
    painter.setBrush(QBrush(QColor("#3A3A3C")))
    painter.setPen(Qt.NoPen)
    painter.drawEllipse(0, 0, size, size)
    
    # Берем 1-2 буквы из ФИО
    parts = [p for p in name.split() if p]
    initials = "".join([p[0].upper() for p in parts[:2]])
    
    # Рисуем белые буквы по центру
    painter.setPen(QColor("#FFFFFF"))
    font = painter.font()
    font.setBold(True)
    font.setPointSize(size // 2 - 2)
    painter.setFont(font)
    painter.drawText(pixmap.rect(), Qt.AlignCenter, initials)
    painter.end()
    
    return QIcon(pixmap)
    
# -----------------------------
# Snackbar / Undo
# -----------------------------
class Snackbar(QFrame):
    def __init__(self, parent: QWidget):
        super().__init__(parent)
        self.setObjectName("Snackbar")
        self.setFrameShape(QFrame.StyledPanel)
        self.setStyleSheet(
            """
            QFrame#Snackbar {
                background: #2C2C2E;
                border-radius: 10px;
                border: 1px solid rgba(255,255,255,0.1);
            }
            QFrame#Snackbar QLabel { color: #FFFFFF; font-size: 14px; font-weight: 500; }
        """
        )
        self.hide()

        self._action: Optional[Callable[[], None]] = None
        self._timer = QTimer(self)
        self._timer.setSingleShot(True)
        self._timer.timeout.connect(self.hide_message)

        # Подготовка для плавной анимации
        self.opacity_effect = QGraphicsOpacityEffect(self)
        self.setGraphicsEffect(self.opacity_effect)
        self.anim = QPropertyAnimation(self.opacity_effect, b"opacity")
        self.anim.setDuration(250)
        self.anim.setEasingCurve(QEasingCurve.InOutQuad)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(20, 12, 20, 12)
        self.lbl = QLabel("")
        lay.addWidget(self.lbl, 1)

    def show_message(
        self,
        text: str,
        duration_ms: int = 4000,
        action_text: str | None = None,
        action: Callable[[], None] | None = None,
    ) -> None:
        self.lbl.setText(text)
        self.adjustSize()
        self.show()
        self.raise_()
        
        # Плавное появление
        self.anim.stop()
        self.anim.setStartValue(0.0)
        self.anim.setEndValue(1.0)
        try: self.anim.finished.disconnect()
        except: pass
        self.anim.start()
        
        self._timer.start(duration_ms)

    def hide_message(self):
        # Плавное исчезновение
        self.anim.stop()
        self.anim.setStartValue(1.0)
        self.anim.setEndValue(0.0)
        self.anim.finished.connect(self.hide)
        self.anim.start()


# -----------------------------
# App config
# -----------------------------
def app_data_dir() -> Path:
    base = Path.home() / f".{APP_NAME.lower()}"
    base.mkdir(parents=True, exist_ok=True)
    return base


def app_db_store_dir() -> Path:
    d = app_data_dir() / "databases"
    d.mkdir(parents=True, exist_ok=True)
    return d


def config_path() -> Path:
    return app_data_dir() / "config.json"


@dataclass
class AppConfig:
    db_paths: list[str]
    last_db_path: Optional[str] = None
    ui: dict[str, Any] = None  # type: ignore

    @staticmethod
    def load() -> "AppConfig":
        p = config_path()
        if not p.exists():
            return AppConfig(db_paths=[], last_db_path=None, ui={})
        data = json.loads(p.read_text(encoding="utf-8"))
        return AppConfig(
            db_paths=data.get("db_paths", []),
            last_db_path=data.get("last_db_path"),
            ui=data.get("ui", {}) or {},
        )

    def save(self) -> None:
        config_path().write_text(
            json.dumps(
                {"db_paths": self.db_paths, "last_db_path": self.last_db_path, "ui": self.ui},
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )


# -----------------------------
# DB
# -----------------------------
class DBError(Exception):
    pass


class CalendarMissingError(DBError):
    pass


class DB:
    def __init__(self, path: str):
        self.path = path
        self.conn = sqlite3.connect(path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys = ON;")
        self._init_or_migrate()

    def close(self) -> None:
        self.conn.close()

    def _has_table(self, name: str) -> bool:
        r = self.conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (name,)).fetchone()
        return r is not None

    def _init_or_migrate(self) -> None:
        if not self._has_table("meta"):
            self._init_schema()
            return
        ver = self.get_meta_int("schema_version", 0)
        if ver == SCHEMA_VERSION:
            return
        if ver == 1 and SCHEMA_VERSION == 2:
            self._migrate_1_to_2()
            return
        if ver == 2 and SCHEMA_VERSION == 3:
            self._migrate_2_to_3()
            return
        raise DBError(f"Схема базы версии {ver}, ожидается {SCHEMA_VERSION}. Миграции не реализованы.")

    def _migrate_1_to_2(self) -> None:
        self.conn.execute("BEGIN;")
        try:
            self.conn.execute("ALTER TABLE employee ADD COLUMN opening_minutes INTEGER NOT NULL DEFAULT 0;")
            self.conn.execute("ALTER TABLE employee ADD COLUMN opening_days INTEGER NOT NULL DEFAULT 0;")
            self.set_meta("schema_version", "2")
            self.conn.execute("COMMIT;")
        except Exception:
            self.conn.execute("ROLLBACK;")
            raise

    def _migrate_2_to_3(self) -> None:
        self.conn.execute("BEGIN;")
        try:
            self.conn.execute(
                """
                CREATE TABLE IF NOT EXISTS duty_break (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    duty_id INTEGER NOT NULL REFERENCES duty(id) ON DELETE CASCADE,
                    start_dt TEXT NOT NULL,
                    end_dt TEXT NOT NULL
                );
                """
            )
            self.set_meta("schema_version", "3")
            self.conn.execute("COMMIT;")
        except Exception:
            self.conn.execute("ROLLBACK;")
            raise

    def _init_schema(self) -> None:
        c = self.conn
        c.execute("BEGIN;")
        try:
            c.execute("CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT NOT NULL);")
            c.execute(
                """
                CREATE TABLE department_settings (
                    id INTEGER PRIMARY KEY CHECK (id=1),
                    department_name TEXT NOT NULL,
                    resp_position TEXT,
                    resp_rank TEXT,
                    resp_last_name TEXT,
                    resp_first_name TEXT,
                    resp_middle_name TEXT
                );
                """
            )
            c.execute(
                """
                CREATE TABLE employee (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    last_name TEXT NOT NULL,
                    first_name TEXT NOT NULL,
                    middle_name TEXT,
                    rank TEXT,
                    position TEXT,
                    start_month TEXT NOT NULL,
                    end_date TEXT,
                    end_reason TEXT,
                    opening_minutes INTEGER NOT NULL DEFAULT 0,
                    opening_days INTEGER NOT NULL DEFAULT 0
                );
                """
            )
            c.execute(
                """
                CREATE TABLE duty (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER NOT NULL REFERENCES employee(id) ON DELETE CASCADE,
                    start_dt TEXT NOT NULL,
                    end_dt TEXT NOT NULL,
                    comment TEXT
                );
                """
            )
            c.execute(
                """
                CREATE TABLE duty_break (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    duty_id INTEGER NOT NULL REFERENCES duty(id) ON DELETE CASCADE,
                    start_dt TEXT NOT NULL,
                    end_dt TEXT NOT NULL
                );
                """
            )            
            c.execute(
                """
                CREATE TABLE compensation (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER NOT NULL REFERENCES employee(id) ON DELETE CASCADE,
                    unit TEXT NOT NULL,
                    method TEXT NOT NULL,
                    event_date TEXT,
                    amount_minutes INTEGER,
                    amount_days INTEGER,
                    order_no TEXT,
                    order_date TEXT,
                    comment TEXT
                );
                """
            )
            c.execute(
                """
                CREATE TABLE comp_day_off_date (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    compensation_id INTEGER NOT NULL REFERENCES compensation(id) ON DELETE CASCADE,
                    employee_id INTEGER NOT NULL REFERENCES employee(id) ON DELETE CASCADE,
                    day_off_date TEXT NOT NULL,
                    UNIQUE(employee_id, day_off_date)
                );
                """
            )
            c.execute(
                """
                CREATE TABLE calendar_day (
                    date TEXT PRIMARY KEY,
                    is_working INTEGER NOT NULL
                );
                """
            )

            self.set_meta("schema_version", str(SCHEMA_VERSION))
            self.set_meta("db_uuid", str(uuid.uuid4()))
            self.set_meta("created_at", datetime.now().isoformat(timespec="seconds"))
            c.execute("INSERT INTO department_settings (id, department_name) VALUES (1, ?)", ("Подразделение",))
            c.execute("COMMIT;")
        except Exception:
            c.execute("ROLLBACK;")
            raise

    def get_meta(self, key: str, default: Optional[str] = None) -> Optional[str]:
        r = self.conn.execute("SELECT value FROM meta WHERE key=?", (key,)).fetchone()
        return r["value"] if r else default

    def get_meta_int(self, key: str, default: int = 0) -> int:
        v = self.get_meta(key)
        return int(v) if v is not None else default

    def set_meta(self, key: str, value: str) -> None:
        self.conn.execute(
            "INSERT INTO meta(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, value),
        )

    def get_department_name(self) -> str:
        r = self.conn.execute("SELECT department_name FROM department_settings WHERE id=1").fetchone()
        return r["department_name"] if r else "Подразделение"

    def get_department_settings(self) -> sqlite3.Row:
        return self.conn.execute("SELECT * FROM department_settings WHERE id=1").fetchone()

    def update_department_settings(self, **fields: Any) -> None:
        if not fields:
            return
        cols = ", ".join([f"{k}=?" for k in fields.keys()])
        vals = list(fields.values())
        self.conn.execute(f"UPDATE department_settings SET {cols} WHERE id=1", vals)

    # calendar
    def calendar_year_complete(self, year: int) -> bool:
        y0 = f"{year:04d}-01-01"
        y1 = f"{year:04d}-12-31"
        r = self.conn.execute("SELECT COUNT(*) AS c FROM calendar_day WHERE date BETWEEN ? AND ?", (y0, y1)).fetchone()
        days = 366 if date(year, 12, 31).timetuple().tm_yday == 366 else 365
        return int(r["c"]) == days

    def ensure_calendar_year_default(self, year: int) -> None:
        """Создает недостающие дни календаря на год (Пн–Пт рабочие, Сб–Вс выходные),
        НЕ затирая уже существующие ручные правки.
        """
        y_start = date(year, 1, 1)
        y_end = date(year, 12, 31)
        rows = []
        cur = y_start
        while cur <= y_end:
            is_working = 1 if cur.weekday() < 5 else 0
            rows.append((d_iso(cur), is_working))
            cur += timedelta(days=1)
    
        self.conn.executemany(
            "INSERT OR IGNORE INTO calendar_day(date,is_working) VALUES (?,?)",
            rows,
        )

    def create_calendar_year_default(self, year: int) -> None:
        y_start = date(year, 1, 1)
        y_end = date(year, 12, 31)
        self.conn.execute("DELETE FROM calendar_day WHERE date BETWEEN ? AND ?", (d_iso(y_start), d_iso(y_end)))
        rows = []
        cur = y_start
        while cur <= y_end:
            is_working = 1 if cur.weekday() < 5 else 0
            rows.append((d_iso(cur), is_working))
            cur += timedelta(days=1)
        self.conn.executemany("INSERT INTO calendar_day(date,is_working) VALUES (?,?)", rows)

    def is_working_day(self, d0: date) -> bool:
        r = self.conn.execute("SELECT is_working FROM calendar_day WHERE date=?", (d_iso(d0),)).fetchone()
        if not r:
            raise CalendarMissingError(f"Календарь не создан/неполный для даты {d_iso(d0)}")
        return bool(r["is_working"])

    def toggle_calendar_day(self, d0: date) -> None:
        r = self.conn.execute("SELECT is_working FROM calendar_day WHERE date=?", (d_iso(d0),)).fetchone()
        if not r:
            raise CalendarMissingError("Календарь года не создан.")
        new_val = 0 if r["is_working"] else 1
        self.conn.execute("UPDATE calendar_day SET is_working=? WHERE date=?", (new_val, d_iso(d0)))

    # employees
    def list_employees_for_month(self, year: int, month: int, active_only: bool, search: str) -> list[sqlite3.Row]:
        m = f"{year:04d}-{month:02d}"
        search = (search or "").strip().lower()
        params: list[Any] = []
        where = []
        if active_only:
            where.append("start_month <= ?")
            where.append("(end_date IS NULL OR substr(end_date,1,7) >= ?)")
            params += [m, m]
        if search:
            where.append(
                "(lower(last_name) LIKE ? OR lower(first_name) LIKE ? OR lower(middle_name) LIKE ? OR lower(rank) LIKE ? OR lower(position) LIKE ?)"
            )
            s = f"%{search}%"
            params += [s, s, s, s, s]
        sql = "SELECT * FROM employee"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY last_name, first_name, middle_name"
        return self.conn.execute(sql, params).fetchall()

    def get_employee(self, employee_id: int) -> sqlite3.Row:
        r = self.conn.execute("SELECT * FROM employee WHERE id=?", (employee_id,)).fetchone()
        if not r:
            raise DBError("Сотрудник не найден.")
        return r

    def add_employee(
        self,
        last: str,
        first: str,
        middle: str,
        rank: str,
        position: str,
        start_month: str,
        opening_minutes: int,
        opening_days: int,
    ) -> int:
        cur = self.conn.execute(
            """
            INSERT INTO employee(last_name,first_name,middle_name,rank,position,start_month,opening_minutes,opening_days)
            VALUES (?,?,?,?,?,?,?,?)
            """,
            (last, first, middle or None, rank or None, position or None, start_month, opening_minutes, opening_days),
        )
        return int(cur.lastrowid)

    def update_employee(self, employee_id: int, **fields: Any) -> None:
        if not fields:
            return
        cols = ", ".join([f"{k}=?" for k in fields.keys()])
        vals = list(fields.values()) + [employee_id]
        self.conn.execute(f"UPDATE employee SET {cols} WHERE id=?", vals)

    def delete_employee(self, employee_id: int) -> None:
        self.conn.execute("DELETE FROM employee WHERE id=?", (employee_id,))

    def employee_data_years(self, employee_id: int) -> set[int]:
        years: set[int] = set()
        rows = self.conn.execute(
            """
            SELECT DISTINCT substr(start_dt,1,4) AS y FROM duty WHERE employee_id=?
            UNION
            SELECT DISTINCT substr(event_date,1,4) AS y FROM compensation WHERE employee_id=? AND event_date IS NOT NULL
            UNION
            SELECT DISTINCT substr(day_off_date,1,4) AS y FROM comp_day_off_date WHERE employee_id=?
            """,
            (employee_id, employee_id, employee_id),
        ).fetchall()
        for r in rows:
            if r["y"]:
                years.add(int(r["y"]))
        return years

    def has_records_after_date(self, employee_id: int, end_date: date) -> bool:
        r1 = self.conn.execute(
            "SELECT 1 FROM duty WHERE employee_id=? AND substr(start_dt,1,10) > ? LIMIT 1",
            (employee_id, d_iso(end_date)),
        ).fetchone()
        if r1:
            return True
        r2 = self.conn.execute(
            "SELECT 1 FROM compensation WHERE employee_id=? AND event_date IS NOT NULL AND event_date > ? LIMIT 1",
            (employee_id, d_iso(end_date)),
        ).fetchone()
        if r2:
            return True
        r3 = self.conn.execute(
            "SELECT 1 FROM comp_day_off_date WHERE employee_id=? AND day_off_date > ? LIMIT 1",
            (employee_id, d_iso(end_date)),
        ).fetchone()
        return bool(r3)

    # duties
    def breaks_for_duty_ids(self, duty_ids: list[int]) -> dict[int, list[tuple[datetime, datetime]]]:
        out: dict[int, list[tuple[datetime, datetime]]] = {}
        duty_ids = [int(x) for x in duty_ids if x is not None]
        if not duty_ids:
            return out

        ph = ",".join(["?"] * len(duty_ids))
        rows = self.conn.execute(
            f"""
            SELECT duty_id, start_dt, end_dt
            FROM duty_break
            WHERE duty_id IN ({ph})
            ORDER BY duty_id, start_dt
            """,
            duty_ids,
        ).fetchall()

        for r in rows:
            did = int(r["duty_id"])
            out.setdefault(did, []).append((dt_parse(r["start_dt"]), dt_parse(r["end_dt"])))

        return out

    def replace_duty_breaks(self, duty_id: int, breaks: list[tuple[datetime, datetime]]) -> None:
        self.conn.execute("DELETE FROM duty_break WHERE duty_id=?", (int(duty_id),))
        rows: list[tuple[int, str, str]] = []
        for s, e in breaks or []:
            rows.append((int(duty_id), dt_iso(s), dt_iso(e)))
        if rows:
            self.conn.executemany("INSERT INTO duty_break(duty_id,start_dt,end_dt) VALUES (?,?,?)", rows)
            
    def list_duties_for_month(self, employee_id: int, year: int, month: int) -> list[sqlite3.Row]:
        start, end = month_bounds_dt(year, month)
        return self.conn.execute(
            """
            SELECT * FROM duty
            WHERE employee_id=?
              AND end_dt > ?
              AND start_dt < ?
            ORDER BY start_dt
            """,
            (employee_id, dt_iso(start), dt_iso(end)),
        ).fetchall()

    def list_duties_for_period(self, employee_id: int, start: datetime, end: datetime) -> list[sqlite3.Row]:
        return self.conn.execute(
            """
            SELECT * FROM duty
            WHERE employee_id=?
              AND end_dt > ?
              AND start_dt < ?
            ORDER BY start_dt
            """,
            (employee_id, dt_iso(start), dt_iso(end)),
        ).fetchall()

    def get_duty(self, duty_id: int) -> sqlite3.Row:
        r = self.conn.execute("SELECT * FROM duty WHERE id=?", (duty_id,)).fetchone()
        if not r:
            raise DBError("Дежурство не найдено.")
        return r

    def find_overlapping_duties(
        self,
        employee_id: int,
        start: datetime,
        end: datetime,
        exclude_duty_id: Optional[int] = None,
    ) -> list[tuple[int, datetime, datetime]]:
        params: list[Any] = [employee_id, dt_iso(start), dt_iso(end)]
        sql = """
            SELECT id, start_dt, end_dt FROM duty
            WHERE employee_id=?
              AND end_dt > ?
              AND start_dt < ?
        """
        if exclude_duty_id is not None:
            sql += " AND id <> ?"
            params.append(exclude_duty_id)
        rows = self.conn.execute(sql, params).fetchall()
        return [(int(r["id"]), dt_parse(r["start_dt"]), dt_parse(r["end_dt"])) for r in rows]

    def add_duty(self, employee_id: int, start: datetime, end: datetime, comment: str) -> int:
        overlaps = self.find_overlapping_duties(employee_id, start, end, exclude_duty_id=None)
        if overlaps:
            raise DBError("Нельзя сохранить дежурство: есть пересечение с существующим дежурством.")

        cur = self.conn.execute(
            "INSERT INTO duty(employee_id,start_dt,end_dt,comment) VALUES (?,?,?,?)",
            (employee_id, dt_iso(start), dt_iso(end), comment or None),
        )
        return int(cur.lastrowid)

    def update_duty(self, duty_id: int, **fields: Any) -> None:
        if not fields:
            return

        if ("start_dt" in fields) or ("end_dt" in fields):
            cur_row = self.get_duty(duty_id)
            employee_id = int(cur_row["employee_id"])

            new_start = dt_parse(fields.get("start_dt", cur_row["start_dt"]))
            new_end = dt_parse(fields.get("end_dt", cur_row["end_dt"]))

            overlaps = self.find_overlapping_duties(employee_id, new_start, new_end, exclude_duty_id=duty_id)
            if overlaps:
                raise DBError("Нельзя сохранить дежурство: есть пересечение с существующим дежурством.")

        cols = ", ".join([f"{k}=?" for k in fields.keys()])
        vals = list(fields.values()) + [duty_id]
        self.conn.execute(f"UPDATE duty SET {cols} WHERE id=?", vals)

    def delete_duty(self, duty_id: int) -> None:
        self.conn.execute("DELETE FROM duty WHERE id=?", (duty_id,))

    def snapshot_duty(self, duty_id: int) -> dict[str, Any]:
        r = self.get_duty(duty_id)

        br = self.conn.execute(
            "SELECT start_dt, end_dt FROM duty_break WHERE duty_id=? ORDER BY start_dt",
            (int(duty_id),),
        ).fetchall()

        return {
            "employee_id": int(r["employee_id"]),
            "start_dt": r["start_dt"],
            "end_dt": r["end_dt"],
            "comment": r["comment"] or "",
            "breaks": [{"start_dt": x["start_dt"], "end_dt": x["end_dt"]} for x in br],
        }

    def restore_duty(self, snap: dict[str, Any]) -> int:
        duty_id = self.add_duty(
            snap["employee_id"],
            dt_parse(snap["start_dt"]),
            dt_parse(snap["end_dt"]),
            snap.get("comment", ""),
        )

        breaks = []
        for b in snap.get("breaks", []) or []:
            try:
                breaks.append((dt_parse(b["start_dt"]), dt_parse(b["end_dt"])))
            except Exception:
                pass

        self.replace_duty_breaks(duty_id, breaks)
        return duty_id
        
            # compensations
    def list_compensations_for_month(self, employee_id: int, year: int, month: int) -> list[sqlite3.Row]:
        start = date(year, month, 1)
        ny, nm = next_month(year, month)
        end = date(ny, nm, 1)
        return self.conn.execute(
            """
            SELECT * FROM compensation
            WHERE employee_id=?
              AND (
                (event_date IS NOT NULL AND event_date >= ? AND event_date < ?)
                OR
                (unit='days' AND method='day_off' AND id IN (
                    SELECT DISTINCT compensation_id FROM comp_day_off_date
                    WHERE employee_id=? AND day_off_date >= ? AND day_off_date < ?
                ))
              )
            ORDER BY COALESCE(event_date,'9999-12-31'), id
            """,
            (employee_id, d_iso(start), d_iso(end), employee_id, d_iso(start), d_iso(end)),
        ).fetchall()

    def get_compensation(self, comp_id: int) -> sqlite3.Row:
        r = self.conn.execute("SELECT * FROM compensation WHERE id=?", (comp_id,)).fetchone()
        if not r:
            raise DBError("Компенсация не найдена.")
        return r

    def get_comp_dates(self, compensation_id: int) -> list[str]:
        rows = self.conn.execute(
            "SELECT day_off_date FROM comp_day_off_date WHERE compensation_id=? ORDER BY day_off_date",
            (compensation_id,),
        ).fetchall()
        return [r["day_off_date"] for r in rows]

    def add_compensation_hours_dayoff(self, employee_id: int, event_date: date, minutes: int, comment: str) -> int:
        cur = self.conn.execute(
            """
            INSERT INTO compensation(employee_id,unit,method,event_date,amount_minutes,comment)
            VALUES (?,?,?,?,?,?)
            """,
            (employee_id, "hours", "day_off", d_iso(event_date), minutes, comment or None),
        )
        return int(cur.lastrowid)

    def add_compensation_days_dayoff(self, employee_id: int, dates: list[date], comment: str) -> int:
        cur = self.conn.execute(
            "INSERT INTO compensation(employee_id,unit,method,amount_days,comment) VALUES (?,?,?,?,?)",
            (employee_id, "days", "day_off", len(dates), comment or None),
        )
        comp_id = int(cur.lastrowid)
        for d0 in dates:
            self.conn.execute(
                "INSERT INTO comp_day_off_date(compensation_id,employee_id,day_off_date) VALUES (?,?,?)",
                (comp_id, employee_id, d_iso(d0)),
            )
        return comp_id

    def add_compensation_money(
        self,
        employee_id: int,
        unit: str,
        amount_minutes: int | None,
        amount_days: int | None,
        order_no: str,
        order_date: date,
        comment: str,
    ) -> int:
        cur = self.conn.execute(
            """
            INSERT INTO compensation(employee_id,unit,method,event_date,amount_minutes,amount_days,order_no,order_date,comment)
            VALUES (?,?,?,?,?,?,?,?,?)
            """,
            (
                employee_id,
                unit,
                "money",
                d_iso(order_date),
                amount_minutes,
                amount_days,
                order_no,
                d_iso(order_date),
                comment or None,
            ),
        )
        return int(cur.lastrowid)

    def update_compensation(self, comp_id: int, **fields: Any) -> None:
        if not fields:
            return
        cols = ", ".join([f"{k}=?" for k in fields.keys()])
        vals = list(fields.values()) + [comp_id]
        self.conn.execute(f"UPDATE compensation SET {cols} WHERE id=?", vals)

    def replace_comp_dayoff_dates(self, comp_id: int, employee_id: int, dates: list[date]) -> None:
        self.conn.execute("DELETE FROM comp_day_off_date WHERE compensation_id=?", (comp_id,))
        for d0 in dates:
            self.conn.execute(
                "INSERT INTO comp_day_off_date(compensation_id,employee_id,day_off_date) VALUES (?,?,?)",
                (comp_id, employee_id, d_iso(d0)),
            )
        self.update_compensation(comp_id, amount_days=len(dates))

    def delete_compensation(self, comp_id: int) -> None:
        self.conn.execute("DELETE FROM compensation WHERE id=?", (comp_id,))

    def snapshot_compensation(self, comp_id: int) -> dict[str, Any]:
        c = self.get_compensation(comp_id)
        snap = {k: c[k] for k in c.keys()}
        snap["id"] = int(c["id"])
        snap["employee_id"] = int(c["employee_id"])
        if c["unit"] == "days" and c["method"] == "day_off":
            snap["day_off_dates"] = self.get_comp_dates(int(c["id"]))
        else:
            snap["day_off_dates"] = []
        return snap

    def restore_compensation(self, snap: dict[str, Any]) -> int:
        employee_id = int(snap["employee_id"])
        unit = snap["unit"]
        method = snap["method"]
        comment = snap.get("comment") or ""
        if method == "money":
            order_no = snap.get("order_no") or ""
            order_date = d_parse(snap.get("order_date") or snap.get("event_date"))
            if unit == "hours":
                return self.add_compensation_money(
                    employee_id, "hours", int(snap.get("amount_minutes") or 0), None, order_no, order_date, comment
                )
            else:
                return self.add_compensation_money(
                    employee_id, "days", None, int(snap.get("amount_days") or 0), order_no, order_date, comment
                )
        else:
            if unit == "hours":
                event_date = d_parse(snap.get("event_date"))
                return self.add_compensation_hours_dayoff(
                    employee_id, event_date, int(snap.get("amount_minutes") or 0), comment
                )
            else:
                dates = [d_parse(x) for x in (snap.get("day_off_dates") or [])]
                return self.add_compensation_days_dayoff(employee_id, dates, comment)

    def list_comp_events_for_year(self, employee_id: int, year: int) -> list[dict[str, Any]]:
        y0 = date(year, 1, 1)
        y1 = date(year + 1, 1, 1)
        events: list[dict[str, Any]] = []

        rows = self.conn.execute(
            """
            SELECT unit, method, event_date, amount_minutes, amount_days, order_no
            FROM compensation
            WHERE employee_id=?
              AND event_date IS NOT NULL
              AND event_date >= ? AND event_date < ?
            """,
            (employee_id, d_iso(y0), d_iso(y1)),
        ).fetchall()
        for r in rows:
            d0 = d_parse(r["event_date"])
            if r["unit"] == "hours":
                events.append(
                    {
                        "date": d0,
                        "hours": int(r["amount_minutes"] or 0),
                        "days": 0,
                        "desc": f"компенсация часов ({'приказ '+(r['order_no'] or '') if r['method']=='money' else 'выходной'}) {d_iso(d0)}",
                    }
                )
            else:
                events.append(
                    {
                        "date": d0,
                        "hours": 0,
                        "days": int(r["amount_days"] or 0),
                        "desc": f"компенсация дней (приказ {(r['order_no'] or '')}) {d_iso(d0)}",
                    }
                )

        rows2 = self.conn.execute(
            """
            SELECT day_off_date
            FROM comp_day_off_date
            WHERE employee_id=?
              AND day_off_date >= ? AND day_off_date < ?
            """,
            (employee_id, d_iso(y0), d_iso(y1)),
        ).fetchall()
        for r in rows2:
            d0 = d_parse(r["day_off_date"])
            events.append({"date": d0, "hours": 0, "days": 1, "desc": f"выходной (день) {d_iso(d0)}"})

        events.sort(key=lambda e: (e["date"], e["desc"]))
        return events


# -----------------------------
# Business logic
# -----------------------------
def employee_period_bounds(emp: sqlite3.Row) -> tuple[date, Optional[date]]:
    y, m = parse_month(emp["start_month"])
    start = date(y, m, 1)
    end = d_parse(emp["end_date"]) if emp["end_date"] else None
    return start, end


def validate_event_in_employee_period(emp: sqlite3.Row, event_date: date) -> Optional[str]:
    start, end = employee_period_bounds(emp)
    if event_date < start:
        return "Дата раньше начала работы сотрудника в этом отделе."
    if end is not None and event_date > end:
        return "Дата позже перевода/увольнения сотрудника."
    return None


def opening_effective(emp: sqlite3.Row, year: int, until_exclusive_dt: datetime) -> tuple[int, int]:
    start_y, start_m = parse_month(emp["start_month"])
    start_dt = datetime(start_y, start_m, 1, 0, 0)
    if year != start_y:
        return 0, 0
    if until_exclusive_dt < start_dt:
        return 0, 0
    return int(emp["opening_minutes"] or 0), int(emp["opening_days"] or 0)


def extract_night_intervals(s: datetime, e: datetime) -> list[tuple[datetime, datetime]]:
    out: list[tuple[datetime, datetime]] = []
    day = s.date() - timedelta(days=1)
    last_day = e.date()
    while day <= last_day:
        w0 = datetime.combine(day, time(22, 0))
        w1 = datetime.combine(day + timedelta(days=1), time(6, 0))
        inter = intersect(s, e, w0, w1)
        if inter:
            out.append(inter)
        day += timedelta(days=1)
    return out


def compute_night_minutes(db: DB, employee_id: int, start: datetime, end: datetime) -> int:
    duties = db.list_duties_for_period(employee_id, start, end)
    duty_ids = [int(d["id"]) for d in duties]
    breaks_map = db.breaks_for_duty_ids(duty_ids)

    intervals: list[tuple[datetime, datetime]] = []
    for d in duties:
        did = int(d["id"])
        s0 = max(dt_parse(d["start_dt"]), start)
        e0 = min(dt_parse(d["end_dt"]), end)
        if s0 >= e0:
            continue

        cuts = breaks_map.get(did, [])
        parts = subtract_intervals((s0, e0), cuts)

        for s, e in parts:
            intervals.extend(extract_night_intervals(s, e))

    merged = merge_intervals(intervals)
    return sum(int((e - s).total_seconds() // 60) for s, e in merged)


def compute_nonworking_days(db: DB, employee_id: int, start: datetime, end: datetime) -> int:
    duties = db.list_duties_for_period(employee_id, start, end)
    duty_ids = [int(d["id"]) for d in duties]
    breaks_map = db.breaks_for_duty_ids(duty_ids)

    hit_dates: set[date] = set()

    for row in duties:
        did = int(row["id"])
        ds0 = max(dt_parse(row["start_dt"]), start)
        de0 = min(dt_parse(row["end_dt"]), end)
        if ds0 >= de0:
            continue

        parts = subtract_intervals((ds0, de0), breaks_map.get(did, []))

        for ds, de in parts:
            cur = ds.date()
            last = de.date()
            while cur <= last:
                w0 = datetime.combine(cur, time(6, 0))
                w1 = datetime.combine(cur, time(22, 0))
                if intersect(ds, de, w0, w1) is not None:
                    if not db.is_working_day(cur):
                        hit_dates.add(cur)
                cur += timedelta(days=1)

    return len(hit_dates)


def compute_accrual_hours_up_to(db: DB, employee_id: int, year: int, inclusive_date: date) -> int:
    emp = db.get_employee(employee_id)
    y_start, _ = year_bounds_dt(year)
    end_dt = datetime.combine(inclusive_date + timedelta(days=1), time(0, 0))
    open_m, _ = opening_effective(emp, year, end_dt)
    return open_m + compute_night_minutes(db, employee_id, y_start, end_dt)


def compute_accrual_days_up_to(db: DB, employee_id: int, year: int, inclusive_date: date) -> int:
    emp = db.get_employee(employee_id)
    y_start, _ = year_bounds_dt(year)
    end_dt = datetime.combine(inclusive_date + timedelta(days=1), time(0, 0))
    _, open_d = opening_effective(emp, year, end_dt)
    return open_d + compute_nonworking_days(db, employee_id, y_start, end_dt)


def validate_non_negative_over_year(db: DB, employee_id: int, year: int) -> tuple[bool, str]:
    events = db.list_comp_events_for_year(employee_id, year)
    spent_minutes = 0
    spent_days = 0
    need_check_days = any(e["days"] > 0 for e in events)

    for ev in events:
        spent_minutes += ev["hours"]
        spent_days += ev["days"]

        acc_minutes = compute_accrual_hours_up_to(db, employee_id, year, ev["date"])
        if spent_minutes > acc_minutes:
            need = spent_minutes - acc_minutes
            return False, f"Изменение не применено: после '{ev['desc']}' не хватает {minutes_to_hhmm(need)} часов."

        if need_check_days:
            try:
                acc_days = compute_accrual_days_up_to(db, employee_id, year, ev["date"])
            except CalendarMissingError:
                return False, "Календарь года не создан/неполный. Невозможно проверить баланс по дням."
            if spent_days > acc_days:
                need = spent_days - acc_days
                return False, f"Изменение не применено: после '{ev['desc']}' не хватает {need} дней."

    return True, ""


def compute_month_summary(db: DB, employee_id: int, year: int, month: int) -> dict[str, Any]:
    emp = db.get_employee(employee_id)
    m_start, m_end = month_bounds_dt(year, month)
    y_start, _ = year_bounds_dt(year)

    month_minutes = compute_night_minutes(db, employee_id, m_start, m_end)
    open_m, open_d = opening_effective(emp, year, m_end)
    ytd_minutes = open_m + compute_night_minutes(db, employee_id, y_start, m_end)

    days_ok = True
    days_err = ""
    month_days = 0
    ytd_days = 0
    try:
        month_days = compute_nonworking_days(db, employee_id, m_start, m_end)
        ytd_days = open_d + compute_nonworking_days(db, employee_id, y_start, m_end)
    except CalendarMissingError:
        days_ok = False
        days_err = "Календарь года не создан/неполный: дни за нерабочие не рассчитываются."

    m0 = m_start.date()
    m1 = m_end.date()
    y0 = date(year, 1, 1)

    r = db.conn.execute(
        """
        SELECT COALESCE(SUM(amount_minutes),0) AS m
        FROM compensation
        WHERE employee_id=?
          AND unit='hours'
          AND event_date IS NOT NULL
          AND event_date >= ? AND event_date < ?
        """,
        (employee_id, d_iso(m0), d_iso(m1)),
    ).fetchone()
    comp_month_minutes = int(r["m"])

    r = db.conn.execute(
        """
        SELECT COALESCE(SUM(amount_minutes),0) AS m
        FROM compensation
        WHERE employee_id=?
          AND unit='hours'
          AND event_date IS NOT NULL
          AND event_date >= ? AND event_date < ?
        """,
        (employee_id, d_iso(y0), d_iso(m1)),
    ).fetchone()
    comp_ytd_minutes = int(r["m"])

    r = db.conn.execute(
        """
        SELECT COALESCE(SUM(amount_days),0) AS d
        FROM compensation
        WHERE employee_id=?
          AND unit='days' AND method='money'
          AND event_date IS NOT NULL
          AND event_date >= ? AND event_date < ?
        """,
        (employee_id, d_iso(m0), d_iso(m1)),
    ).fetchone()
    comp_month_days_money = int(r["d"])

    r = db.conn.execute(
        """
        SELECT COALESCE(SUM(amount_days),0) AS d
        FROM compensation
        WHERE employee_id=?
          AND unit='days' AND method='money'
          AND event_date IS NOT NULL
          AND event_date >= ? AND event_date < ?
        """,
        (employee_id, d_iso(y0), d_iso(m1)),
    ).fetchone()
    comp_ytd_days_money = int(r["d"])

    r = db.conn.execute(
        "SELECT COUNT(*) AS c FROM comp_day_off_date WHERE employee_id=? AND day_off_date >= ? AND day_off_date < ?",
        (employee_id, d_iso(m0), d_iso(m1)),
    ).fetchone()
    comp_month_days_dayoff = int(r["c"])

    r = db.conn.execute(
        "SELECT COUNT(*) AS c FROM comp_day_off_date WHERE employee_id=? AND day_off_date >= ? AND day_off_date < ?",
        (employee_id, d_iso(y0), d_iso(m1)),
    ).fetchone()
    comp_ytd_days_dayoff = int(r["c"])

    comp_month_days = comp_month_days_money + comp_month_days_dayoff
    comp_ytd_days = comp_ytd_days_money + comp_ytd_days_dayoff

    bal_minutes = ytd_minutes - comp_ytd_minutes
    bal_days = ytd_days - comp_ytd_days

    return {
        "days_ok": days_ok,
        "days_err": days_err,
        "opening_minutes": int(emp["opening_minutes"] or 0),
        "opening_days": int(emp["opening_days"] or 0),
        "month_minutes": month_minutes,
        "month_days": month_days,
        "comp_month_minutes": comp_month_minutes,
        "comp_month_days": comp_month_days,
        "bal_minutes": bal_minutes,
        "bal_days": bal_days,
    }


# -----------------------------
# Export (Excel)
# -----------------------------
def ensure_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore

        return openpyxl
    except Exception as e:
        raise DBError("Для экспорта нужен пакет openpyxl. Установите: pip install openpyxl") from e


def export_simple_xlsx(db: DB, employee_id: int, year: int, month: int, out_path: str) -> None:
    openpyxl = ensure_openpyxl()
    from openpyxl.styles import Font

    emp = db.get_employee(employee_id)
    summ = compute_month_summary(db, employee_id, year, month)
    dept = db.get_department_settings()

    duties = db.list_duties_for_month(employee_id, year, month)
    comps = db.list_compensations_for_month(employee_id, year, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Итоги"

    fio = f"{emp['last_name']} {emp['first_name']} {emp['middle_name'] or ''}".strip()
    resp_fio = fio_initials_last(dept["resp_last_name"], dept["resp_first_name"], dept["resp_middle_name"])

    ws["A1"] = "Отдел:"
    ws["B1"] = dept["department_name"]
    ws["A2"] = "Сотрудник:"
    ws["B2"] = fio
    ws["A3"] = "Звание:"
    ws["B3"] = emp["rank"] or ""
    ws["A4"] = "Должность:"
    ws["B4"] = emp["position"] or ""
    ws["A5"] = "Период:"
    ws["B5"] = f"{year:04d}-{month:02d}"

    ws["A7"] = "Начислено за месяц (часы):"
    ws["B7"] = minutes_to_hhmm(summ["month_minutes"])
    ws["A8"] = "Начислено за месяц (дни):"
    ws["B8"] = summ["month_days"] if summ["days_ok"] else ""
    ws["A9"] = "Списано за месяц (часы):"
    ws["B9"] = minutes_to_hhmm(summ["comp_month_minutes"])
    ws["A10"] = "Списано за месяц (дни):"
    ws["B10"] = summ["comp_month_days"] if summ["days_ok"] else ""
    ws["A11"] = "Остаток на конец месяца (часы):"
    ws["B11"] = minutes_to_hhmm(summ["bal_minutes"])
    ws["A12"] = "Остаток на конец месяца (дни):"
    ws["B12"] = summ["bal_days"] if summ["days_ok"] else ""

    ws["A14"] = "Ответственный:"
    ws["B14"] = f"{dept['resp_position'] or ''} {dept['resp_rank'] or ''} {resp_fio}".strip()

    for cell in ["A1", "A2", "A3", "A4", "A5", "A7", "A8", "A9", "A10", "A11", "A12", "A14"]:
        ws[cell].font = Font(bold=True)

    ws2 = wb.create_sheet("Дежурства")
    ws2.append(["Начало", "Конец", "Комментарий"])
    for r in duties:
        ws2.append([fmt_dt_iso(r["start_dt"]), fmt_dt_iso(r["end_dt"]), r["comment"] or ""])

    ws3 = wb.create_sheet("Компенсации")
    ws3.append(["Тип", "Способ", "Дата/Приказ", "Количество", "Комментарий"])
    for r in comps:
        unit = r["unit"]
        method = r["method"]
        typ = "Часы" if unit == "hours" else "Дни"
        mth = "Выходной" if method == "day_off" else "Деньги"
        when = ""
        amount = ""
        if method == "money":
            when = f"№{r['order_no']} от {fmt_date_iso(r['order_date'])}"
            amount = minutes_to_hhmm(int(r["amount_minutes"] or 0)) if unit == "hours" else str(int(r["amount_days"] or 0))
        else:
            if unit == "hours":
                when = fmt_date_iso(r["event_date"] or "")
                amount = minutes_to_hhmm(int(r["amount_minutes"] or 0))
            else:
                dates = db.get_comp_dates(int(r["id"]))
                when = ", ".join([fmt_date_iso(x) for x in dates])
                amount = str(len(dates))
        ws3.append([typ, mth, when, amount, r["comment"] or ""])

    wb.save(out_path)
    
class TemplateExporter:
    @staticmethod
    def _hide_daytime_intervals_in_workdays(
        db: DB,
        intervals_by_day: dict[date, list[tuple[datetime, datetime]]],
    ) -> dict[date, list[tuple[datetime, datetime]]]:
        """
        Для ЭКСПОРТА: в рабочие дни скрываем часть 06:00–22:00.
        Оставляем только 00:00–06:00 и 22:00–24:00.
        В нерабочие дни оставляем как есть.
        """
        out: dict[date, list[tuple[datetime, datetime]]] = {}

        for d0, intervals in (intervals_by_day or {}).items():
            try:
                is_working = bool(db.is_working_day(d0))
            except Exception:
                # если календаря нет/неполный — ничего не скрываем
                is_working = False

            if not is_working:
                out[d0] = intervals
                continue

            # рабочий день: оставляем только ночь в пределах этого дня
            w0 = datetime.combine(d0, time(0, 0))
            w1 = datetime.combine(d0, time(6, 0))
            w2 = datetime.combine(d0, time(22, 0))
            w3 = datetime.combine(d0 + timedelta(days=1), time(0, 0))

            parts: list[tuple[datetime, datetime]] = []
            for a, b in intervals:
                inter1 = intersect(a, b, w0, w1)  # 00:00–06:00
                if inter1:
                    parts.append(inter1)

                inter2 = intersect(a, b, w2, w3)  # 22:00–24:00
                if inter2:
                    parts.append(inter2)

            out[d0] = merge_intervals(parts)

        return out    
    
    @staticmethod
    def _safe_cell(ws, row: int, col: int):
        from openpyxl.cell.cell import MergedCell

        cell = ws.cell(row=row, column=col)
        if not isinstance(cell, MergedCell):
            return cell

        for r in ws.merged_cells.ranges:
            if int(r.min_row) <= row <= int(r.max_row) and int(r.min_col) <= col <= int(r.max_col):
                return ws.cell(row=int(r.min_row), column=int(r.min_col))

        return None
        
    @staticmethod
    def month_name_ru(month: int) -> str:
        names = [
            "январь", "февраль", "март", "апрель", "май", "июнь",
            "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
        ]
        return names[month - 1] if 1 <= month <= 12 else str(month)

    @staticmethod
    def _fmt_minutes_ru(minutes: int) -> str:
        sign = "-" if minutes < 0 else ""
        m = abs(int(minutes))
        h = m // 60
        mm = m % 60
        if mm == 0:
            return f"{sign}{h} ч."
        return f"{sign}{h} ч. {mm} мин."

    @staticmethod
    def _fmt_days_ru(days: int) -> str:
        sign = "-" if days < 0 else ""
        d = abs(int(days))
        return f"{sign}{d} д."

    @staticmethod
    def _find_first_cell_with_value(ws, value: str):
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() == value:
                    return cell
        return None

    @staticmethod
    def _replace_markers_everywhere(ws, mapping: dict[str, str]) -> None:
        for row in ws.iter_rows():
            for cell0 in row:
                if not isinstance(cell0.value, str):
                    continue

                s = cell0.value
                s2 = s
                for k, v in mapping.items():
                    if k in s2:
                        s2 = s2.replace(k, v)

                if s2 != s:
                    cell = TemplateExporter._safe_cell(ws, int(cell0.row), int(cell0.column))
                    if cell is not None:
                        cell.value = s2

    @staticmethod
    def _collect_marker_columns_in_row(ws, row: int, markers: set[str]) -> dict[str, list[int]]:
        out: dict[str, list[int]] = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, str):
                s = v.strip()
                if s in markers:
                    out.setdefault(s, []).append(col)
        return out

    @staticmethod
    def _clear_markers_in_row(ws, row: int, markers: set[str]) -> None:
        for col in range(1, ws.max_column + 1):
            cell0 = ws.cell(row=row, column=col)
            v = cell0.value
            if isinstance(v, str) and v.strip() in markers:
                cell = TemplateExporter._safe_cell(ws, row, col)
                if cell is not None:
                    cell.value = None

    @staticmethod
    def _copy_row_style(ws, src_row: int, dst_row: int) -> None:
        from copy import copy
        from openpyxl.cell.cell import MergedCell

        if ws.row_dimensions.get(src_row) and ws.row_dimensions[src_row].height:
            ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

        for col in range(1, ws.max_column + 1):
            s = ws.cell(row=src_row, column=col)
            d = ws.cell(row=dst_row, column=col)

            if isinstance(d, MergedCell):
                continue

            d._style = copy(s._style)
            d.font = copy(s.font)
            d.border = copy(s.border)
            d.fill = copy(s.fill)
            d.number_format = s.number_format
            d.protection = copy(s.protection)
            d.alignment = copy(s.alignment)

    @staticmethod
    def _row_merge_patterns(ws, row: int) -> list[tuple[int, int]]:
        """Список объединений (min_col, max_col), которые лежат ТОЛЬКО в этой строке."""
        out: list[tuple[int, int]] = []
        try:
            ranges = list(ws.merged_cells.ranges)
        except Exception:
            ranges = []
        for r in ranges:
            if int(r.min_row) == row and int(r.max_row) == row:
                out.append((int(r.min_col), int(r.max_col)))
        return out

    @staticmethod
    def _apply_row_merge_patterns(ws, row: int, patterns: list[tuple[int, int]]) -> None:
        for min_c, max_c in patterns:
            try:
                ws.merge_cells(start_row=row, end_row=row, start_column=min_c, end_column=max_c)
            except Exception:
                pass

    @staticmethod
    def _sum_day_minutes(intervals: list[tuple[datetime, datetime]]) -> int:
        return sum(int((b - a).total_seconds() // 60) for a, b in intervals)

    @staticmethod
    def _sum_comp_rest_hours_minutes(db: DB, employee_id: int, year: int, month: int) -> int:
        m0 = date(year, month, 1)
        ny, nm = next_month(year, month)
        m1 = date(ny, nm, 1)
        r = db.conn.execute(
            """
            SELECT COALESCE(SUM(amount_minutes),0) AS m
            FROM compensation
            WHERE employee_id=?
              AND unit='hours' AND method='day_off'
              AND event_date IS NOT NULL
              AND event_date >= ? AND event_date < ?
            """,
            (employee_id, d_iso(m0), d_iso(m1)),
        ).fetchone()
        return int(r["m"])

    @staticmethod
    def _count_comp_rest_days(db: DB, employee_id: int, year: int, month: int) -> int:
        m0 = date(year, month, 1)
        ny, nm = next_month(year, month)
        m1 = date(ny, nm, 1)
        r = db.conn.execute(
            """
            SELECT COUNT(*) AS c
            FROM comp_day_off_date
            WHERE employee_id=?
              AND day_off_date >= ? AND day_off_date < ?
            """,
            (employee_id, d_iso(m0), d_iso(m1)),
        ).fetchone()
        return int(r["c"])

    @staticmethod
    def _money_comp_text_for_month(db: DB, employee_id: int, year: int, month: int) -> str:
        m0 = date(year, month, 1)
        ny, nm = next_month(year, month)
        m1 = date(ny, nm, 1)

        rows = db.conn.execute(
            """
            SELECT unit, amount_minutes, amount_days, order_no, order_date, event_date
            FROM compensation
            WHERE employee_id=?
              AND method='money'
              AND event_date IS NOT NULL
              AND event_date >= ? AND event_date < ?
            ORDER BY COALESCE(order_date,event_date), order_no, id
            """,
            (employee_id, d_iso(m0), d_iso(m1)),
        ).fetchall()

        grouped: dict[tuple[str, str], dict[str, int]] = {}
        for r in rows:
            order_no = (r["order_no"] or "").strip()
            order_date = (r["order_date"] or r["event_date"] or "").strip()
            key = (order_date, order_no)

            g = grouped.setdefault(key, {"minutes": 0, "days": 0})
            if r["unit"] == "hours":
                g["minutes"] += int(r["amount_minutes"] or 0)
            else:
                g["days"] += int(r["amount_days"] or 0)

        out: list[str] = []
        for (order_date, order_no), g in sorted(grouped.items(), key=lambda x: (x[0][0], x[0][1])):
            parts: list[str] = []
            if g["minutes"] > 0:
                parts.append(TemplateExporter._fmt_minutes_ru(g["minutes"]))
            if g["days"] > 0:
                parts.append(TemplateExporter._fmt_days_ru(g["days"]))

            first_part = " и ".join(parts).strip()
            if not first_part:
                continue

            od = fmt_date_iso(order_date)
            if order_no:
                out.append(f"{first_part} Приказом от {od} №{order_no}")
            else:
                out.append(f"{first_part} Приказом от {od}")

        return "\n".join(out)

    @staticmethod
    def _balance_at_start_of_month(db: DB, employee_id: int, year: int, month: int) -> tuple[int, int]:
        emp = db.get_employee(employee_id)
        y_start, _ = year_bounds_dt(year)
        m_start_dt = datetime(year, month, 1, 0, 0)

        open_m, open_d = opening_effective(emp, year, m_start_dt)
        acc_m = open_m + compute_night_minutes(db, employee_id, y_start, m_start_dt)

        r = db.conn.execute(
            """
            SELECT COALESCE(SUM(amount_minutes),0) AS m
            FROM compensation
            WHERE employee_id=?
              AND unit='hours'
              AND event_date IS NOT NULL
              AND event_date >= ? AND event_date < ?
            """,
            (employee_id, d_iso(date(year, 1, 1)), d_iso(m_start_dt.date())),
        ).fetchone()
        spent_m = int(r["m"])
        bal_m = acc_m - spent_m

        bal_d = 0
        try:
            acc_d = open_d + compute_nonworking_days(db, employee_id, y_start, m_start_dt)

            r1 = db.conn.execute(
                """
                SELECT COALESCE(SUM(amount_days),0) AS d
                FROM compensation
                WHERE employee_id=?
                  AND unit='days' AND method='money'
                  AND event_date IS NOT NULL
                  AND event_date >= ? AND event_date < ?
                """,
                (employee_id, d_iso(date(year, 1, 1)), d_iso(m_start_dt.date())),
            ).fetchone()
            spent_d_money = int(r1["d"])

            r2 = db.conn.execute(
                """
                SELECT COUNT(*) AS c
                FROM comp_day_off_date
                WHERE employee_id=?
                  AND day_off_date >= ? AND day_off_date < ?
                """,
                (employee_id, d_iso(date(year, 1, 1)), d_iso(m_start_dt.date())),
            ).fetchone()
            spent_d_dayoff = int(r2["c"])

            bal_d = acc_d - (spent_d_money + spent_d_dayoff)
        except CalendarMissingError:
            bal_d = 0

        return bal_m, bal_d

    @staticmethod
    def export(db: DB, year: int, month: int, template_path: str, out_path: str, sheet_name: str = "Лист1") -> None:
        openpyxl = ensure_openpyxl()
        from openpyxl.styles import Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.cell_range import CellRange

        wb = openpyxl.load_workbook(template_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        start_cell = TemplateExporter._find_first_cell_with_value(ws, "{{EMP_START}}")
        if not start_cell:
            raise DBError("В шаблоне не найдена метка {{EMP_START}} (строка первого сотрудника).")

        start_row = int(start_cell.row)
        start_cell.value = None

        dept = db.get_department_settings()
        resp_fio = fio_initials_last(dept["resp_last_name"], dept["resp_first_name"], dept["resp_middle_name"])

        base_markers: set[str] = {
            "{{EMP_NO}}",
            "{{EMP_FIO_RANK_POS}}",
            "{{OPEN_HOURS}}",
            "{{OPEN_DAYS}}",
            "{{COMP_REST_HOURS}}",
            "{{COMP_REST_DAYS}}",
            "{{COMP_MONEY}}",
            "{{BAL_END_HOURS}}",
            "{{BAL_END_DAYS}}",
        }
        day_markers = {f"{{{{DAY_{d:02d}}}}}" for d in range(1, 32)}
        all_markers = set(base_markers) | set(day_markers)

        marker_cols = TemplateExporter._collect_marker_columns_in_row(ws, start_row, all_markers)
        if not marker_cols:
            raise DBError("В строке {{EMP_START}} не найдены метки ({{EMP_NO}}, {{EMP_FIO_RANK_POS}}, {{DAY_01}}...).")

        emps = db.list_employees_for_month(year, month, active_only=True, search="")
        n = len(emps)

        if n > 1:
            try:
                ws.insert_rows(start_row + 1, amount=n - 1)
            except TypeError:
                for _ in range(n - 1):
                    ws.insert_rows(start_row + 1)

        merge_patterns = TemplateExporter._row_merge_patterns(ws, start_row)
        for r in range(start_row + 1, start_row + n):
            TemplateExporter._copy_row_style(ws, start_row, r)
            TemplateExporter._apply_row_merge_patterns(ws, r, merge_patterns)

        TemplateExporter._replace_markers_everywhere(
            ws,
            {
                "{{HDR_DEPARTMENT}}": dept["department_name"],
                "{{HDR_MONTH_NAME}}": TemplateExporter.month_name_ru(month),
                "{{HDR_YEAR}}": str(year),
            },
        )

        ny, nm = next_month(year, month)
        last_day = (date(ny, nm, 1) - timedelta(days=1)).day
        m0 = date(year, month, 1)
        m1 = date(ny, nm, 1)

        for d in range(last_day + 1, 32):
            marker = f"{{{{DAY_{d:02d}}}}}"
            cols = marker_cols.get(marker, [])
            for col in cols:
                ws.column_dimensions[get_column_letter(col)].hidden = True

        yellow_fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")

        nonwork_days: set[int] = set()
        try:
            rows = db.conn.execute(
                "SELECT date FROM calendar_day WHERE is_working=0 AND date>=? AND date<?",
                (d_iso(m0), d_iso(m1)),
            ).fetchall()
            for r0 in rows:
                try:
                    nonwork_days.add(d_parse(r0["date"]).day)
                except Exception:
                    pass
        except Exception:
            nonwork_days = set()

        top_fill_row = max(1, start_row - 3)
        for d in sorted(nonwork_days):
            if not (1 <= d <= last_day):
                continue
            marker = f"{{{{DAY_{d:02d}}}}}"
            cols = marker_cols.get(marker, [])
            if not cols:
                continue
            for col in cols:
                for rr in range(top_fill_row, start_row + n):
                    c = TemplateExporter._safe_cell(ws, rr, col)
                    if c is None:
                        continue
                    c.fill = yellow_fill

        slash = "⁄"

        sup_map = {
            "0": "⁰", "1": "¹", "2": "²", "3": "³", "4": "⁴",
            "5": "⁵", "6": "⁶", "7": "⁷", "8": "⁸", "9": "⁹",
        }
        sub_map = {
            "0": "₀", "1": "₁", "2": "₂", "3": "₃", "4": "₄",
            "5": "₅", "6": "₆", "7": "₇", "8": "₈", "9": "₉",
        }

        def to_sup(s: str) -> str:
            return "".join(sup_map.get(ch, ch) for ch in s)

        def to_sub(s: str) -> str:
            return "".join(sub_map.get(ch, ch) for ch in s)

        def fmt_t(dt: datetime) -> str:
            hh = dt.strftime("%H")
            mm = dt.strftime("%M")
            return hh if mm == "00" else f"{hh}:{mm}"

        def duty_fraction(day_date: date, intervals: list[tuple[datetime, datetime]]) -> str:
            parts: list[str] = []
            for a, b in intervals:
                parts.append(f"{to_sup(fmt_t(a))}{slash}{to_sub(fmt_t(b))}")
            return "\n".join(parts)

        def fmt_comp_hours(minutes: int) -> str:
            m = int(minutes)
            h = m // 60
            mm = m % 60
            return f"{h}" if mm == 0 else f"{h}:{mm:02d}"

        def comp_fraction(label: str, bottom: str) -> str:
            return "В"

        for i, emp in enumerate(emps):
            row = start_row + i
            eid = int(emp["id"])

            fio = f"{emp['last_name']} {emp['first_name']} {emp['middle_name'] or ''}".strip()
            rank = (emp["rank"] or "").strip()
            pos = (emp["position"] or "").strip()

            parts: list[str] = [fio]
            if rank:
                parts.append(rank)
            parts.append("")
            if pos:
                parts.append(pos)
            fio_rank_pos = "\n".join(parts).rstrip()

            start_bal_m, start_bal_d = TemplateExporter._balance_at_start_of_month(db, eid, year, month)
            summ = compute_month_summary(db, eid, year, month)

            values: dict[str, str] = {
                "{{EMP_NO}}": str(i + 1),
                "{{EMP_FIO_RANK_POS}}": fio_rank_pos,
                "{{OPEN_HOURS}}": TemplateExporter._fmt_minutes_ru(start_bal_m),
                "{{OPEN_DAYS}}": TemplateExporter._fmt_days_ru(start_bal_d),
                "{{COMP_REST_HOURS}}": TemplateExporter._fmt_minutes_ru(
                    TemplateExporter._sum_comp_rest_hours_minutes(db, eid, year, month)
                ),
                "{{COMP_REST_DAYS}}": TemplateExporter._fmt_days_ru(
                    TemplateExporter._count_comp_rest_days(db, eid, year, month)
                ),
                "{{COMP_MONEY}}": TemplateExporter._money_comp_text_for_month(db, eid, year, month),
                "{{BAL_END_HOURS}}": TemplateExporter._fmt_minutes_ru(int(summ["bal_minutes"])),
                "{{BAL_END_DAYS}}": (TemplateExporter._fmt_days_ru(int(summ["bal_days"])) if summ["days_ok"] else ""),
            }

            intervals_by_day, _ = compute_day_intervals_in_month(db, eid, year, month)
            intervals_by_day = TemplateExporter._hide_daytime_intervals_in_workdays(db, intervals_by_day)
            
            comp_by_day: dict[date, list[str]] = {}

            rows_h = db.conn.execute(
                """
                SELECT event_date, amount_minutes
                FROM compensation
                WHERE employee_id=?
                  AND unit='hours' AND method='day_off'
                  AND event_date IS NOT NULL
                  AND event_date >= ? AND event_date < ?
                """,
                (eid, d_iso(m0), d_iso(m1)),
            ).fetchall()

            for rr in rows_h:
                try:
                    d0 = d_parse(rr["event_date"])
                except Exception:
                    continue
                mins = int(rr["amount_minutes"] or 0)
                if mins <= 0:
                    continue
                comp_by_day.setdefault(d0, []).append(comp_fraction("ДВО", fmt_comp_hours(mins)))

            rows_d = db.conn.execute(
                """
                SELECT cd.day_off_date, c.amount_days
                FROM comp_day_off_date cd
                JOIN compensation c ON c.id = cd.compensation_id
                WHERE cd.employee_id=?
                  AND cd.day_off_date >= ? AND cd.day_off_date < ?
                """,
                (eid, d_iso(m0), d_iso(m1)),
            ).fetchall()

            for rr in rows_d:
                try:
                    d0 = d_parse(rr["day_off_date"])
                except Exception:
                    continue
                total = int(rr["amount_days"] or 1)
                comp_by_day.setdefault(d0, []).append(comp_fraction("ДДО", "1"))

            for marker, cols in marker_cols.items():
                if marker in values:
                    for col in cols:
                        c = TemplateExporter._safe_cell(ws, row, col)
                        if c is None:
                            continue
                        c.value = values[marker]
                        if marker in ("{{EMP_FIO_RANK_POS}}",):
                            c.alignment = Alignment(wrap_text=True, vertical="top")

            for d in range(1, 32):
                marker = f"{{{{DAY_{d:02d}}}}}"
                cols = marker_cols.get(marker, [])
                if not cols:
                    continue

                txt = ""
                if d <= last_day:
                    day_date = date(year, month, d)
                    intervals = intervals_by_day.get(day_date, [])
                    if intervals:
                        txt = duty_fraction(day_date, intervals)

                    comps = comp_by_day.get(day_date, [])
                    if comps:
                        if txt:
                            txt = txt + "\n" + "\n".join(comps)
                        else:
                            txt = "\n".join(comps)

                for col in cols:
                    c = TemplateExporter._safe_cell(ws, row, col)
                    if c is None:
                        continue
                    c.value = txt
                    if txt:
                        c.alignment = c.alignment.copy(wrap_text=True)

        sign_rows: list[int] = []
        for marker in ("{{SIGN_POS}}", "{{SIGN_RANK}}", "{{SIGN_FIO}}"):
            for row0 in ws.iter_rows():
                for cell0 in row0:
                    if isinstance(cell0.value, str) and cell0.value.strip() == marker:
                        sign_rows.append(int(cell0.row))

        bottom_row = max(sign_rows) if sign_rows else ws.max_row
        bottom_row = bottom_row + 2

        TemplateExporter._replace_markers_everywhere(
            ws,
            {
                "{{SIGN_POS}}": dept["resp_position"] or "",
                "{{SIGN_RANK}}": dept["resp_rank"] or "",
                "{{SIGN_FIO}}": resp_fio,
            },
        )

        current_pa = ws.print_area
        if current_pa:
            if isinstance(current_pa, (list, tuple)):
                current_pa = current_pa[0]

            if isinstance(current_pa, str):
                pa_str = current_pa.split(",")[0]
                pa_str = pa_str.split("!")[-1].replace("$", "")

                try:
                    cr = CellRange(pa_str)
                    left_col = get_column_letter(int(cr.min_col))
                    right_col = get_column_letter(int(cr.max_col))
                    top_row_pa = int(cr.min_row)
                    ws.print_area = f"{left_col}{top_row_pa}:{right_col}{bottom_row}"
                except Exception:
                    ws.print_area = f"A1:{get_column_letter(ws.max_column)}{bottom_row}"
            else:
                ws.print_area = f"A1:{get_column_letter(ws.max_column)}{bottom_row}"
        else:
            ws.print_area = f"A1:{get_column_letter(ws.max_column)}{bottom_row}"

        TemplateExporter._clear_markers_in_row(ws, start_row, all_markers)
        wb.save(out_path)
    
def _which(cmd: str) -> Optional[str]:
    try:
        return shutil.which(cmd)
    except Exception:
        return None


def convert_xlsx_to_pdf_libreoffice(xlsx_path: str, out_pdf_path: str) -> None:
    soffice = _which("soffice") or _which("libreoffice")
    if not soffice:
        raise DBError("Не найден LibreOffice (soffice). Установите LibreOffice для печати.")

    out_dir = str(Path(out_pdf_path).resolve().parent)
    xlsx_path = str(Path(xlsx_path).resolve())

    p = subprocess.run(
        [
            soffice,
            "--headless",
            "--nologo",
            "--nolockcheck",
            "--norestore",
            "--convert-to",
            "pdf",
            "--outdir",
            out_dir,
            xlsx_path,
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )
    if p.returncode != 0:
        raise DBError(f"LibreOffice не смог конвертировать в PDF.\n{p.stderr or p.stdout}")

    produced = str(Path(out_dir) / (Path(xlsx_path).stem + ".pdf"))
    if not os.path.exists(produced):
        raise DBError("LibreOffice не создал PDF (файл не найден после конвертации).")

    if str(Path(produced).resolve()) != str(Path(out_pdf_path).resolve()):
        shutil.move(produced, out_pdf_path)


def convert_xlsx_to_pdf_excel_windows(xlsx_path: str, out_pdf_path: str) -> None:
    try:
        import win32com.client  # type: ignore
    except Exception as e:
        raise DBError("pywin32 не установлен, конвертация через Excel недоступна.") from e

    xlsx_path = str(Path(xlsx_path).resolve())
    out_pdf_path = str(Path(out_pdf_path).resolve())

    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(xlsx_path)
        wb.ExportAsFixedFormat(0, out_pdf_path)
    except Exception as e:
        raise DBError(f"Не удалось конвертировать через Excel: {e}")
    finally:
        try:
            if wb is not None:
                wb.Close(False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass


def convert_xlsx_to_pdf_auto(xlsx_path: str, out_pdf_path: str) -> None:
    if sys.platform.startswith("win"):
        try:
            convert_xlsx_to_pdf_excel_windows(xlsx_path, out_pdf_path)
            return
        except Exception:
            convert_xlsx_to_pdf_libreoffice(xlsx_path, out_pdf_path)
            return

    convert_xlsx_to_pdf_libreoffice(xlsx_path, out_pdf_path)
    
    # -----------------------------
# Dialogs helpers
# -----------------------------
def program_base_dir() -> Path:
    """
    Папка, где лежит программа:
    - если приложение собрано (PyInstaller), то рядом с .exe
    - если запущен .py, то папка скрипта
    """
    try:
        if getattr(sys, "frozen", False):
            return Path(sys.executable).resolve().parent
    except Exception:
        pass
    return Path(sys.argv[0]).resolve().parent


def sanitize_filename(name: str) -> str:
    s = (name or "").strip()
    bad = '<>:"/\\|?*'
    for ch in bad:
        s = s.replace(ch, "_")
    s = " ".join(s.split())
    return s or "database"


def can_write_dir(d: Path) -> bool:
    try:
        d.mkdir(parents=True, exist_ok=True)
        test = d / f".write_test_{uuid.uuid4().hex}.tmp"
        test.write_text("ok", encoding="utf-8")
        test.unlink()
        return True
    except Exception:
        return False


def make_unique_db_path(base_dir: Path, base_name: str) -> Path:
    base_dir.mkdir(parents=True, exist_ok=True)
    safe = sanitize_filename(base_name)
    p = base_dir / f"{safe}.sqlite"
    i = 2
    while p.exists():
        p = base_dir / f"{safe}_{i}.sqlite"
        i += 1
    return p


class WelcomeWidget(QWidget):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(40, 40, 40, 40) # БОЛЬШЕ ВОЗДУХА
        lay.setSpacing(12)

        self.lbl_img = QLabel()
        self.lbl_img.setAlignment(Qt.AlignCenter)
        lay.addWidget(self.lbl_img, 1)

        self.lbl_title = QLabel("Добро пожаловать")
        self.lbl_title.setAlignment(Qt.AlignCenter)
        # Современная типографика (крупно и жирно)
        self.lbl_title.setStyleSheet("font-size: 28px; font-weight: 700; color: #FFFFFF;")
        lay.addWidget(self.lbl_title, 0)

        self.lbl_sub = QLabel("OvertimeTab загружается…")
        self.lbl_sub.setAlignment(Qt.AlignCenter)
        # Серый цвет текста для второстепенной информации
        self.lbl_sub.setStyleSheet("font-size: 15px; color: #8E8E93;")
        lay.addWidget(self.lbl_sub, 0)

        lay.addStretch(2)
        self._pix_orig: QPixmap | None = None
        self._try_load_pixmap()

    def _try_load_pixmap(self) -> None:
        p = program_base_dir() / "welcome.png"
        if p.exists():
            pm = QPixmap(str(p))
            if not pm.isNull():
                self._pix_orig = pm
        self._update_pixmap()

    def _update_pixmap(self) -> None:
        if not self._pix_orig:
            self.lbl_img.clear()
            return

        max_w = max(200, min(self.width() - 48, 900))
        max_h = max(120, min(int(self.height() * 0.45), 360))

        pm = self._pix_orig.scaled(max_w, max_h, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.lbl_img.setPixmap(pm)

    def resizeEvent(self, event) -> None:  # type: ignore[override]
        super().resizeEvent(event)
        self._update_pixmap()


class CreateSubdivisionDialog(QDialog):
    """
    Первый запуск: создаём базу подразделения и сразу задаём ответственного.
    Файл базы пытаемся создать рядом с программой (если можно), иначе в папке данных пользователя.
    """

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle("Создание подразделения")
        self.resize(480, 400) # Чуть изменили пропорции

        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(20)

        # Главный заголовок внутри диалога (стиль macOS)
        title = QLabel("Настройка базы")
        title.setStyleSheet("font-size: 20px; font-weight: bold;")
        lay.addWidget(title)

        form = QFormLayout()
        form.setVerticalSpacing(12)

        self.ed_name = QLineEdit()
        self.ed_name.setPlaceholderText("Например: Отдел разработки")
        form.addRow("Название:", self.ed_name)

        # Визуальное разделение без жестких линий
        lbl_resp = QLabel("Ответственный за табель")
        lbl_resp.setStyleSheet("font-size: 14px; font-weight: 600; color: #8E8E93; margin-top: 15px;")
        form.addRow(lbl_resp)

        self.ed_pos = QLineEdit()
        self.ed_rank = QLineEdit()
        self.ed_last = QLineEdit()
        self.ed_first = QLineEdit()
        self.ed_mid = QLineEdit()

        form.addRow("Должность:", self.ed_pos)
        form.addRow("Звание:", self.ed_rank)
        form.addRow("Фамилия:", self.ed_last)
        form.addRow("Имя:", self.ed_first)
        form.addRow("Отчество:", self.ed_mid)

        lay.addLayout(form)

        self.lbl_hint = QLabel("База данных будет создана локально.")
        self.lbl_hint.setStyleSheet("color: #636366; font-size: 12px;")
        lay.addWidget(self.lbl_hint)

        lay.addStretch()

        bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        # Делаем кнопку Save акцентной синей
        bb.button(QDialogButtonBox.Save).setObjectName("AccentButton") 
        lay.addWidget(bb)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)

    def values(self) -> dict[str, Any]:
        return {
            "department_name": (self.ed_name.text() or "").strip(),
            "resp_position": (self.ed_pos.text() or "").strip() or None,
            "resp_rank": (self.ed_rank.text() or "").strip() or None,
            "resp_last_name": (self.ed_last.text() or "").strip() or None,
            "resp_first_name": (self.ed_first.text() or "").strip() or None,
            "resp_middle_name": (self.ed_mid.text() or "").strip() or None,
        }

    def accept(self) -> None:
        v = self.values()
        if not v["department_name"]:
            QMessageBox.warning(self, "Проверка", "Введите название подразделения.")
            return
        super().accept()

def simple_text_input(parent: QWidget, title: str, prompt: str) -> tuple[str, bool]:
    dlg = QDialog(parent)
    dlg.setWindowTitle(title)
    lay = QVBoxLayout(dlg)
    lay.addWidget(QLabel(prompt))
    edit = QLineEdit()
    lay.addWidget(edit)
    bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
    lay.addWidget(bb)
    bb.accepted.connect(dlg.accept)
    bb.rejected.connect(dlg.reject)
    ok = dlg.exec() == QDialog.Accepted
    return edit.text(), ok

class PdfPreviewDialog(QDialog):
    def __init__(self, pdf_path: str, parent: QWidget | None = None):
        super().__init__(parent)
        self.pdf_path = pdf_path

        self.setWindowTitle("Предпросмотр печати")
        self.resize(900, 700)

        lay = QVBoxLayout(self)

        self.doc = QPdfDocument(self)
        err = self.doc.load(pdf_path)
        if err != QPdfDocument.Error.None_:
            raise DBError(f"Не удалось открыть PDF для предпросмотра: {err}")

        self.view = QPdfView(self)
        self.view.setDocument(self.doc)
        self.view.setZoomMode(QPdfView.ZoomMode.FitToWidth)
        lay.addWidget(self.view, 1)

        btns = QHBoxLayout()
        self.btn_print = QPushButton("Печать…")
        self.btn_save = QPushButton("Сохранить PDF…")
        self.btn_close = QPushButton("Закрыть")
        btns.addWidget(self.btn_print)
        btns.addWidget(self.btn_save)
        btns.addStretch(1)
        btns.addWidget(self.btn_close)
        lay.addLayout(btns)

        self.btn_print.clicked.connect(self.on_print)
        self.btn_save.clicked.connect(self.on_save_pdf)
        self.btn_close.clicked.connect(self.reject)

    def on_save_pdf(self) -> None:
        out, _ = QFileDialog.getSaveFileName(self, "Сохранить PDF", "табель.pdf", "PDF (*.pdf)")
        if not out:
            return
        if not out.lower().endswith(".pdf"):
            out += ".pdf"
        shutil.copy2(self.pdf_path, out)

    def on_print(self) -> None:
        printer = QPrinter(QPrinter.HighResolution)
        dlg = QPrintDialog(printer, self)
        if dlg.exec() != QDialog.Accepted:
            return
    
        paint_pdf_to_printer(self.doc, printer)

def paint_pdf_to_printer(doc: QPdfDocument, printer: QPrinter) -> None:
    """
    Печатает PDF так, чтобы поля были как в самом PDF:
    - ориентация по 1-й странице PDF
    - печать в размер бумаги (paperRect), а не в pageRect
    - масштаб по физическому размеру PDF (points -> DPI)
    """
    from PySide6.QtCore import QSize, QRect, QMarginsF
    from PySide6.QtGui import QPainter, QPageLayout

    # Пытаемся убрать "системные" поля Qt (но аппаратные поля принтера всё равно могут остаться)
    try:
        printer.setFullPage(True)
    except Exception:
        pass

    try:
        printer.setPageMargins(QMarginsF(0, 0, 0, 0), QPageLayout.Unit.Millimeter)
    except Exception:
        try:
            printer.setPageMargins(QMarginsF(0, 0, 0, 0))
        except Exception:
            pass

    # Ориентация по PDF (по 1-й странице)
    try:
        ps0 = doc.pagePointSize(0)  # QSizeF (в поинтах)
        if ps0.width() > ps0.height():
            printer.setPageOrientation(QPageLayout.Orientation.Landscape)
        else:
            printer.setPageOrientation(QPageLayout.Orientation.Portrait)
    except Exception:
        pass

    painter = QPainter()
    if not painter.begin(printer):
        raise DBError("Не удалось начать печать.")

    try:
        dpi = int(printer.resolution())  # DPI принтера

        page_count = int(doc.pageCount())
        for page in range(page_count):
            if page > 0:
                printer.newPage()

            # Берем ВСЮ бумагу, а не область печати (pageRect)
            paperf = printer.paperRect(QPrinter.Unit.DevicePixel)  # обычно QRectF
            paper_w = int(paperf.width())
            paper_h = int(paperf.height())

            # Размер PDF страницы в поинтах (1 point = 1/72 inch)
            ps = doc.pagePointSize(page)  # QSizeF
            pdf_w_points = float(ps.width())
            pdf_h_points = float(ps.height())

            # Переводим "физический" размер PDF в пиксели принтера
            # inches = points / 72
            pdf_w_px = int((pdf_w_points / 72.0) * dpi) if pdf_w_points > 0 else paper_w
            pdf_h_px = int((pdf_h_points / 72.0) * dpi) if pdf_h_points > 0 else paper_h

            pdf_w_px = max(1, pdf_w_px)
            pdf_h_px = max(1, pdf_h_px)

            # Если физически не помещается на бумагу (из-за принтера/драйвера) — уменьшаем, но пропорционально
            scale = min(paper_w / pdf_w_px, paper_h / pdf_h_px) if paper_w > 0 and paper_h > 0 else 1.0
            scale = min(1.0, scale)  # увеличивать НЕ будем, только уменьшать при необходимости

            w = max(1, int(pdf_w_px * scale))
            h = max(1, int(pdf_h_px * scale))

            img = doc.render(page, QSize(w, h))
            if img.isNull():
                continue

            # Центрируем на листе
            x = int(paperf.x()) + (paper_w - w) // 2
            y = int(paperf.y()) + (paper_h - h) // 2

            painter.drawImage(QRect(x, y, w, h), img)

    finally:
        painter.end()

class BaseSelectorDialog(QDialog):
    def __init__(self, cfg: AppConfig, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle("Выбор базы (отдела)")
        self.resize(550, 400)
        self.cfg = cfg
        self.selected_path: Optional[str] = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        title = QLabel("Выберите базу данных")
        title.setStyleSheet("font-size: 18px; font-weight: bold;")
        layout.addWidget(title)

        self.list = QListWidget()
        self.list.setStyleSheet("QListWidget { font-size: 14px; padding: 5px; } QListWidget::item { padding: 10px; border-radius: 6px; }")
        layout.addWidget(self.list)

        btns = QHBoxLayout()
        
        # Кнопка с выпадающим меню (Mac-style)
        self.btn_more = QPushButton("Дополнительно ▾")
        menu = QMenu(self)
        menu.addAction("Подключить существующую...", self.on_attach)
        menu.addAction("Импортировать (копировать)...", self.on_import_copy)
        menu.addAction("Экспорт копии базы...", self.on_export_copy)
        menu.addSeparator()
        menu.addAction("Убрать из списка", self.on_remove)
        self.btn_more.setMenu(menu)

        self.btn_create = QPushButton("Создать базу...")
        self.btn_open = QPushButton("Открыть")
        self.btn_open.setObjectName("AccentButton") # Синяя акцентная кнопка

        btns.addWidget(self.btn_more)
        btns.addStretch(1)
        btns.addWidget(self.btn_create)
        btns.addWidget(self.btn_open)
        layout.addLayout(btns)

        self.btn_create.clicked.connect(self.on_create)
        self.btn_open.clicked.connect(self.on_open)
        self.list.itemDoubleClicked.connect(self.on_open) # Открытие по двойному клику

        self.refresh()

    def refresh(self) -> None:
        self.list.clear()
        self.cfg.db_paths = [p for p in self.cfg.db_paths if os.path.exists(p)]
        for p in self.cfg.db_paths:
            name = p
            try:
                db = DB(p)
                name = db.get_department_name()
                db.close()
            except Exception:
                name = f"(неизвестная база) {Path(p).name}"
            it = QListWidgetItem(f"{name}\n{p}")
            it.setData(Qt.UserRole, p)
            self.list.addItem(it)

        if self.cfg.last_db_path:
            for i in range(self.list.count()):
                if self.list.item(i).data(Qt.UserRole) == self.cfg.last_db_path:
                    self.list.setCurrentRow(i)
                    break

    def current_path(self) -> Optional[str]:
        it = self.list.currentItem()
        return it.data(Qt.UserRole) if it else None

    def on_create(self) -> None:
        name, ok = simple_text_input(self, "Название отдела", "Введите название отдела:")
        if not ok or not name.strip():
            return
        default_dir = str(app_db_store_dir())
        path, _ = QFileDialog.getSaveFileName(
            self, "Создать базу отдела", os.path.join(default_dir, f"{name}.sqlite"), "SQLite (*.sqlite)"
        )
        if not path:
            return
        if not path.endswith(".sqlite"):
            path += ".sqlite"
        if os.path.exists(path):
            QMessageBox.warning(self, "Ошибка", "Файл уже существует.")
            return

        db = DB(path)
        db.update_department_settings(department_name=name.strip())
        db.conn.commit()
        db.close()

        if path not in self.cfg.db_paths:
            self.cfg.db_paths.append(path)
        self.cfg.last_db_path = path
        self.cfg.save()
        self.refresh()

    def on_attach(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Подключить базу", "", "SQLite (*.sqlite *.db);;All files (*.*)")
        if not path:
            return
        try:
            db = DB(path)
            _ = db.get_department_name()
            db.close()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть базу:\n{e}")
            return

        if path not in self.cfg.db_paths:
            self.cfg.db_paths.append(path)
        self.cfg.last_db_path = path
        self.cfg.save()
        self.refresh()

    def on_import_copy(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Импортировать базу (копировать)", "", "SQLite (*.sqlite *.db);;All files (*.*)"
        )
        if not path:
            return
        try:
            db = DB(path)
            name = db.get_department_name()
            db.close()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть базу:\n{e}")
            return

        dest = app_db_store_dir() / f"{name}_{uuid.uuid4().hex[:8]}.sqlite"
        shutil.copy2(path, dest)
        dest_path = str(dest)

        if dest_path not in self.cfg.db_paths:
            self.cfg.db_paths.append(dest_path)
        self.cfg.last_db_path = dest_path
        self.cfg.save()
        self.refresh()

    def on_export_copy(self) -> None:
        src = self.current_path()
        if not src:
            QMessageBox.information(self, "Экспорт", "Выберите базу из списка.")
            return
        name = Path(src).stem
        try:
            db = DB(src)
            name = db.get_department_name()
            db.close()
        except Exception:
            pass

        out, _ = QFileDialog.getSaveFileName(self, "Экспорт копии базы", f"{name}.sqlite", "SQLite (*.sqlite)")
        if not out:
            return
        if not out.endswith(".sqlite"):
            out += ".sqlite"
        shutil.copy2(src, out)
        QMessageBox.information(self, "Экспорт", "Копия базы сохранена.")

    def on_remove(self) -> None:
        p = self.current_path()
        if not p:
            return
        self.cfg.db_paths = [x for x in self.cfg.db_paths if x != p]
        if self.cfg.last_db_path == p:
            self.cfg.last_db_path = None
        self.cfg.save()
        self.refresh()

    def on_open(self) -> None:
        p = self.current_path()
        if not p:
            QMessageBox.information(self, "Выбор базы", "Выберите базу из списка.")
            return
        self.selected_path = p
        self.cfg.last_db_path = p
        self.cfg.save()
        self.accept()


class DepartmentSettingsDialog(QDialog):
    def __init__(self, db: DB, parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Настройки отдела")
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        s = db.get_department_settings()
        self.ed_name = QLineEdit(s["department_name"])
        form.addRow("Название отдела:", self.ed_name)

        form.addRow(QLabel("<b>Ответственный за ведение табеля</b>"))
        self.ed_pos = QLineEdit(s["resp_position"] or "")
        self.ed_rank = QLineEdit(s["resp_rank"] or "")
        self.ed_last = QLineEdit(s["resp_last_name"] or "")
        self.ed_first = QLineEdit(s["resp_first_name"] or "")
        self.ed_mid = QLineEdit(s["resp_middle_name"] or "")
        form.addRow("Должность:", self.ed_pos)
        form.addRow("Звание:", self.ed_rank)
        form.addRow("Фамилия:", self.ed_last)
        form.addRow("Имя:", self.ed_first)
        form.addRow("Отчество:", self.ed_mid)

        bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        lay.addWidget(bb)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)

    def accept(self) -> None:
        self.db.conn.execute("BEGIN;")
        try:
            self.db.update_department_settings(
                department_name=self.ed_name.text().strip() or "Отдел",
                resp_position=self.ed_pos.text().strip() or None,
                resp_rank=self.ed_rank.text().strip() or None,
                resp_last_name=self.ed_last.text().strip() or None,
                resp_first_name=self.ed_first.text().strip() or None,
                resp_middle_name=self.ed_mid.text().strip() or None,
            )
            self.db.conn.execute("COMMIT;")
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))
            return
        super().accept()


class EmployeeDialog(QDialog):
    def __init__(self, title: str, default_year: int, default_month: int, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle(title)
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.ed_last = QLineEdit()
        self.ed_first = QLineEdit()
        self.ed_mid = QLineEdit()
        self.ed_rank = QLineEdit()
        self.ed_pos = QLineEdit()

        self.cb_year = QComboBox()
        self.cb_month = QComboBox()
        for y in range(2000, 2100):
            self.cb_year.addItem(str(y), y)
        for m in range(1, 13):
            self.cb_month.addItem(f"{m:02d}", m)

        yi = self.cb_year.findData(default_year)
        if yi >= 0:
            self.cb_year.setCurrentIndex(yi)
        self.cb_month.setCurrentIndex(max(0, default_month - 1))

        form.addRow("Фамилия:", self.ed_last)
        form.addRow("Имя:", self.ed_first)
        form.addRow("Отчество:", self.ed_mid)
        form.addRow("Звание:", self.ed_rank)
        form.addRow("Должность:", self.ed_pos)

        row = QHBoxLayout()
        row.addWidget(self.cb_year)
        row.addWidget(self.cb_month)
        w = QWidget()
        w.setLayout(row)
        form.addRow("Месяц приема (ГГГГ-ММ):", w)

        self.opening_time = DurationPicker(max_hours=999)
        self.opening_time.set_minutes_total(0)

        self.opening_days = QSpinBox()
        self.opening_days.setRange(0, 366)
        configure_spinbox(self.opening_days)
        self.opening_days.setValue(0)

        form.addRow("Начальные часы (перенос):", self.opening_time)
        form.addRow("Начальные дни (перенос):", self.opening_days)

        self.bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        lay.addWidget(self.bb)
        self.bb.accepted.connect(self.accept)
        self.bb.rejected.connect(self.reject)

    def set_from_employee(self, emp: sqlite3.Row) -> None:
        self.ed_last.setText(emp["last_name"])
        self.ed_first.setText(emp["first_name"])
        self.ed_mid.setText(emp["middle_name"] or "")
        self.ed_rank.setText(emp["rank"] or "")
        self.ed_pos.setText(emp["position"] or "")

        y, m = parse_month(emp["start_month"])
        yi = self.cb_year.findData(y)
        if yi >= 0:
            self.cb_year.setCurrentIndex(yi)
        self.cb_month.setCurrentIndex(m - 1)

        self.opening_time.set_minutes_total(int(emp["opening_minutes"] or 0))
        self.opening_days.setValue(int(emp["opening_days"] or 0))

    def get_values(self) -> dict[str, Any]:
        y = int(self.cb_year.currentData())
        m = int(self.cb_month.currentData())
        start_month = f"{y:04d}-{m:02d}"
        return {
            "last_name": self.ed_last.text().strip(),
            "first_name": self.ed_first.text().strip(),
            "middle_name": self.ed_mid.text().strip(),
            "rank": self.ed_rank.text().strip(),
            "position": self.ed_pos.text().strip(),
            "start_month": start_month,
            "opening_minutes": self.opening_time.minutes_total(),
            "opening_days": int(self.opening_days.value()),
        }

    def accept(self) -> None:
        v = self.get_values()
        if not v["last_name"] or not v["first_name"]:
            QMessageBox.warning(self, "Проверка", "Фамилия и имя обязательны.")
            return
        super().accept()


class EndDateDialog(QDialog):
    def __init__(self, title: str, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle(title)
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.de = QDateEdit()
        self.de.setCalendarPopup(True)
        self.de.setDate(QDate.currentDate())
        configure_spinbox(self.de)

        form.addRow("Дата:", self.de)
        self.bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lay.addWidget(self.bb)
        self.bb.accepted.connect(self.accept)
        self.bb.rejected.connect(self.reject)

    def get_date(self) -> date:
        return qdate_to_date(self.de.date())

class BreakIntervalDialog(QDialog):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle("Перерыв")
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.p_start = DateTimePicker()
        self.p_end = DateTimePicker()

        now = datetime.now().replace(minute=0, second=0, microsecond=0)
        self.p_start.set_datetime(now)
        self.p_end.set_datetime(now + timedelta(hours=1))

        form.addRow("Начало перерыва:", self.p_start)
        form.addRow("Конец перерыва:", self.p_end)

        bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        lay.addWidget(bb)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)

    def set_values(self, start: datetime, end: datetime) -> None:
        self.p_start.set_datetime(start)
        self.p_end.set_datetime(end)

    def values(self) -> tuple[datetime, datetime]:
        return self.p_start.get_datetime(), self.p_end.get_datetime()

    def accept(self) -> None:
        s, e = self.values()
        if s >= e:
            QMessageBox.warning(self, "Проверка", "Конец перерыва должен быть позже начала.")
            return
        super().accept()

class DutyDialog(QDialog):
    def __init__(self, db: DB, employee_id: int, exclude_duty_id: Optional[int] = None, parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.employee_id = employee_id
        self.exclude_duty_id = exclude_duty_id

        self.setWindowTitle("Дежурство")
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.p_start = DateTimePicker()
        self.p_end = DateTimePicker()
        now = datetime.now().replace(second=0, microsecond=0)
        self.p_start.set_datetime(now)
        self.p_end.set_datetime(now + timedelta(hours=1))

        self.ed_comment = QLineEdit()

        form.addRow("Начало:", self.p_start)
        form.addRow("Конец:", self.p_end)
        form.addRow("Комментарий:", self.ed_comment)
        self.chk_breaks = QCheckBox("Перерывы")
        lay.addWidget(self.chk_breaks)

        self.breaks_panel = QWidget()
        vb = QVBoxLayout(self.breaks_panel)
        vb.setContentsMargins(0, 0, 0, 0)

        self.list_breaks = QListWidget()
        vb.addWidget(self.list_breaks, 1)

        hb = QHBoxLayout()
        self.btn_break_add = QPushButton("Добавить перерыв…")
        self.btn_break_edit = QPushButton("Редактировать")
        self.btn_break_del = QPushButton("Удалить")
        hb.addWidget(self.btn_break_add)
        hb.addWidget(self.btn_break_edit)
        hb.addWidget(self.btn_break_del)
        hb.addStretch(1)
        vb.addLayout(hb)

        lay.addWidget(self.breaks_panel)
        self.breaks_panel.setVisible(False)

        self._breaks: list[tuple[datetime, datetime]] = []        

        self.lbl_warn = QLabel("")
        self.lbl_warn.setStyleSheet("color:#b35a00;")
        lay.addWidget(self.lbl_warn)

        bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        lay.addWidget(bb)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)

        for w in [self.p_start.de, self.p_end.de]:
            w.dateChanged.connect(lambda _=None: self.update_warning())
        for w in [self.p_start.tp.sp_h, self.p_start.tp.sp_m, self.p_end.tp.sp_h, self.p_end.tp.sp_m]:
            w.valueChanged.connect(lambda _=None: self.update_warning())

        self.update_warning()
        self.chk_breaks.toggled.connect(self._on_breaks_toggled)
        self.btn_break_add.clicked.connect(self._add_break)
        self.btn_break_edit.clicked.connect(self._edit_break)
        self.btn_break_del.clicked.connect(self._delete_break)
        self.list_breaks.itemDoubleClicked.connect(lambda _=None: self._edit_break())
        
    def set_values(self, start: datetime, end: datetime, comment: str, breaks: Optional[list[tuple[datetime, datetime]]] = None) -> None:
        self.p_start.set_datetime(start)
        self.p_end.set_datetime(end)
        self.ed_comment.setText(comment or "")
        self.set_breaks(breaks or [])
        self.update_warning()

    def get_values(self) -> tuple[datetime, datetime, str]:
        return self.p_start.get_datetime(), self.p_end.get_datetime(), self.ed_comment.text().strip()

    def set_breaks(self, breaks: list[tuple[datetime, datetime]]) -> None:
        self._breaks = sorted(breaks or [], key=lambda x: x[0])
        self.chk_breaks.setChecked(bool(self._breaks))
        self._rebuild_breaks_list()

    def get_breaks(self) -> list[tuple[datetime, datetime]]:
        return list(self._breaks) if self.chk_breaks.isChecked() else []

    def _on_breaks_toggled(self, v: bool) -> None:
        self.breaks_panel.setVisible(v)
        if not v:
            self._breaks = []
            self._rebuild_breaks_list()

    def _rebuild_breaks_list(self) -> None:
        self.list_breaks.clear()
        for s, e in sorted(self._breaks, key=lambda x: x[0]):
            txt = f"{s.strftime('%d.%m.%Y %H:%M')} – {e.strftime('%d.%m.%Y %H:%M')}"
            it = QListWidgetItem(txt)
            it.setData(Qt.UserRole, (dt_iso(s), dt_iso(e)))
            self.list_breaks.addItem(it)

    def _selected_break_index(self) -> Optional[int]:
        it = self.list_breaks.currentItem()
        if not it:
            return None
        iso_pair = it.data(Qt.UserRole)
        if not iso_pair:
            return None
        try:
            s_iso, e_iso = iso_pair
            s0 = dt_parse(s_iso)
            e0 = dt_parse(e_iso)
        except Exception:
            return None

        for i, (s, e) in enumerate(self._breaks):
            if s == s0 and e == e0:
                return i
        return None

    def _add_break(self) -> None:
        s_duty, e_duty, _ = self.get_values()

        dlg = BreakIntervalDialog(self)
        # дефолт: середина дежурства на 30 минут
        mid = s_duty + (e_duty - s_duty) / 2 if e_duty > s_duty else s_duty
        mid = mid.replace(minute=0, second=0, microsecond=0)
        dlg.set_values(mid, mid + timedelta(hours=1))

        if dlg.exec() != QDialog.Accepted:
            return

        s, e = dlg.values()
        self._breaks.append((s, e))
        self._breaks = sorted(self._breaks, key=lambda x: x[0])
        self.chk_breaks.setChecked(True)
        self._rebuild_breaks_list()

    def _edit_break(self) -> None:
        idx = self._selected_break_index()
        if idx is None:
            return

        s0, e0 = self._breaks[idx]
        dlg = BreakIntervalDialog(self)
        dlg.set_values(s0, e0)
        if dlg.exec() != QDialog.Accepted:
            return

        s, e = dlg.values()
        self._breaks[idx] = (s, e)
        self._breaks = sorted(self._breaks, key=lambda x: x[0])
        self._rebuild_breaks_list()

    def _delete_break(self) -> None:
        idx = self._selected_break_index()
        if idx is None:
            return
        del self._breaks[idx]
        self._rebuild_breaks_list()
        if not self._breaks:
            self.chk_breaks.setChecked(False)

    def _validate_breaks(self, duty_start: datetime, duty_end: datetime) -> Optional[str]:
        br = self.get_breaks()
        if not br:
            return None

        # внутри дежурства + порядок
        for s, e in br:
            if s >= e:
                return "Есть перерыв с неверным интервалом (конец раньше/равен началу)."
            if s < duty_start or e > duty_end:
                return "Перерывы должны полностью находиться внутри интервала дежурства."

        # не пересекаться между собой
        merged = merge_intervals(sorted(br, key=lambda x: x[0]))
        if len(merged) != len(br):
            return "Перерывы не должны пересекаться между собой."

        return None

    def update_warning(self) -> None:
        s, e, _ = self.get_values()
        if s >= e:
            self.lbl_warn.setText("")
            return
        overlaps = self.db.find_overlapping_duties(self.employee_id, s, e, exclude_duty_id=self.exclude_duty_id)
        if not overlaps:
            self.lbl_warn.setText("")
            return
        self.lbl_warn.setText("Есть пересечение с существующим дежурством. Сохранение будет запрещено.")

    def accept(self) -> None:
        s, e, _ = self.get_values()
        if s >= e:
            QMessageBox.warning(self, "Проверка", "Конец должен быть позже начала.")
            return
        err = self._validate_breaks(s, e)
        if err:
            QMessageBox.warning(self, "Проверка", err)
            return            

        overlaps = self.db.find_overlapping_duties(self.employee_id, s, e, exclude_duty_id=self.exclude_duty_id)
        if overlaps:
            QMessageBox.warning(self, "Проверка", "Нельзя сохранить дежурство: есть пересечение с существующим дежурством.")
            return

        super().accept()


class FillPeriodDialog(QDialog):
    def __init__(self, parent: QWidget | None = None, default_only_working: bool = True):
        super().__init__(parent)
        self.setWindowTitle("Заполнить периодом")
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.d1 = QDateEdit()
        self.d1.setCalendarPopup(True)
        self.d1.setDate(QDate.currentDate())
        configure_spinbox(self.d1)

        self.d2 = QDateEdit()
        self.d2.setCalendarPopup(True)
        self.d2.setDate(QDate.currentDate())
        configure_spinbox(self.d2)

        self.rb_only_working = QRadioButton("Только рабочие дни")
        self.rb_only_working.setChecked(default_only_working)

        form.addRow("Начало:", self.d1)
        form.addRow("Конец:", self.d2)
        lay.addWidget(self.rb_only_working)

        self.lbl_warn = QLabel("Внимание: выходные/праздничные дни обычно нельзя указывать как компенсационный выходной.")
        self.lbl_warn.setStyleSheet("color:#b00020;")
        lay.addWidget(self.lbl_warn)
        self.lbl_warn.setVisible(not self.rb_only_working.isChecked())
        self.rb_only_working.toggled.connect(lambda v: self.lbl_warn.setVisible(not v))

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lay.addWidget(bb)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)

    def values(self) -> tuple[date, date, bool]:
        return qdate_to_date(self.d1.date()), qdate_to_date(self.d2.date()), self.rb_only_working.isChecked()


class FillCountDialog(QDialog):
    def __init__(self, parent: QWidget | None = None, default_only_working: bool = True):
        super().__init__(parent)
        self.setWindowTitle("Заполнить по количеству")
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.d1 = QDateEdit()
        self.d1.setCalendarPopup(True)
        self.d1.setDate(QDate.currentDate())
        configure_spinbox(self.d1)

        self.sp = QSpinBox()
        self.sp.setRange(1, 366)
        self.sp.setValue(1)
        configure_spinbox(self.sp)

        self.rb_only_working = QRadioButton("Только рабочие дни")
        self.rb_only_working.setChecked(default_only_working)

        form.addRow("Старт:", self.d1)
        form.addRow("Количество N:", self.sp)
        lay.addWidget(self.rb_only_working)

        self.lbl_warn = QLabel("Внимание: выходные/праздничные дни обычно нельзя указывать как компенсационный выходной.")
        self.lbl_warn.setStyleSheet("color:#b00020;")
        lay.addWidget(self.lbl_warn)
        self.lbl_warn.setVisible(not self.rb_only_working.isChecked())
        self.rb_only_working.toggled.connect(lambda v: self.lbl_warn.setVisible(not v))

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lay.addWidget(bb)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)

    def values(self) -> tuple[date, int, bool]:
        return qdate_to_date(self.d1.date()), int(self.sp.value()), self.rb_only_working.isChecked()


class MultiDateCalendarDialog(QDialog):
    """Клик по дате = выделить/снять. Можно выбирать даты разных месяцев."""

    def __init__(self, db: DB, year: int, parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.year = year
        self.selected: set[date] = set()

        self.setWindowTitle("Выбор дат выходных")
        lay = QVBoxLayout(self)

        top = QHBoxLayout()
        self.chk_only_working = QCheckBox("Только рабочие дни")
        self.chk_only_working.setChecked(True)
        top.addWidget(self.chk_only_working)
        top.addStretch(1)
        self.lbl_counter = QLabel("Выбрано: 0")
        top.addWidget(self.lbl_counter)
        lay.addLayout(top)

        self.lbl_warn = QLabel("Внимание: выходные/праздничные дни обычно нельзя указывать как компенсационный выходной.")
        self.lbl_warn.setStyleSheet("color:#b00020;")
        lay.addWidget(self.lbl_warn)
        self.lbl_warn.setVisible(False)

        self.cal = QCalendarWidget()
        self.cal.setMinimumDate(QDate(year, 1, 1))
        self.cal.setMaximumDate(QDate(year, 12, 31))
        lay.addWidget(self.cal)

        btns = QHBoxLayout()
        self.btn_clear = QPushButton("Очистить")
        btns.addWidget(self.btn_clear)
        btns.addStretch(1)
        self.bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.addWidget(self.bb)
        lay.addLayout(btns)

        self.bb.accepted.connect(self.accept)
        self.bb.rejected.connect(self.reject)
        self.btn_clear.clicked.connect(self.clear_all)
        self.cal.clicked.connect(self.on_date_clicked)
        self.chk_only_working.toggled.connect(self.on_only_working_toggled)

        self.fmt_selected = QTextCharFormat()
        self.fmt_selected.setBackground(QBrush(QColor("#9ecbff")))

    def on_only_working_toggled(self, v: bool) -> None:
        self.lbl_warn.setVisible(not v)

    def refresh_counter(self) -> None:
        self.lbl_counter.setText(f"Выбрано: {len(self.selected)}")

    def set_selected_dates(self, dates: list[date]) -> None:
        self.selected = set(dates)
        self.repaint_formats()
        self.refresh_counter()

    def repaint_formats(self) -> None:
        d0 = date(self.year, 1, 1)
        d1 = date(self.year, 12, 31)
        cur = d0
        while cur <= d1:
            qd = QDate(cur.year, cur.month, cur.day)
            self.cal.setDateTextFormat(qd, self.fmt_selected if cur in self.selected else QTextCharFormat())
            cur += timedelta(days=1)

    def clear_all(self) -> None:
        self.selected.clear()
        self.repaint_formats()
        self.refresh_counter()

    def on_date_clicked(self, qd: QDate) -> None:
        d0 = qdate_to_date(qd)

        if self.chk_only_working.isChecked():
            try:
                if not self.db.is_working_day(d0):
                    QMessageBox.information(
                        self,
                        "Выбор дат",
                        "Это нерабочий день. Снимите галочку 'Только рабочие дни', если хотите выбрать его.",
                    )
                    return
            except CalendarMissingError as e:
                QMessageBox.warning(self, "Календарь", str(e))
                return

        if d0 in self.selected:
            self.selected.remove(d0)
            self.cal.setDateTextFormat(qd, QTextCharFormat())
        else:
            self.selected.add(d0)
            self.cal.setDateTextFormat(qd, self.fmt_selected)

        self.refresh_counter()

    def get_dates(self) -> list[date]:
        return sorted(self.selected)


# -----------------------------
# Dialogs: Compensation
# -----------------------------
class CompensationDialog(QDialog):
    def __init__(
        self,
        db: DB,
        employee: sqlite3.Row,
        year: int,
        workday_minutes: int,
        comp_id: Optional[int] = None,
        parent: QWidget | None = None,
    ):
        super().__init__(parent)
        self.db = db
        self.employee = employee
        self.year = year
        self.workday_minutes = int(workday_minutes)
        self.comp_id = comp_id
        self.edit_mode = comp_id is not None

        self.setWindowTitle("Компенсация" + (" (редактирование)" if self.edit_mode else ""))
        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.cb_unit = QComboBox()
        self.cb_unit.addItem("Часы", "hours")
        self.cb_unit.addItem("Дни", "days")
        form.addRow("Списать:", self.cb_unit)

        self.ed_comment = QLineEdit()
        form.addRow("Комментарий:", self.ed_comment)

        # -----------------
        # HOURS (day_off) — много дат + минут на каждую дату
        # -----------------
        self.grp_hours = QGroupBox("Выходной (часы)")
        vh = QVBoxLayout(self.grp_hours)

        self.lbl_hours_hint = QLabel("")
        self.lbl_hours_hint.setStyleSheet("color:#808080;")
        vh.addWidget(self.lbl_hours_hint)

        fh = QFormLayout()
        self.hours_amount = DurationPicker(max_hours=24)
        fh.addRow("Количество на одну дату (чч:мм):", self.hours_amount)
        vh.addLayout(fh)

        self.lbl_hours_count = QLabel("Выбрано дат: 0")
        vh.addWidget(self.lbl_hours_count)

        self.hours_dates = QListWidget()
        vh.addWidget(self.hours_dates, 1)

        bh = QHBoxLayout()
        self.btn_hours_fill_period = QPushButton("Заполнить периодом")
        self.btn_hours_fill_count = QPushButton("Заполнить по количеству")
        self.btn_hours_pick_calendar = QPushButton("Выбрать в календаре…")
        self.btn_hours_clear = QPushButton("Очистить")
        bh.addWidget(self.btn_hours_fill_period)
        bh.addWidget(self.btn_hours_fill_count)
        bh.addWidget(self.btn_hours_pick_calendar)
        bh.addWidget(self.btn_hours_clear)
        bh.addStretch(1)
        vh.addLayout(bh)

        # -----------------
        # DAYS (day_off) — как было
        # -----------------
        self.grp_days = QGroupBox("Выходной (дни)")
        vd = QVBoxLayout(self.grp_days)

        self.lbl_days_count = QLabel("Количество дней: 0")
        vd.addWidget(self.lbl_days_count)

        self.days_dates = QListWidget()
        vd.addWidget(self.days_dates, 1)

        bd = QHBoxLayout()
        self.btn_days_fill_period = QPushButton("Заполнить периодом")
        self.btn_days_fill_count = QPushButton("Заполнить по количеству")
        self.btn_days_pick_calendar = QPushButton("Выбрать в календаре…")
        self.btn_days_clear = QPushButton("Очистить")
        bd.addWidget(self.btn_days_fill_period)
        bd.addWidget(self.btn_days_fill_count)
        bd.addWidget(self.btn_days_pick_calendar)
        bd.addWidget(self.btn_days_clear)
        bd.addStretch(1)
        vd.addLayout(bd)

        lay.addWidget(self.grp_hours, 1)
        lay.addWidget(self.grp_days, 1)

        self.bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        lay.addWidget(self.bb)
        self.bb.accepted.connect(self.accept)
        self.bb.rejected.connect(self.reject)

        self.cb_unit.currentIndexChanged.connect(self.update_visibility)

        self.btn_hours_fill_period.clicked.connect(lambda: self._fill_period_into(self.hours_dates))
        self.btn_hours_fill_count.clicked.connect(lambda: self._fill_count_into(self.hours_dates))
        self.btn_hours_pick_calendar.clicked.connect(lambda: self._pick_calendar_into(self.hours_dates))
        self.btn_hours_clear.clicked.connect(lambda: self._clear_list(self.hours_dates))

        self.btn_days_fill_period.clicked.connect(lambda: self._fill_period_into(self.days_dates))
        self.btn_days_fill_count.clicked.connect(lambda: self._fill_count_into(self.days_dates))
        self.btn_days_pick_calendar.clicked.connect(lambda: self._pick_calendar_into(self.days_dates))
        self.btn_days_clear.clicked.connect(lambda: self._clear_list(self.days_dates))

        # дефолт: по длине рабочего дня (но не больше 24ч)
        self._update_hours_hint()
        self.hours_amount.set_minutes_total(min(max(1, self.workday_minutes), 24 * 60))

        if self.edit_mode:
            self.load_existing()

        self.update_visibility()

        if self.edit_mode:
            self.cb_unit.setEnabled(False)

    def prefill_date(self, d0: date) -> None:
        # предзаполняем обе вкладки, чтобы при переключении "Часы/Дни" дата уже была
        if d0.year != self.year:
            return
        self._set_dates_to_list(self.hours_dates, [d0])
        self._set_dates_to_list(self.days_dates, [d0])
        self._update_counts()

    def _update_hours_hint(self) -> None:
        self.lbl_hours_hint.setText(f"Максимум на одну дату: {fmt_minutes_ru_words(self.workday_minutes)}")

    def _clear_list(self, lw: QListWidget) -> None:
        lw.clear()
        self._update_counts()

    def _update_counts(self) -> None:
        self.lbl_hours_count.setText(f"Выбрано дат: {self.hours_dates.count()}")
        self.lbl_days_count.setText(f"Количество дней: {self.days_dates.count()}")

    def _dates_from_list(self, lw: QListWidget) -> list[date]:
        out: list[date] = []
        for i in range(lw.count()):
            try:
                out.append(d_parse(lw.item(i).text()))
            except Exception:
                pass
        # уберем дубликаты
        out2 = sorted(set(out))
        return out2

    def _set_dates_to_list(self, lw: QListWidget, dates: list[date]) -> None:
        lw.clear()
        for d0 in sorted(set(dates)):
            lw.addItem(d_iso(d0))
        self._update_counts()

    def load_existing(self) -> None:
        c = self.db.get_compensation(int(self.comp_id))
        unit = c["unit"]
        method = c["method"]

        if method == "money":
            raise DBError("Денежные компенсации редактируются в отдельном окне.")

        self.ed_comment.setText(c["comment"] or "")

        if unit == "hours":
            self.cb_unit.setCurrentIndex(0)
            if c["event_date"]:
                d0 = d_parse(c["event_date"])
                self._set_dates_to_list(self.hours_dates, [d0])
            self.hours_amount.set_minutes_total(int(c["amount_minutes"] or 0))
        else:
            self.cb_unit.setCurrentIndex(1)
            dates = [d_parse(x) for x in self.db.get_comp_dates(int(c["id"]))]
            self._set_dates_to_list(self.days_dates, dates)

    def update_visibility(self) -> None:
        unit = self.cb_unit.currentData()
        self.grp_hours.setVisible(unit == "hours")
        self.grp_days.setVisible(unit == "days")
        self._update_counts()

    def _fill_period_into(self, lw: QListWidget) -> None:
        if not self.db.calendar_year_complete(self.year):
            QMessageBox.warning(self, "Календарь", "Календарь года не создан/неполный. Создайте календарь.")
            return

        dlg = FillPeriodDialog(self, default_only_working=True)
        if dlg.exec() != QDialog.Accepted:
            return

        d1, d2, only_working = dlg.values()
        if d1.year != self.year or d2.year != self.year:
            QMessageBox.warning(self, "Проверка", f"Выберите даты в пределах {self.year} года.")
            return
        if d2 < d1:
            QMessageBox.warning(self, "Проверка", "Конец периода раньше начала.")
            return

        dates: list[date] = []
        cur = d1
        while cur <= d2:
            if only_working:
                if self.db.is_working_day(cur):
                    dates.append(cur)
            else:
                dates.append(cur)
            cur += timedelta(days=1)

        self._set_dates_to_list(lw, dates)

    def _fill_count_into(self, lw: QListWidget) -> None:
        if not self.db.calendar_year_complete(self.year):
            QMessageBox.warning(self, "Календарь", "Календарь года не создан/неполный. Создайте календарь.")
            return

        dlg = FillCountDialog(self, default_only_working=True)
        if dlg.exec() != QDialog.Accepted:
            return

        start, n, only_working = dlg.values()
        if start.year != self.year:
            QMessageBox.warning(self, "Проверка", f"Выберите стартовую дату в пределах {self.year} года.")
            return

        dates: list[date] = []
        cur = start
        while len(dates) < n:
            if cur.year != self.year:
                break
            if only_working:
                if self.db.is_working_day(cur):
                    dates.append(cur)
            else:
                dates.append(cur)
            cur += timedelta(days=1)

        self._set_dates_to_list(lw, dates)

    def _pick_calendar_into(self, lw: QListWidget) -> None:
        if not self.db.calendar_year_complete(self.year):
            QMessageBox.warning(self, "Календарь", "Календарь года не создан/неполный. Создайте календарь.")
            return

        current = self._dates_from_list(lw)

        dlg = MultiDateCalendarDialog(self.db, self.year, parent=self)
        dlg.set_selected_dates(current)
        if dlg.exec() != QDialog.Accepted:
            return

        self._set_dates_to_list(lw, dlg.get_dates())

    def _sum_hours_dayoff_on_date(self, employee_id: int, d0: date, exclude_comp_id: Optional[int]) -> int:
        if exclude_comp_id is None:
            r = self.db.conn.execute(
                """
                SELECT COALESCE(SUM(amount_minutes),0) AS m
                FROM compensation
                WHERE employee_id=?
                  AND unit='hours' AND method='day_off'
                  AND event_date=?
                """,
                (employee_id, d_iso(d0)),
            ).fetchone()
        else:
            r = self.db.conn.execute(
                """
                SELECT COALESCE(SUM(amount_minutes),0) AS m
                FROM compensation
                WHERE employee_id=?
                  AND unit='hours' AND method='day_off'
                  AND event_date=?
                  AND id <> ?
                """,
                (employee_id, d_iso(d0), int(exclude_comp_id)),
            ).fetchone()
        return int(r["m"])

    def _validate_employee_all_years(self, employee_id: int) -> tuple[bool, str]:
        years = self.db.employee_data_years(employee_id)
        if not years:
            years = {self.year}
        for y in sorted(years):
            ok, msg = validate_non_negative_over_year(self.db, employee_id, y)
            if not ok:
                return False, msg
        return True, ""

    def accept(self) -> None:
        unit = self.cb_unit.currentData()
        comment = self.ed_comment.text().strip()
        eid = int(self.employee["id"])

        if unit == "hours":
            minutes = int(self.hours_amount.minutes_total())
            if minutes <= 0:
                QMessageBox.warning(self, "Проверка", "Количество часов должно быть больше 0.")
                return
            if minutes > self.workday_minutes:
                QMessageBox.warning(
                    self,
                    "Проверка",
                    f"Нельзя списать {fmt_minutes_ru_words(minutes)} на одну дату.\n"
                    f"Максимум: {fmt_minutes_ru_words(self.workday_minutes)}.",
                )
                return

            dates = self._dates_from_list(self.hours_dates)
            if not dates:
                QMessageBox.warning(self, "Проверка", "Выберите хотя бы одну дату.")
                return

            # проверка периода сотрудника + лимита на дату
            exclude_id = int(self.comp_id) if self.edit_mode else None
            for d0 in dates:
                msg = validate_event_in_employee_period(self.employee, d0)
                if msg:
                    QMessageBox.warning(self, "Проверка", f"Дата {d_iso(d0)}: {msg}")
                    return

                already = self._sum_hours_dayoff_on_date(eid, d0, exclude_id)
                if already + minutes > self.workday_minutes:
                    rest = max(0, self.workday_minutes - already)
                    QMessageBox.warning(
                        self,
                        "Проверка",
                        f"Дата {fmt_date_iso(d_iso(d0))}: уже списано {fmt_minutes_ru_words(already)}.\n"
                        f"Можно добавить не больше {fmt_minutes_ru_words(rest)}.",
                    )
                    return

            self.db.conn.execute("BEGIN;")
            try:
                if self.edit_mode:
                    # проще и логичнее: заменить редактируемую запись на набор записей по датам
                    self.db.delete_compensation(int(self.comp_id))

                for d0 in dates:
                    self.db.add_compensation_hours_dayoff(eid, d0, minutes, comment)

                ok, msg2 = self._validate_employee_all_years(eid)
                if not ok:
                    self.db.conn.execute("ROLLBACK;")
                    QMessageBox.warning(self, "Недостаточно", msg2)
                    return

                self.db.conn.execute("COMMIT;")
                super().accept()
            except Exception as e:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.critical(self, "Ошибка", str(e))
            return

        # unit == "days"
        if not self.db.calendar_year_complete(self.year):
            QMessageBox.warning(self, "Календарь", "Календарь года не создан/неполный. Нельзя сохранять выходные днями.")
            return

        dates = self._dates_from_list(self.days_dates)
        if not dates:
            QMessageBox.warning(self, "Проверка", "Выберите хотя бы одну дату выходного.")
            return

        for d0 in dates:
            msg = validate_event_in_employee_period(self.employee, d0)
            if msg:
                QMessageBox.warning(self, "Проверка", f"Дата {d_iso(d0)}: {msg}")
                return

        self.db.conn.execute("BEGIN;")
        try:
            if self.edit_mode:
                self.db.update_compensation(int(self.comp_id), unit="days", method="day_off", comment=comment or None)
                self.db.replace_comp_dayoff_dates(int(self.comp_id), eid, dates)
            else:
                self.db.add_compensation_days_dayoff(eid, dates, comment)

            ok, msg2 = self._validate_employee_all_years(eid)
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return

            self.db.conn.execute("COMMIT;")
            super().accept()
        except sqlite3.IntegrityError:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.warning(self, "Проверка", "Нельзя поставить два выходных дня на одну и ту же дату.")
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

class MoneyCompensationDialog(QDialog):
    """Компенсация деньгами (приказ). Отдельно от обычных 'выходных' компенсаций."""

    def __init__(self, db: DB, employee: sqlite3.Row, default_year: int, comp_id: Optional[int] = None, parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.employee = employee
        self.default_year = default_year
        self.comp_id = comp_id
        self.edit_mode = comp_id is not None

        self.setWindowTitle("Компенсация деньгами (приказ)" + (" — редактирование" if self.edit_mode else ""))

        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.cb_unit = QComboBox()
        self.cb_unit.addItem("Часы", "hours")
        self.cb_unit.addItem("Дни", "days")
        form.addRow("Тип:", self.cb_unit)

        self.ed_order_no = QLineEdit()
        form.addRow("№ приказа:", self.ed_order_no)

        self.de_order_date = QDateEdit()
        self.de_order_date.setCalendarPopup(True)
        self.de_order_date.setDate(QDate.currentDate())
        configure_spinbox(self.de_order_date)
        form.addRow("Дата приказа:", self.de_order_date)

        self.sp_days = QSpinBox()
        self.sp_days.setRange(1, 366)
        self.sp_days.setValue(1)
        configure_spinbox(self.sp_days)

        self.amount_hours = DurationPicker(max_hours=999)
        self.amount_hours.set_minutes_total(60)

        self.lbl_amount = QLabel("")
        form.addRow("Количество:", self.lbl_amount)
        form.addRow("Дни:", self.sp_days)
        form.addRow("Часы (чч:мм):", self.amount_hours)

        self.ed_comment = QLineEdit()
        form.addRow("Комментарий:", self.ed_comment)

        self.bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        lay.addWidget(self.bb)
        self.bb.accepted.connect(self.accept)
        self.bb.rejected.connect(self.reject)

        self.cb_unit.currentIndexChanged.connect(self._update_visibility)

        if self.edit_mode:
            self._load_existing()

        self._update_visibility()

        if self.edit_mode:
            # чтобы не путать конвертациями "дни<->часы" при редактировании
            self.cb_unit.setEnabled(False)

    def _update_visibility(self) -> None:
        unit = self.cb_unit.currentData()
        if unit == "days":
            self.lbl_amount.setText("Дни")
            self.sp_days.setVisible(True)
            self.amount_hours.setVisible(False)
        else:
            self.lbl_amount.setText("Часы")
            self.sp_days.setVisible(False)
            self.amount_hours.setVisible(True)

    def _load_existing(self) -> None:
        c = self.db.get_compensation(int(self.comp_id))
        if c["method"] != "money":
            raise DBError("Это не денежная компенсация (method != money).")

        unit = c["unit"]
        self.cb_unit.setCurrentIndex(0 if unit == "hours" else 1)
        self.ed_order_no.setText(c["order_no"] or "")
        if c["order_date"]:
            d0 = d_parse(c["order_date"])
            self.de_order_date.setDate(QDate(d0.year, d0.month, d0.day))
        self.ed_comment.setText(c["comment"] or "")

        if unit == "hours":
            self.amount_hours.set_minutes_total(int(c["amount_minutes"] or 0))
        else:
            self.sp_days.setValue(int(c["amount_days"] or 1))

    def accept(self) -> None:
        unit = self.cb_unit.currentData()
        order_no = self.ed_order_no.text().strip()
        if not order_no:
            QMessageBox.warning(self, "Проверка", "Укажите № приказа.")
            return

        od = qdate_to_date(self.de_order_date.date())
        msg = validate_event_in_employee_period(self.employee, od)
        if msg:
            QMessageBox.warning(self, "Проверка", msg)
            return

        # если списываем ДНИ — нужен календарь года (иначе баланс дней не проверить)
        if unit == "days" and not self.db.calendar_year_complete(od.year):
            QMessageBox.warning(self, "Календарь", f"Календарь {od.year} года не создан/неполный. Нельзя списывать дни.")
            return

        comment = self.ed_comment.text().strip()

        self.db.conn.execute("BEGIN;")
        try:
            if unit == "days":
                days = int(self.sp_days.value())
                if self.edit_mode:
                    self.db.update_compensation(
                        int(self.comp_id),
                        unit="days",
                        method="money",
                        event_date=d_iso(od),
                        order_no=order_no,
                        order_date=d_iso(od),
                        amount_days=days,
                        amount_minutes=None,
                        comment=comment or None,
                    )
                else:
                    self.db.add_compensation_money(int(self.employee["id"]), "days", None, days, order_no, od, comment)
            else:
                minutes = int(self.amount_hours.minutes_total())
                if minutes <= 0:
                    QMessageBox.warning(self, "Проверка", "Количество часов должно быть больше 0.")
                    self.db.conn.execute("ROLLBACK;")
                    return

                if self.edit_mode:
                    self.db.update_compensation(
                        int(self.comp_id),
                        unit="hours",
                        method="money",
                        event_date=d_iso(od),
                        order_no=order_no,
                        order_date=d_iso(od),
                        amount_minutes=minutes,
                        amount_days=None,
                        comment=comment or None,
                    )
                else:
                    self.db.add_compensation_money(int(self.employee["id"]), "hours", minutes, None, order_no, od, comment)

            ok, msg2 = validate_non_negative_over_year(self.db, int(self.employee["id"]), od.year)
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return

            self.db.conn.execute("COMMIT;")
            super().accept()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

class MoneyCompListDialog(QDialog):
    def __init__(self, db: DB, main: "MainWindow", employee: sqlite3.Row, parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.main = main
        self.employee = employee

        fio = f"{employee['last_name']} {employee['first_name']} {employee['middle_name'] or ''}".strip()
        self.setWindowTitle(f"Денежные компенсации — {fio}")
        self.resize(900, 500)

        lay = QVBoxLayout(self)

        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["Дата приказа", "Тип", "Количество", "№ приказа", "Комментарий"])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.cellDoubleClicked.connect(lambda r, c: self.edit_selected())
        lay.addWidget(self.table, 1)

        h = self.table.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        h.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        h.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        h.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        h.setSectionResizeMode(4, QHeaderView.Stretch)

        btns = QHBoxLayout()
        self.btn_add = QPushButton("Добавить")
        self.btn_edit = QPushButton("Редактировать")
        self.btn_del = QPushButton("Удалить")
        btns.addWidget(self.btn_add)
        btns.addWidget(self.btn_edit)
        btns.addWidget(self.btn_del)
        btns.addStretch(1)
        lay.addLayout(btns)

        self.btn_add.clicked.connect(self.add_new)
        self.btn_edit.clicked.connect(self.edit_selected)
        self.btn_del.clicked.connect(self.delete_selected)

        self.refresh()

    def selected_id(self) -> Optional[int]:
        sel = self.table.selectedItems()
        if not sel:
            return None
        return int(sel[0].data(Qt.UserRole))

    def refresh(self) -> None:
        eid = int(self.employee["id"])
        rows = self.db.conn.execute(
            """
            SELECT id, unit, order_no, order_date, amount_minutes, amount_days, comment
            FROM compensation
            WHERE employee_id=?
              AND method='money'
            ORDER BY COALESCE(order_date,event_date,'9999-12-31'), id
            """,
            (eid,),
        ).fetchall()

        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            cid = int(r["id"])
            unit = r["unit"]
            dt = fmt_date_iso(r["order_date"] or "")
            typ = "Часы" if unit == "hours" else "Дни"
            if unit == "hours":
                amount = minutes_to_hhmm(int(r["amount_minutes"] or 0))
            else:
                amount = f"{int(r['amount_days'] or 0)}"
            order_no = r["order_no"] or ""
            comment = r["comment"] or ""

            it0 = QTableWidgetItem(dt)
            it0.setData(Qt.UserRole, cid)
            self.table.setItem(i, 0, it0)
            self.table.setItem(i, 1, QTableWidgetItem(typ))
            self.table.setItem(i, 2, QTableWidgetItem(amount))
            self.table.setItem(i, 3, QTableWidgetItem(order_no))
            self.table.setItem(i, 4, QTableWidgetItem(comment))

    def add_new(self) -> None:
        y, _ = self.main.current_year_month()
        dlg = MoneyCompensationDialog(self.db, self.employee, default_year=y, comp_id=None, parent=self)
        if dlg.exec() == QDialog.Accepted:
            self.main.toast("Сохранено", 4000)
            self.refresh()
            self.main.refresh_context()

    def edit_selected(self) -> None:
        cid = self.selected_id()
        if cid is None:
            return
        y, _ = self.main.current_year_month()
        dlg = MoneyCompensationDialog(self.db, self.employee, default_year=y, comp_id=cid, parent=self)
        if dlg.exec() == QDialog.Accepted:
            self.main.toast("Сохранено", 4000)
            self.refresh()
            self.main.refresh_context()

    def delete_selected(self) -> None:
        cid = self.selected_id()
        if cid is None:
            return
        if QMessageBox.question(self, "Удаление", "Удалить денежную компенсацию?") != QMessageBox.Yes:
            return

        eid = int(self.employee["id"])

        self.db.conn.execute("BEGIN;")
        try:
            self.db.delete_compensation(cid)

            ok, msg2 = self.main.validate_employee_all_years(eid)
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return

            self.db.conn.execute("COMMIT;")
            self.main.toast("Удалено", 4000)
            self.refresh()
            self.main.refresh_context()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

class CalendarDialog(QDialog):
    def __init__(self, db: DB, main: "MainWindow", parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.main = main
        self.setWindowTitle("Календарь (рабочие/нерабочие)")
        lay = QVBoxLayout(self)

        top = QHBoxLayout()
        self.cb_year = QComboBox()
        for y in range(2000, 2100):
            self.cb_year.addItem(str(y), y)
        self.cb_month = QComboBox()
        for m in range(1, 13):
            self.cb_month.addItem(f"{m:02d}", m)
        self.btn_create_year = QPushButton("Создать календарь на год")
        top.addWidget(QLabel("Год:"))
        top.addWidget(self.cb_year)
        top.addWidget(QLabel("Месяц:"))
        top.addWidget(self.cb_month)
        top.addStretch(1)
        top.addWidget(self.btn_create_year)
        lay.addLayout(top)

        self.grid = QTableWidget(6, 7)
        self.grid.setHorizontalHeaderLabels(["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"])
        self.grid.verticalHeader().setVisible(False)
        self.grid.setEditTriggers(QTableWidget.NoEditTriggers)
        self.grid.setSelectionMode(QTableWidget.NoSelection)
        self.grid.cellClicked.connect(self.on_cell_clicked)
        lay.addWidget(self.grid)

        self.btn_create_year.clicked.connect(self.on_create_year)
        self.cb_year.currentIndexChanged.connect(self.refresh)
        self.cb_month.currentIndexChanged.connect(self.refresh)

        self.cb_year.setCurrentIndex(max(0, self.cb_year.findData(date.today().year)))
        self.cb_month.setCurrentIndex(date.today().month - 1)
        self.refresh()

    def on_create_year(self) -> None:
        y = int(self.cb_year.currentData())
        self.db.conn.execute("BEGIN;")
        try:
            self.db.create_calendar_year_default(y)
            ok_all = self.main.validate_all_employees_year(y)
            if not ok_all[0]:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", ok_all[1])
                return
            self.db.conn.execute("COMMIT;")
            self.refresh()
            self.main.refresh_all()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    def refresh(self) -> None:
        y = int(self.cb_year.currentData())
        m = int(self.cb_month.currentData())
        first = date(y, m, 1)
        start_col = first.weekday()
        ny, nm = next_month(y, m)
        last_day = (date(ny, nm, 1) - timedelta(days=1)).day

        for r in range(6):
            for c in range(7):
                self.grid.setItem(r, c, QTableWidgetItem(""))

        day_num = 1
        row = 0
        col = start_col
        while day_num <= last_day:
            d0 = date(y, m, day_num)
            it = QTableWidgetItem(str(day_num))
            it.setData(Qt.UserRole, d_iso(d0))
            try:
                if not self.db.is_working_day(d0):
                    it.setBackground(QBrush(QColor("#ffe0e0")))
            except CalendarMissingError:
                it.setBackground(QBrush(QColor("#e0e0e0")))
            self.grid.setItem(row, col, it)
            day_num += 1
            col += 1
            if col >= 7:
                col = 0
                row += 1

    def on_cell_clicked(self, r: int, c: int) -> None:
        it = self.grid.item(r, c)
        if not it or not it.text().strip():
            return
        d0 = d_parse(it.data(Qt.UserRole))

        self.db.conn.execute("BEGIN;")
        try:
            self.db.toggle_calendar_day(d0)
            ok_all = self.main.validate_all_employees_year(d0.year)
            if not ok_all[0]:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", ok_all[1])
                return
            self.db.conn.execute("COMMIT;")
            self.refresh()
            self.main.refresh_all()
        except CalendarMissingError as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.warning(self, "Календарь", str(e))
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))


# -----------------------------
# New UI: Quick defaults
# -----------------------------
class DefaultBreakTimeDialog(QDialog):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle("Перерыв (по умолчанию)")

        lay = QVBoxLayout(self)
        form = QFormLayout()
        lay.addLayout(form)

        self.t_start = TimeOfDayPicker()
        self.t_end = TimeOfDayPicker()

        self.t_start.set_time(time(13, 0))
        self.t_end.set_time(time(13, 30))

        form.addRow("Начало (время):", self.t_start)
        form.addRow("Конец (время):", self.t_end)

        self.lbl_hint = QLabel(
            "Если время начала/конца получается раньше начала дежурства,\n"
            "при применении будет автоматически перенос на следующие сутки."
        )
        self.lbl_hint.setStyleSheet("color:#808080;")
        lay.addWidget(self.lbl_hint)

        bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        lay.addWidget(bb)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)

    def set_values(self, start_t: time, end_t: time) -> None:
        self.t_start.set_time(start_t)
        self.t_end.set_time(end_t)

    def values(self) -> tuple[time, time]:
        return self.t_start.get_time(), self.t_end.get_time()

    def accept(self) -> None:
        s, e = self.values()
        # Тут не проверяем s<e, потому что e может быть "на следующие сутки"
        # (например 23:30–00:10). Проверка будет в применении/в DutyDialog.
        super().accept()

class QuickDefaultsWidget(QGroupBox):
    """Единый источник стандартов быстрых действий (с сохранением в cfg.ui)."""

    def __init__(self, cfg: AppConfig, parent: QWidget | None = None):
        super().__init__("Стандарты быстрых действий", parent)
        self.cfg = cfg

        lay = QFormLayout(self)

        self.std_start = TimeOfDayPicker()
        self.std_duration = DurationPicker(max_hours=48)
        self.std_comp_hours = DurationPicker(max_hours=24)

        # --- Перерывы по умолчанию: механика как в DutyDialog (чекбокс + список + Add/Edit/Delete) ---
        self.chk_breaks = QCheckBox("Перерывы по умолчанию")
        self.breaks_panel = QWidget()
        vb = QVBoxLayout(self.breaks_panel)
        vb.setContentsMargins(0, 0, 0, 0)

        self.list_breaks = QListWidget()
        vb.addWidget(self.list_breaks, 1)

        hb = QHBoxLayout()
        self.btn_break_add = QPushButton("Добавить…")
        self.btn_break_edit = QPushButton("Редактировать")
        self.btn_break_del = QPushButton("Удалить")
        self.btn_break_clear = QPushButton("Очистить")
        hb.addWidget(self.btn_break_add)
        hb.addWidget(self.btn_break_edit)
        hb.addWidget(self.btn_break_del)
        hb.addWidget(self.btn_break_clear)
        hb.addStretch(1)
        vb.addLayout(hb)

        self.breaks_panel.setVisible(False)

        # хранение: список (start_time, end_time)
        self._break_times: list[tuple[time, time]] = []

        self._load()

        lay.addRow("Дежурство (двойной ЛКМ) старт:", self.std_start)
        lay.addRow("Дежурство (двойной ЛКМ) длительность:", self.std_duration)
        lay.addRow("Выходной (быстро) часы:", self.std_comp_hours)

        lay.addRow(self.chk_breaks)
        lay.addRow(self.breaks_panel)

        for w in [
            self.std_start.sp_h,
            self.std_start.sp_m,
            self.std_duration.sp_h,
            self.std_duration.sp_m,
            self.std_comp_hours.sp_h,
            self.std_comp_hours.sp_m,
        ]:
            w.valueChanged.connect(self._save)

        self.chk_breaks.toggled.connect(self._on_breaks_toggled)
        self.btn_break_add.clicked.connect(self._add_break)
        self.btn_break_edit.clicked.connect(self._edit_break)
        self.btn_break_del.clicked.connect(self._delete_break)
        self.btn_break_clear.clicked.connect(self._clear_breaks)
        self.list_breaks.itemDoubleClicked.connect(lambda _=None: self._edit_break())

    def _on_breaks_toggled(self, v: bool) -> None:
        self.breaks_panel.setVisible(v)
        if not v:
            self._break_times = []
            self._rebuild_breaks_list()
        self._save()

    def _clear_breaks(self) -> None:
        self._break_times = []
        self._rebuild_breaks_list()
        self.chk_breaks.setChecked(False)
        self._save()

    def _selected_break_index(self) -> Optional[int]:
        it = self.list_breaks.currentItem()
        if not it:
            return None
        data = it.data(Qt.UserRole)
        if not data:
            return None
        try:
            s_str, e_str = data
            s = parse_hhmm(s_str)
            e = parse_hhmm(e_str)
        except Exception:
            return None

        for i, (a, b) in enumerate(self._break_times):
            if a == s and b == e:
                return i
        return None

    def _rebuild_breaks_list(self) -> None:
        self.list_breaks.clear()
        for s, e in self._break_times:
            note = ""
            if fmt_hhmm(e) <= fmt_hhmm(s):
                note = " (до следующих суток)"
            it = QListWidgetItem(f"{fmt_hhmm(s)}–{fmt_hhmm(e)}{note}")
            it.setData(Qt.UserRole, (fmt_hhmm(s), fmt_hhmm(e)))
            self.list_breaks.addItem(it)

    def _add_break(self) -> None:
        dlg = DefaultBreakTimeDialog(self)
        if dlg.exec() != QDialog.Accepted:
            return
        s, e = dlg.values()

        self._break_times.append((s, e))
        self._break_times = sorted(self._break_times, key=lambda x: (x[0].hour, x[0].minute))

        self.chk_breaks.setChecked(True)
        self._rebuild_breaks_list()
        self._save()

    def _edit_break(self) -> None:
        idx = self._selected_break_index()
        if idx is None:
            return

        s0, e0 = self._break_times[idx]
        dlg = DefaultBreakTimeDialog(self)
        dlg.set_values(s0, e0)
        if dlg.exec() != QDialog.Accepted:
            return

        s, e = dlg.values()
        self._break_times[idx] = (s, e)
        self._break_times = sorted(self._break_times, key=lambda x: (x[0].hour, x[0].minute))

        self._rebuild_breaks_list()
        self._save()

    def _delete_break(self) -> None:
        idx = self._selected_break_index()
        if idx is None:
            return
        del self._break_times[idx]
        self._rebuild_breaks_list()
        if not self._break_times:
            self.chk_breaks.setChecked(False)
        self._save()

    def _load(self) -> None:
        ui = (self.cfg.ui or {})

        start_str = ui.get("duty_std_start", "08:00")
        try:
            hh, mm = start_str.split(":")
            self.std_start.sp_h.setValue(int(hh))
            self.std_start.sp_m.setValue(int(mm))
        except Exception:
            self.std_start.sp_h.setValue(8)
            self.std_start.sp_m.setValue(0)

        dur_min = int(ui.get("duty_std_duration_minutes", 24 * 60))
        self.std_duration.set_minutes_total(dur_min)

        comp_min = int(ui.get("comp_std_minutes", 8 * 60))
        self.std_comp_hours.set_minutes_total(comp_min)

        self._break_times = []
        enabled = bool(ui.get("default_breaks_enabled", False))

        raw = ui.get("default_breaks", [])
        if isinstance(raw, list):
            for item in raw:
                if isinstance(item, dict):
                    s = parse_hhmm(item.get("start", "00:00"))
                    e = parse_hhmm(item.get("end", "00:00"))
                    self._break_times.append((s, e))

        self._break_times = sorted(self._break_times, key=lambda x: (x[0].hour, x[0].minute))
        self.chk_breaks.setChecked(bool(enabled and self._break_times))
        self.breaks_panel.setVisible(self.chk_breaks.isChecked())
        self._rebuild_breaks_list()

    def _save(self) -> None:
        ui = (self.cfg.ui or {})

        ui["duty_std_start"] = f"{int(self.std_start.sp_h.value()):02d}:{int(self.std_start.sp_m.value()):02d}"
        ui["duty_std_duration_minutes"] = int(self.std_duration.minutes_total())
        ui["comp_std_minutes"] = int(self.std_comp_hours.minutes_total())

        ui["default_breaks_enabled"] = bool(self.chk_breaks.isChecked())
        ui["default_breaks"] = [{"start": fmt_hhmm(s), "end": fmt_hhmm(e)} for s, e in self._break_times]

        self.cfg.ui = ui
        self.cfg.save()

    def duty_start_time(self) -> time:
        return self.std_start.get_time()

    def duty_duration_minutes(self) -> int:
        return int(self.std_duration.minutes_total())

    def comp_minutes(self) -> int:
        return int(self.std_comp_hours.minutes_total())

    def default_break_times(self) -> list[tuple[time, time]]:
        if not self.chk_breaks.isChecked():
            return []
        return list(self._break_times)
        
        # -----------------------------
# Calendar widget with marks (NO week numbers, NO other-month days)
# -----------------------------
class MarkedCalendarWidget(QCalendarWidget):
    rmb_requested = Signal(object, object, object)  # date, (row,col), global QPoint
    clipboard_action = Signal(object, object)  # date, "copy" | "paste"
    rmb_double_requested = Signal(object, object, object)  # date, (row,col), global QPoint
    
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setFocusPolicy(Qt.StrongFocus)        
        self.setGridVisible(True)
        self.setNavigationBarVisible(False)

        try:
            self.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)
        except Exception:
            pass

        self._duty_intervals: dict[date, list[tuple[datetime, datetime]]] = {}
        self._comp_dates: set[date] = set()
        self._overlap_dates: set[date] = set()

        self._fmt_nonwork = QTextCharFormat()
        self._fmt_nonwork.setForeground(QBrush(QColor("#c00000")))  # красный номер

        self._fmt_work_weekend = QTextCharFormat()
        self._fmt_work_weekend.setForeground(QBrush(QColor("#ffffff")))  # белый номер
        
        # запрет прокрутки колесом (QCalendarWidget иногда ловит wheel во внутреннем QTableView)
        v = self.findChild(QTableView, "qt_calendar_calendarview") or self.findChild(QTableView)
        if v:
            v.installEventFilter(self)
            v.viewport().installEventFilter(self)
    
    def wheelEvent(self, event):  # type: ignore[override]
        event.ignore()

    def eventFilter(self, obj, event):  # type: ignore[override]
        if event.type() == QEvent.Wheel:
            event.ignore()
            return True
            
        if event.type() in (QEvent.MouseButtonPress, QEvent.MouseButtonDblClick):
            try:
                if event.button() == Qt.MouseButton.RightButton:
                    v = self._view()
                    if not v:
                        return True

                    # координата клика должна быть в координатах viewport
                    pos = event.position().toPoint()
                    if obj is v:
                        pos = v.viewport().mapFrom(v, pos)

                    idx = v.indexAt(pos)
                    if not idx.isValid():
                        return True

                    row = int(idx.row())
                    col = int(idx.column())

                    shown = idx.data(Qt.ItemDataRole.DisplayRole)
                    try:
                        shown_day = int(shown)
                    except Exception:
                        shown_day = 0

                    # пустая ячейка/не число
                    if not (1 <= shown_day <= 31):
                        return True

                    # игнор хвостов других месяцев (как у тебя было)
                    if (row <= 1 and shown_day > 20) or (row >= 4 and shown_day < 15):
                        return True

                    y = self.yearShown()
                    m = self.monthShown()
                    try:
                        d0 = date(y, m, shown_day)
                    except ValueError:
                        return True

                    gp = event.globalPosition().toPoint()

                    if event.type() == QEvent.MouseButtonDblClick:
                        self.rmb_double_requested.emit(d0, (row, col), gp)
                    else:
                        self.rmb_requested.emit(d0, (row, col), gp)

                    event.accept()
                    return True
            except Exception:
                return False

        if event.type() == QEvent.KeyPress:
            if self._handle_copy_paste_key(event):
                return True

        return super().eventFilter(obj, event)

    def _handle_copy_paste_key(self, event) -> bool:
        try:
            if event.matches(QKeySequence.Copy):
                d0 = qdate_to_date(self.selectedDate())
                self.clipboard_action.emit(d0, "copy")
                event.accept()
                return True

            if event.matches(QKeySequence.Paste):
                d0 = qdate_to_date(self.selectedDate())
                self.clipboard_action.emit(d0, "paste")
                event.accept()
                return True
        except Exception:
            pass

        return False

    def keyPressEvent(self, event) -> None:  # type: ignore[override]
        if self._handle_copy_paste_key(event):
            return
        super().keyPressEvent(event)

    def set_marks(
        self,
        duty_intervals: dict[date, list[tuple[datetime, datetime]]],
        comp_dates: set[date],
        overlap_dates: set[date],
    ) -> None:
        self._duty_intervals = duty_intervals or {}
        self._comp_dates = comp_dates or set()
        self._overlap_dates = overlap_dates or set()
        self.updateCells()

    def apply_calendar_formats_for_month(
        self,
        year: int,
        month: int,
        nonworking_dates: set[date],
        working_weekend_dates: set[date],
    ) -> None:
        ny, nm = next_month(year, month)
        last_day = (date(ny, nm, 1) - timedelta(days=1)).day

        empty = QTextCharFormat()
        for day in range(1, last_day + 1):
            self.setDateTextFormat(QDate(year, month, day), empty)

        # 1) выходные, которые сделали рабочими -> белый номер
        for d0 in working_weekend_dates:
            if d0.year == year and d0.month == month:
                self.setDateTextFormat(QDate(d0.year, d0.month, d0.day), self._fmt_work_weekend)

        # 2) нерабочие -> красный номер (перекрывает)
        for d0 in nonworking_dates:
            if d0.year == year and d0.month == month:
                self.setDateTextFormat(QDate(d0.year, d0.month, d0.day), self._fmt_nonwork)

    def _fmt_end_in_day(self, day: date, dt: datetime) -> str:
        return dt.strftime("%H:%M")

    def _view(self) -> Optional[QTableView]:
        return self.findChild(QTableView, "qt_calendar_calendarview") or self.findChild(QTableView)

    def _date_for_cell(self, row: int, col: int) -> date:
        y = self.yearShown()
        m = self.monthShown()
        first = date(y, m, 1)
        fd = int(self.firstDayOfWeek().value) - 1  # 0..6
        fw = first.weekday()  # 0..6
        offset = (fw - fd) % 7
        day_offset = row * 7 + col - offset
        return first + timedelta(days=day_offset)

    def contextMenuEvent(self, event):  # type: ignore[override]
        event.ignore()

    def paintCell(self, painter: QPainter, rect: QRect, qd: QDate) -> None:  # type: ignore[override]
        d0 = qdate_to_date(qd)

        # дни другого месяца не показываем
        if d0.year != self.yearShown() or d0.month != self.monthShown():
            painter.save()
            painter.fillRect(rect, self.palette().color(QPalette.Base))
            painter.restore()
            return

        super().paintCell(painter, rect, qd)

        intervals = self._duty_intervals.get(d0, [])
        has_duty = len(intervals) > 0
        has_comp = d0 in self._comp_dates
        is_overlap = d0 in self._overlap_dates

        if not (has_duty or has_comp or is_overlap):
            return

        painter.save()
        painter.setRenderHint(QPainter.Antialiasing, True)

        # подложка под дежурства
        if has_duty:
            painter.fillRect(rect.adjusted(1, 1, -1, -1), QColor(195, 247, 195, 110))

        # маркер компенсации
        if has_comp:
            painter.setPen(Qt.NoPen)
            painter.setBrush(QColor(158, 203, 255, 220))
            r = 4
            painter.drawEllipse(QPoint(rect.right() - 7, rect.bottom() - 7), r, r)

        # интервалы дежурств текстом
        if has_duty:
            parts: list[str] = []
            for a, b in intervals:
                parts.append(f"{a.strftime('%H:%M')}–{self._fmt_end_in_day(d0, b)}")
            txt = "\n".join(parts)

            f = painter.font()
            f.setPointSize(max(6, f.pointSize() - 2))
            painter.setFont(f)
            painter.setPen(QColor(60, 60, 60, 230))
            painter.drawText(rect.adjusted(2, 12, -2, -2), Qt.AlignLeft | Qt.AlignBottom, txt)

        if is_overlap:
            pen = QPen(QColor("#f0ad4e"))
            pen.setWidth(2)
            painter.setPen(pen)
            painter.setBrush(Qt.NoBrush)
            painter.drawRect(rect.adjusted(1, 1, -2, -2))

        painter.restore()


def compute_day_intervals_in_month(
    db: DB, employee_id: int, year: int, month: int
) -> tuple[dict[date, list[tuple[datetime, datetime]]], set[date]]:
    """date -> list of merged intervals within that day (00:00..24:00), с учетом перерывов."""
    m_start, m_end = month_bounds_dt(year, month)
    duties = db.list_duties_for_month(employee_id, year, month)

    duty_ids = [int(r["id"]) for r in duties]
    breaks_map = db.breaks_for_duty_ids(duty_ids)

    per_day: dict[date, list[tuple[datetime, datetime]]] = {}

    for r in duties:
        did = int(r["id"])

        s0 = max(dt_parse(r["start_dt"]), m_start)
        e0 = min(dt_parse(r["end_dt"]), m_end)
        if s0 >= e0:
            continue

        parts = subtract_intervals((s0, e0), breaks_map.get(did, []))

        for s, e in parts:
            cur = s.date()
            last = (e - timedelta(seconds=1)).date()
            while cur <= last:
                d0 = datetime.combine(cur, time(0, 0))
                d1 = d0 + timedelta(days=1)
                inter = intersect(s, e, d0, d1)
                if inter:
                    per_day.setdefault(cur, []).append(inter)
                cur += timedelta(days=1)

    merged_by_day: dict[date, list[tuple[datetime, datetime]]] = {}
    overlap_days: set[date] = set()

    for d, intervals in per_day.items():
        raw_cnt = len(intervals)
        merged = merge_intervals(intervals)
        merged_by_day[d] = merged
        if len(merged) < raw_cnt:
            overlap_days.add(d)

    return merged_by_day, overlap_days


def compute_comp_dates_in_month(db: DB, employee_id: int, year: int, month: int) -> set[date]:
    rows = db.list_compensations_for_month(employee_id, year, month)
    out: set[date] = set()

    for r in rows:
        unit = r["unit"]
        method = r["method"]

        if method == "money":
            continue

        if unit == "hours":
            if r["event_date"]:
                out.add(d_parse(r["event_date"]))
        else:
            dates = db.get_comp_dates(int(r["id"]))
            for x in dates:
                try:
                    out.add(d_parse(x))
                except Exception:
                    pass

    return {d0 for d0 in out if d0.year == year and d0.month == month}


# -----------------------------
# Month summary panel (ONLY month totals)
# -----------------------------
class MonthSummaryPanel(QGroupBox):
    def __init__(self, db: DB, main: "MainWindow", parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.main = main
        self.employee: Optional[sqlite3.Row] = None
        self.year = date.today().year
        self.month = date.today().month

        # Главный контейнер (карточка)
        self.setObjectName("SummaryCard")
        self.setStyleSheet("""
            QWidget#SummaryCard {
                background-color: #2C2C2E;
                border-radius: 12px;
            }
        """)
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)

        # Заголовок дашборда
        title_row = QHBoxLayout()
        title = QLabel("Итоги месяца")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #FFFFFF;")
        title_row.addWidget(title)
        
        self.lbl_days_warn = QLabel("")
        self.lbl_days_warn.setStyleSheet("color: #FF453A; font-size: 12px;") # Красный цвет Apple
        title_row.addWidget(self.lbl_days_warn, 1, Qt.AlignRight)
        main_layout.addLayout(title_row)

        # Сетка метрик (Дашборд)
        grid = QGridLayout()
        grid.setHorizontalSpacing(30)
        grid.setVerticalSpacing(15)

        # Хелпер для создания красивых блоков
        def make_metric_block(title: str, val_h: QLabel, val_d: QLabel):
            vbox = QVBoxLayout()
            vbox.setSpacing(4)
            lbl_t = QLabel(title)
            lbl_t.setStyleSheet("font-size: 12px; font-weight: 600; color: #8E8E93; text-transform: uppercase;")
            vbox.addWidget(lbl_t)
            
            hbox = QHBoxLayout()
            val_h.setStyleSheet("font-size: 22px; font-weight: 500; color: #FFFFFF;")
            val_d.setStyleSheet("font-size: 22px; font-weight: 500; color: #FFFFFF;")
            hbox.addWidget(val_h)
            hbox.addWidget(val_d)
            hbox.addStretch()
            vbox.addLayout(hbox)
            return vbox

        self.lbl_start_hours = QLabel("—")
        self.lbl_start_days = QLabel("—")
        self.lbl_comp_hours = QLabel("—")
        self.lbl_comp_days = QLabel("—")
        self.lbl_end_hours = QLabel("—")
        self.lbl_end_days = QLabel("—")
        self.lbl_money_comp = QLabel("—")
        self.lbl_money_comp.setStyleSheet("font-size: 14px; color: #32D74B;") # Зеленый цвет для денег

        # Расставляем блоки
        l1 = make_metric_block("На начало", self.lbl_start_hours, self.lbl_start_days)
        l2 = make_metric_block("Списано (выходные)", self.lbl_comp_hours, self.lbl_comp_days)
        l3 = make_metric_block("Остаток", self.lbl_end_hours, self.lbl_end_days)

        grid.addLayout(l1, 0, 0)
        grid.addLayout(l2, 0, 1)
        grid.addLayout(l3, 0, 2)

        # Блок денег
        money_box = QVBoxLayout()
        lbl_m = QLabel("ДЕНЬГАМИ:")
        lbl_m.setStyleSheet("font-size: 12px; font-weight: 600; color: #8E8E93;")
        money_box.addWidget(lbl_m)
        money_box.addWidget(self.lbl_money_comp)
        grid.addLayout(money_box, 1, 0, 1, 3)

        main_layout.addLayout(grid)

        # Кнопки аккуратно внизу
        btn_row = QHBoxLayout()
        self.btn_add_money = QPushButton("+ Деньгами")
        self.btn_manage_money = QPushButton("Все приказы...")
        btn_row.addWidget(self.btn_add_money)
        btn_row.addWidget(self.btn_manage_money)
        btn_row.addStretch()
        main_layout.addLayout(btn_row)

        self.btn_add_money.clicked.connect(self.main.add_money_compensation_from_summary)
        self.btn_manage_money.clicked.connect(self.main.open_money_comp_list)

        self.setEnabled(False)

    def set_context(self, emp: Optional[sqlite3.Row], year: int, month: int) -> None:
        self.employee = emp
        self.year = year
        self.month = month
        self.refresh()

    def _money_comp_totals_for_month(self, employee_id: int, year: int, month: int) -> tuple[int, int]:
        m0 = date(year, month, 1)
        ny, nm = next_month(year, month)
        m1 = date(ny, nm, 1)

        r1 = self.db.conn.execute(
            """
            SELECT COALESCE(SUM(amount_minutes),0) AS m
            FROM compensation
            WHERE employee_id=?
              AND method='money'
              AND unit='hours'
              AND event_date IS NOT NULL
              AND event_date >= ? AND event_date < ?
            """,
            (employee_id, d_iso(m0), d_iso(m1)),
        ).fetchone()
        money_minutes = int(r1["m"])

        r2 = self.db.conn.execute(
            """
            SELECT COALESCE(SUM(amount_days),0) AS d
            FROM compensation
            WHERE employee_id=?
              AND method='money'
              AND unit='days'
              AND event_date IS NOT NULL
              AND event_date >= ? AND event_date < ?
            """,
            (employee_id, d_iso(m0), d_iso(m1)),
        ).fetchone()
        money_days = int(r2["d"])

        return money_minutes, money_days

    def refresh(self) -> None:
        if not self.employee:
            self.setEnabled(False)
            self.lbl_start_hours.setText("—")
            self.lbl_start_days.setText("—")
            self.lbl_comp_hours.setText("—")
            self.lbl_comp_days.setText("—")
            self.lbl_money_comp.setText("—")
            self.lbl_end_hours.setText("—")
            self.lbl_end_days.setText("—")
            self.lbl_days_warn.setText("")
            return

        self.setEnabled(True)

        try:
            eid = int(self.employee["id"])

            # баланс на начало месяца
            start_bal_m, start_bal_d = TemplateExporter._balance_at_start_of_month(self.db, eid, self.year, self.month)

            # итоги месяца (в т.ч. баланс на конец)
            summ = compute_month_summary(self.db, eid, self.year, self.month)

            # компенсации "выходными" (не деньги) за месяц
            comp_rest_minutes = TemplateExporter._sum_comp_rest_hours_minutes(self.db, eid, self.year, self.month)
            comp_rest_days = TemplateExporter._count_comp_rest_days(self.db, eid, self.year, self.month)

            # денежные компенсации за месяц
            money_minutes, money_days = self._money_comp_totals_for_month(eid, self.year, self.month)

            # --- заполняем UI ---
            self.lbl_start_hours.setText(fmt_minutes_ru_words(int(start_bal_m)))
            self.lbl_end_hours.setText(fmt_minutes_ru_words(int(summ["bal_minutes"])))

            self.lbl_comp_hours.setText(fmt_minutes_ru_words(int(comp_rest_minutes)))

            # дни показываем только если календарь ок
            if summ["days_ok"]:
                self.lbl_start_days.setText(fmt_days_ru_words(int(start_bal_d)))
                self.lbl_comp_days.setText(fmt_days_ru_words(int(comp_rest_days)))
                self.lbl_end_days.setText(fmt_days_ru_words(int(summ["bal_days"])))
                self.lbl_days_warn.setText("")
            else:
                self.lbl_start_days.setText("—")
                self.lbl_comp_days.setText("—")
                self.lbl_end_days.setText("—")
                self.lbl_days_warn.setText(summ["days_err"] or "")

            # деньги: одной строкой (часы и/или дни)
            parts: list[str] = []
            if money_minutes:
                parts.append(fmt_minutes_ru_words(int(money_minutes)))
            if money_days:
                parts.append(fmt_days_ru_words(int(money_days)))

            if parts:
                self.lbl_money_comp.setText(", ".join(parts))
            else:
                self.lbl_money_comp.setText("0")

        except Exception as e:
            self.lbl_days_warn.setText(f"Ошибка расчета: {e}")


# -----------------------------
# Day dialog (opened via RMB on calendar day)
# -----------------------------
class DayDialog(QDialog):
    def __init__(self, db: DB, main: "MainWindow", employee: sqlite3.Row, year: int, day: date, parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.main = main
        self.employee = employee
        self.year = year
        self.day = day

        self.setWindowTitle(f"День — {self.day.strftime('%d.%m.%Y')}")
        self.resize(850, 600)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 20, 20, 20)
        lay.setSpacing(15)

        self.lbl_day = QLabel(self.day.strftime("%d %B %Y").lower())
        self.lbl_day.setStyleSheet("font-size: 20px; font-weight: bold; color: #FFFFFF;")
        lay.addWidget(self.lbl_day)

        # --- Хелпер для создания секций с кнопкой + ---
        def create_section_header(title_text, add_callback):
            row = QHBoxLayout()
            lbl = QLabel(title_text)
            lbl.setStyleSheet("font-size: 14px; font-weight: 600; color: #8E8E93;")
            btn_add = QPushButton("+ Добавить")
            btn_add.setStyleSheet("background: transparent; color: #0A84FF; font-weight: bold; padding: 0;")
            btn_add.setCursor(Qt.PointingHandCursor)
            btn_add.clicked.connect(add_callback)
            row.addWidget(lbl)
            row.addStretch()
            row.addWidget(btn_add)
            return row

        # --- Duties ---
        lay.addLayout(create_section_header("ДЕЖУРСТВА", self.add_duty))
        
        self.tbl_duty = QTableWidget(0, 3)
        self.tbl_duty.setHorizontalHeaderLabels(["Начало", "Конец", "Комментарий"])
        self.tbl_duty.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_duty.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_duty.cellDoubleClicked.connect(lambda r, c: self.edit_duty())
        
        # Настройка контекстного меню для дежурств
        self.tbl_duty.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tbl_duty.customContextMenuRequested.connect(self.show_duty_context_menu)
        
        hd = self.tbl_duty.horizontalHeader()
        hd.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hd.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hd.setSectionResizeMode(2, QHeaderView.Stretch)
        lay.addWidget(self.tbl_duty, 1)


        # --- Compensations ---
        lay.addSpacing(10)
        lay.addLayout(create_section_header("КОМПЕНСАЦИИ", self.add_comp))

        self.tbl_comp = QTableWidget(0, 5)
        self.tbl_comp.setHorizontalHeaderLabels(["Тип", "Способ", "Дата/Приказ", "Количество", "Комментарий"])
        self.tbl_comp.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_comp.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_comp.cellDoubleClicked.connect(lambda r, c: self.edit_comp())
        
        # Настройка контекстного меню для компенсаций
        self.tbl_comp.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tbl_comp.customContextMenuRequested.connect(self.show_comp_context_menu)

        hc = self.tbl_comp.horizontalHeader()
        hc.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hc.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hc.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        hc.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        hc.setSectionResizeMode(4, QHeaderView.Stretch)
        lay.addWidget(self.tbl_comp, 1)

        bb = QDialogButtonBox(QDialogButtonBox.Close)
        bb.rejected.connect(self.reject)
        bb.accepted.connect(self.accept)
        lay.addWidget(bb)

        self.refresh_tables()

    # --- Добавляем методы контекстного меню ---
    def show_duty_context_menu(self, pos):
        if not self.tbl_duty.itemAt(pos): return
        menu = QMenu(self)
        menu.addAction("Редактировать...", self.edit_duty)
        menu.addSeparator()
        del_action = menu.addAction("Удалить")
        # В QSS можно настроить красный цвет для критических действий
        action = menu.exec(self.tbl_duty.viewport().mapToGlobal(pos))
        if action == del_action:
            self.delete_duty()

    def show_comp_context_menu(self, pos):
        if not self.tbl_comp.itemAt(pos): return
        menu = QMenu(self)
        menu.addAction("Редактировать...", self.edit_comp)
        menu.addSeparator()
        del_action = menu.addAction("Удалить")
        action = menu.exec(self.tbl_comp.viewport().mapToGlobal(pos))
        if action == del_action:
            self.delete_comp()

    def refresh_tables(self) -> None:
        s_day = datetime.combine(self.day, time(0, 0))
        e_day = s_day + timedelta(days=1)
        duties = self.db.list_duties_for_period(int(self.employee["id"]), s_day, e_day)

        self.tbl_duty.setRowCount(len(duties))
        for i, r in enumerate(duties):
            it0 = QTableWidgetItem(fmt_dt_iso(r["start_dt"]))
            it1 = QTableWidgetItem(fmt_dt_iso(r["end_dt"]))
            it2 = QTableWidgetItem(r["comment"] or "")
            it0.setData(Qt.UserRole, int(r["id"]))
            self.tbl_duty.setItem(i, 0, it0)
            self.tbl_duty.setItem(i, 1, it1)
            self.tbl_duty.setItem(i, 2, it2)

        d_iso0 = d_iso(self.day)
        rows = self.db.conn.execute(
            """
            SELECT * FROM compensation
            WHERE employee_id=?
              AND method <> 'money'
              AND (
                    (event_date IS NOT NULL AND event_date = ?)
                    OR
                    (unit='days' AND method='day_off' AND id IN (
                        SELECT DISTINCT compensation_id FROM comp_day_off_date
                        WHERE employee_id=? AND day_off_date = ?
                    ))
              )
            ORDER BY id
            """,
            (int(self.employee["id"]), d_iso0, int(self.employee["id"]), d_iso0),
        ).fetchall()

        self.tbl_comp.setRowCount(len(rows))
        for i, r in enumerate(rows):
            unit = r["unit"]
            method = r["method"]
            typ = "Часы" if unit == "hours" else "Дни"
            mth = "Выходной" if method == "day_off" else "Деньги"
            when = ""
            amount = ""

            if unit == "hours":
                when = fmt_date_iso(r["event_date"] or "")
                amount = minutes_to_hhmm(int(r["amount_minutes"] or 0))
            else:
                dates = self.db.get_comp_dates(int(r["id"]))
                when = fmt_date_iso(d_iso0) + f" (в записи всего: {len(dates)})"
                amount = "1"

            it0 = QTableWidgetItem(typ)
            it0.setData(Qt.UserRole, int(r["id"]))
            self.tbl_comp.setItem(i, 0, it0)
            self.tbl_comp.setItem(i, 1, QTableWidgetItem(mth))
            self.tbl_comp.setItem(i, 2, QTableWidgetItem(when))
            self.tbl_comp.setItem(i, 3, QTableWidgetItem(amount))
            self.tbl_comp.setItem(i, 4, QTableWidgetItem(r["comment"] or ""))

        self.main.refresh_context()

    def _selected_duty_id(self) -> Optional[int]:
        sel = self.tbl_duty.selectedItems()
        if not sel:
            return None
        return int(sel[0].data(Qt.UserRole))

    def _selected_comp_id(self) -> Optional[int]:
        sel = self.tbl_comp.selectedItems()
        if not sel:
            return None
        return int(sel[0].data(Qt.UserRole))

    def add_duty(self) -> None:
        start_t = self.main.defaults_widget.duty_start_time()
        dur_min = self.main.defaults_widget.duty_duration_minutes()
        start = datetime.combine(self.day, start_t)
        end = start + timedelta(minutes=dur_min)

        dlg = DutyDialog(self.db, int(self.employee["id"]), parent=self)
        dlg.set_values(start, end, "", self.main.default_breaks_for_duty(start, end))
        if dlg.exec() != QDialog.Accepted:
            return
        s, e, comment = dlg.get_values()
        breaks = dlg.get_breaks()
        
        msg = validate_event_in_employee_period(self.employee, s.date())
        if msg:
            QMessageBox.warning(self, "Проверка", msg)
            return

        self.db.conn.execute("BEGIN;")
        try:
            duty_id = self.db.add_duty(int(self.employee["id"]), s, e, comment)
            self.db.replace_duty_breaks(duty_id, breaks)
            ok, msg2 = self.main.validate_employee_all_years(int(self.employee["id"]))
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")
            self.main.toast("Дежурство добавлено")
            self.refresh_tables()
        except Exception as ex:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(ex))

    def edit_duty(self) -> None:
        did = self._selected_duty_id()
        if did is None:
            return
        row = self.db.get_duty(did)
        br_map = self.db.breaks_for_duty_ids([did])
        breaks0 = br_map.get(did, [])        
        dlg = DutyDialog(self.db, int(self.employee["id"]), exclude_duty_id=did, parent=self)
        dlg.set_values(dt_parse(row["start_dt"]), dt_parse(row["end_dt"]), row["comment"] or "", breaks0)
        if dlg.exec() != QDialog.Accepted:
            return
        s, e, comment = dlg.get_values()
        breaks = dlg.get_breaks()        

        msg = validate_event_in_employee_period(self.employee, s.date())
        if msg:
            QMessageBox.warning(self, "Проверка", msg)
            return

        self.db.conn.execute("BEGIN;")
        try:
            self.db.update_duty(did, start_dt=dt_iso(s), end_dt=dt_iso(e), comment=comment or None)
            self.db.replace_duty_breaks(did, breaks)            
            ok, msg2 = self.main.validate_employee_all_years(int(self.employee["id"]))
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")
            self.main.toast("Сохранено")
            self.refresh_tables()
        except Exception as ex:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(ex))

    def delete_duty(self) -> None:
        did = self._selected_duty_id()
        if did is None:
            return
        if QMessageBox.question(self, "Удаление", "Удалить дежурство?") != QMessageBox.Yes:
            return

        snap = self.db.snapshot_duty(did)
        self.db.conn.execute("BEGIN;")
        try:
            self.db.delete_duty(did)
            ok, msg2 = self.main.validate_employee_all_years(int(self.employee["id"]))
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")
            self.main.toast("Удалено")
            self.refresh_tables()
        except Exception as ex:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(ex))

    def add_comp(self) -> None:
        dlg = CompensationDialog(
            self.db,
            self.employee,
            self.year,
            self.main.workday_minutes(),
            comp_id=None,
            parent=self,
        )
        dlg.prefill_date(self.day)

        if dlg.exec() == QDialog.Accepted:
            self.main.toast("Сохранено")
            self.refresh_tables()

    def edit_comp(self) -> None:
        cid = self._selected_comp_id()
        if cid is None:
            return
    
        c = self.db.get_compensation(cid)
        if c["method"] == "money":
            dlg = MoneyCompensationDialog(self.db, self.employee, default_year=self.year, comp_id=cid, parent=self)
        else:
            dlg = CompensationDialog(self.db, self.employee, self.year, self.main.workday_minutes(), comp_id=cid, parent=self)
    
        if dlg.exec() == QDialog.Accepted:
            self.main.toast("Сохранено")
            self.refresh_tables()
            
    def delete_comp(self) -> None:
        cid = self._selected_comp_id()
        if cid is None:
            return
        if QMessageBox.question(self, "Удаление", "Удалить компенсацию?") != QMessageBox.Yes:
            return

        self.db.conn.execute("BEGIN;")
        try:
            self.db.delete_compensation(cid)
            ok, msg2 = self.main.validate_employee_all_years(int(self.employee["id"]))
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")
            self.main.toast("Удалено")
            self.refresh_tables()
        except Exception as ex:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(ex))


# -----------------------------
# Calendar panel (center top)
# -----------------------------
class CalendarPanel(QWidget):
    """Шапка + календарь. ПКМ: меню, двойной ПКМ: быстрая компенсация."""
    def _on_left_click(self, qd: QDate) -> None:
        if self._syncing:
            return

        d0 = qdate_to_date(qd)

        # если кликнули "чужой месяц" — сразу откатываем выделение назад
        if d0.year != self.year or d0.month != self.month:
            self._syncing = True
            try:
                self.cal.setCurrentPage(self.year, self.month)
                back = self._last_in_month_selected
                self.cal.setSelectedDate(QDate(back.year, back.month, back.day))
            finally:
                self._syncing = False
            return

        self._last_in_month_selected = d0
        
    def _on_rmb_double_requested(self, d0: date, cell: tuple[int, int], global_pos: QPoint) -> None:
        # если было запланировано меню по 1-му ПКМ — отменяем
        try:
            self._rmb_timer.stop()
        except Exception:
            pass

        self._rmb_pending_date = None
        self._rmb_pending_global_pos = None

        # на всякий случай закрываем меню, если оно вдруг открыто
        if self._ctx_menu is not None:
            try:
                self._ctx_menu.close()
                self._ctx_menu.deleteLater()
            except Exception:
                pass
            self._ctx_menu = None

        self.quick_add_hour_compensation(d0)

    def __init__(self, db: DB, main: "MainWindow", parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.main = main
        self.employee: Optional[sqlite3.Row] = None
        self.year = date.today().year
        self.month = date.today().month
        self._syncing = False

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0) # Убираем лишние отступы

        # --- НОВАЯ ШАПКА КАЛЕНДАРЯ ---
        header = QHBoxLayout()
        header.setContentsMargins(0, 0, 0, 10)
        header.setSpacing(12)

        # Аккуратные круглые кнопки навигации
        self.btn_prev = QPushButton("〈")
        self.btn_next = QPushButton("〉")
        for b in (self.btn_prev, self.btn_next):
            b.setFixedSize(32, 32)
            b.setStyleSheet("""
                QPushButton { border-radius: 16px; background: #3A3A3C; font-weight: bold; font-size: 16px; }
                QPushButton:hover { background: #48484A; }
            """)
            b.setCursor(Qt.PointingHandCursor)

        # Заголовок месяца в стиле Apple (крупно, без рамки)
        self.btn_period = QPushButton("")
        self.btn_period.setStyleSheet("background: transparent; font-size: 22px; font-weight: bold; color: #FFFFFF;")
        self.btn_period.setCursor(Qt.PointingHandCursor)
        self.btn_period.clicked.connect(self.open_month_picker)

        header.addWidget(self.btn_prev)
        header.addWidget(self.btn_period)
        header.addWidget(self.btn_next)
        
        # Сохраняем ссылку на layout шапки, чтобы MainWindow мог добавить туда Печать/Экспорт
        self.header_layout = header 
        lay.addLayout(header)

        self.cal = MarkedCalendarWidget()
        lay.addWidget(self.cal, 1)

        self.btn_prev.clicked.connect(lambda: self.main.shift_month(-1))
        self.btn_next.clicked.connect(lambda: self.main.shift_month(+1))

        self.cal.selectionChanged.connect(self._on_selection_changed)
        self.cal.activated.connect(self._on_left_double_click)
        self._last_in_month_selected: date = date(self.year, self.month, 1)
        self.cal.clicked.connect(self._on_left_click)

        # ПКМ от календаря
        self.cal.rmb_requested.connect(self._on_rmb_requested)
        self.cal.clipboard_action.connect(self._on_calendar_clipboard_action)
        self.cal.rmb_double_requested.connect(self._on_rmb_double_requested)
        self._duty_clipboard: Optional[dict[str, Any]] = None        
        self._ctx_menu: Optional[QMenu] = None
        self._ctx_menu_date: Optional[date] = None
        self._ctx_menu_cell: Optional[tuple[int, int]] = None

        self._rmb_last_ts_ms: int = 0
        self._rmb_last_cell: Optional[tuple[int, int]] = None
        self._rmb_pending_date: Optional[date] = None
        self._rmb_pending_global_pos: Optional[QPoint] = None

        self._rmb_timer = QTimer(self)
        self._rmb_timer.setSingleShot(True)
        self._rmb_timer.timeout.connect(self._show_rmb_menu)

        self.refresh_header()

    def set_context(self, emp: Optional[sqlite3.Row], year: int, month: int) -> None:
        self.employee = emp
        self.year = year
        self.month = month

        self._syncing = True
        try:
            self.cal.setCurrentPage(year, month)
            sel = qdate_to_date(self.cal.selectedDate())
            if sel.year != year or sel.month != month:
                today = date.today()
                d0 = today if (today.year == year and today.month == month) else date(year, month, 1)
                self.cal.setSelectedDate(QDate(d0.year, d0.month, d0.day))
        finally:
            self._syncing = False

        self.refresh_marks()
        self.refresh_header()
        self._last_in_month_selected = self.selected_date()
        
    def selected_date(self) -> date:
        return qdate_to_date(self.cal.selectedDate())

    def refresh_header(self) -> None:
        names = [
            "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
            "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
        ]
        month_name = names[self.month - 1] if 1 <= self.month <= 12 else str(self.month)
        # Убрали слово "года" для минимализма
        self.btn_period.setText(f"{month_name} {self.year}")

    def open_month_picker(self) -> None:
        dlg = QDialog(self, Qt.Popup)
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(8, 8, 8, 8)

        top = QHBoxLayout()
        top.addWidget(QLabel("Год:"))
        sp_year = QSpinBox()
        sp_year.setRange(2000, 2099)
        sp_year.setValue(self.year)
        configure_spinbox(sp_year)
        top.addWidget(sp_year)
        top.addStretch(1)
        lay.addLayout(top)

        names = [
            "Январь", "Февраль", "Март", "Апрель",
            "Май", "Июнь", "Июль", "Август",
            "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
        ]

        grid = QGridLayout()
        grid.setSpacing(6)
        lay.addLayout(grid)

        buttons: list[QPushButton] = []

        def apply_highlight() -> None:
            y = int(sp_year.value())
            for i, b in enumerate(buttons, start=1):
                if y == self.year and i == self.month:
                    b.setStyleSheet("font-weight:600;")
                else:
                    b.setStyleSheet("")

        def pick_month(mm: int) -> None:
            y = int(sp_year.value())
            dlg.accept()
            self.main.set_period(y, mm)

        for i, name in enumerate(names, start=1):
            btn = QPushButton(name)
            btn.clicked.connect(lambda _=False, mm=i: pick_month(mm))
            buttons.append(btn)
            r = (i - 1) // 3
            c = (i - 1) % 3
            grid.addWidget(btn, r, c)

        sp_year.valueChanged.connect(lambda _=None: apply_highlight())
        apply_highlight()

        dlg.adjustSize()
        gp = self.btn_period.mapToGlobal(QPoint(0, self.btn_period.height()))
        dlg.move(gp)
        dlg.exec()

    def on_today(self) -> None:
        today = date.today()
        self.main.set_period(today.year, today.month)
        self.cal.setSelectedDate(QDate(today.year, today.month, today.day))

    def refresh_marks(self) -> None:
        # 1) Форматирование номеров дней (раб/нераб)
        nonworking: set[date] = set()
        working_weekend: set[date] = set()
        try:
            m_start, m_end = month_bounds_dt(self.year, self.month)
            rows = self.db.conn.execute(
                "SELECT date, is_working FROM calendar_day WHERE date>=? AND date<?",
                (d_iso(m_start.date()), d_iso(m_end.date())),
            ).fetchall()
            for r in rows:
                d0 = d_parse(r["date"])
                if int(r["is_working"]) == 0:
                    nonworking.add(d0)
                else:
                    if d0.weekday() >= 5:  # Сб/Вс, но рабочий -> белый номер
                        working_weekend.add(d0)
        except Exception:
            pass
    
        self.cal.apply_calendar_formats_for_month(self.year, self.month, nonworking, working_weekend)
    
        # 2) Метки дежурств/компенсаций (как раньше)
        if not self.employee:
            self.cal.set_marks({}, set(), set())
            return
    
        eid = int(self.employee["id"])
        duty_intervals, overlap_days = compute_day_intervals_in_month(self.db, eid, self.year, self.month)
        comp_dates = compute_comp_dates_in_month(self.db, eid, self.year, self.month)
        self.cal.set_marks(duty_intervals, comp_dates, overlap_days)

    def _on_selection_changed(self) -> None:
        if self._syncing:
            return

        d0 = self.selected_date()

        # если выбор улетел в другой месяц (клавиатура/хвост) — возвращаем назад
        if d0.year != self.year or d0.month != self.month:
            self._syncing = True
            try:
                self.cal.setCurrentPage(self.year, self.month)
                back = self._last_in_month_selected
                self.cal.setSelectedDate(QDate(back.year, back.month, back.day))
            finally:
                self._syncing = False
            return

        self._last_in_month_selected = d0

    def _on_left_double_click(self, qd: QDate) -> None:
        d0 = qdate_to_date(qd)
        if d0.year != self.year or d0.month != self.month:
            self.main.toast("Дата вне выбранного месяца", 2500)
            return
        self.quick_add_default_duty(d0)

    def _on_rmb_requested(self, d0: date, cell: tuple[int, int], global_pos: QPoint) -> None:
        self._rmb_pending_date = d0
        self._rmb_pending_global_pos = global_pos

        # Ставим короткую задержку (150 мс), чтобы меню открывалось шустро
        self._rmb_timer.start(150)

    def _show_rmb_menu(self) -> None:
        d0 = self._rmb_pending_date
        gp = self._rmb_pending_global_pos

        if d0 is None or gp is None:
            return

        # сбрасываем "ожидание второго ПКМ"
        self._rmb_pending_date = None
        self._rmb_pending_global_pos = None

        # закрыть старое меню, если было
        if self._ctx_menu is not None:
            try:
                self._ctx_menu.close()
                self._ctx_menu.deleteLater()
            except Exception:
                pass
            self._ctx_menu = None

        self._ctx_menu_date = d0
        self._ctx_menu_cell = self._rmb_last_cell

        menu = QMenu(self)
        self._ctx_menu = menu

        menu.addAction("Открыть день…", lambda: self.main.open_day_dialog(d0))
        menu.addSeparator()
        menu.addAction("Добавить дежурство…", lambda: self.add_duty_dialog(d0))
        menu.addAction("Добавить компенсацию…", lambda: self.add_compensation_dialog(d0))
        menu.addSeparator()
        menu.addAction("Удалить дежурства за день", lambda: self.delete_duty_on_day(d0))
        menu.addAction("Удалить компенсации за день", lambda: self.delete_comp_on_day(d0))
        menu.addSeparator()

        is_work = True
        try:
            is_work = self.db.is_working_day(d0)
        except Exception:
            pass

        if is_work:
            self._add_colored_menu_action(menu, "Сделать нерабочим", "#c00000", lambda: self.toggle_working_day(d0))
        else:
            self._add_colored_menu_action(menu, "Сделать рабочим", "#008000", lambda: self.toggle_working_day(d0))

        # ПОКАЗЫВАЕМ МЕНЮ ТОЛЬКО В КОНЦЕ (важно)
        menu.popup(gp)

    def add_duty_dialog(self, d0: date) -> None:
        if not self.employee:
            self.main.toast("Выберите сотрудника", 2500)
            return
    
        # Предзаполняем по стандартам, но пользователь может всё изменить в окне
        start_t = self.main.defaults_widget.duty_start_time()
        dur_min = self.main.defaults_widget.duty_duration_minutes()
        start = datetime.combine(d0, start_t)
        end = start + timedelta(minutes=dur_min)
    
        dlg = DutyDialog(self.db, int(self.employee["id"]), parent=self)
        dlg.set_values(start, end, "", self.main.default_breaks_for_duty(start, end))
    
        if dlg.exec() != QDialog.Accepted:
            return
    
        s, e, comment = dlg.get_values()
        breaks = dlg.get_breaks()
        
        msg = validate_event_in_employee_period(self.employee, s.date())
        if msg:
            QMessageBox.warning(self, "Проверка", msg)
            return
    
        self.db.conn.execute("BEGIN;")
        try:
            duty_id = self.db.add_duty(int(self.employee["id"]), s, e, comment)
            self.db.replace_duty_breaks(duty_id, breaks)
    
            ok, msg2 = self.main.validate_employee_all_years(int(self.employee["id"]))
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
    
            self.db.conn.execute("COMMIT;")
            self.main.toast("Дежурство добавлено", 4000)
            self.main.refresh_context()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))
    
    
    def add_compensation_dialog(self, d0: date) -> None:
        if not self.employee:
            self.main.toast("Выберите сотрудника", 2500)
            return
    
        # Откроем полный диалог компенсации (выбор типа/способа/количества)
        y = self.year  # год текущего выбранного месяца
        dlg = CompensationDialog(self.db, self.employee, y, self.main.workday_minutes(), comp_id=None, parent=self)
        dlg.prefill_date(d0)
    
        # Предзаполним самый частый вариант: "Часы" + "Выходной" на выбранную дату
        try:
            dlg.cb_unit.setCurrentIndex(0)   # "hours"
            dlg.hours_amount.set_minutes_total(
                min(int(self.main.defaults_widget.comp_minutes()), int(self.main.workday_minutes()))
            )
            dlg.ed_comment.setText("")
            dlg.update_visibility()
        except Exception:
            pass
    
        if dlg.exec() == QDialog.Accepted:
            self.main.toast("Компенсация добавлена", 4000)
            self.main.refresh_context()
   
    def quick_add_default_duty(self, d: date) -> None:
        if not self.employee:
            return

        start_t = self.main.defaults_widget.duty_start_time()
        dur_min = self.main.defaults_widget.duty_duration_minutes()
        start = datetime.combine(d, start_t)
        end = start + timedelta(minutes=dur_min)

        msg = validate_event_in_employee_period(self.employee, start.date())
        if msg:
            QMessageBox.warning(self, "Проверка", msg)
            return

        self.db.conn.execute("BEGIN;")
        try:
            duty_id = self.db.add_duty(int(self.employee["id"]), start, end, "")
            self.db.replace_duty_breaks(duty_id, self.main.default_breaks_for_duty(start, end))
            ok, msg2 = self.main.validate_employee_all_years(int(self.employee["id"]))
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")
            self.main.toast("Дежурство добавлено", 4000)
            self.main.refresh_context()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    def quick_add_hour_compensation(self, d0: date) -> None:
        if not self.employee:
            return

        msg = validate_event_in_employee_period(self.employee, d0)
        if msg:
            QMessageBox.warning(self, "Проверка", msg)
            return

        workday = int(self.main.workday_minutes())

        r = self.db.conn.execute(
            """
            SELECT COALESCE(SUM(amount_minutes),0) AS m
            FROM compensation
            WHERE employee_id=?
              AND unit='hours' AND method='day_off'
              AND event_date=?
            """,
            (int(self.employee["id"]), d_iso(d0)),
        ).fetchone()
        already = int(r["m"])
        remaining = workday - already

        if remaining <= 0:
            self.main.toast(f"На {d0.strftime('%d.%m.%Y')} уже списано {fmt_minutes_ru_words(already)}", 5000)
            return

        minutes_default = int(self.main.defaults_widget.comp_minutes())
        if minutes_default <= 0:
            QMessageBox.warning(
                self,
                "Проверка",
                "Стандарт компенсации должен быть больше 0 (Сервис → Стандарты быстрых действий).",
            )
            return

        minutes = min(minutes_default, remaining, workday)
        if minutes <= 0:
            return

        self.db.conn.execute("BEGIN;")
        try:
            self.db.add_compensation_hours_dayoff(int(self.employee["id"]), d0, minutes, "")
            ok, msg2 = self.main.validate_employee_all_years(int(self.employee["id"]))
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg2)
                return
            self.db.conn.execute("COMMIT;")
            self.main.toast(
                f"Списано {fmt_minutes_ru_words(minutes)} на {d0.strftime('%d.%m.%Y')}",
                5000,
            )
            self.main.refresh_context()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    def _on_calendar_clipboard_action(self, d0: date, action: str) -> None:
        if d0.year != self.year or d0.month != self.month:
            self.main.toast("Дата вне выбранного месяца", 2500)
            return

    def on_calendar_clipboard_action(self, d0: date, action: str) -> None:
        self._on_calendar_clipboard_action(d0, action)

        if action == "copy":
            self.copy_duties_from_day(d0)
            return

        if action == "paste":
            self.paste_duties_to_day(d0)
            return

    def copy_duties_from_day(self, d0: date) -> None:
        if not self.employee:
            self.main.toast("Выберите сотрудника", 2500)
            return

        eid = int(self.employee["id"])

        s_day = datetime.combine(d0, time(0, 0))
        e_day = s_day + timedelta(days=1)

        rows_all = self.db.list_duties_for_period(eid, s_day, e_day)

        # Копируем только дежурства, которые НАЧИНАЮТСЯ в этот день
        rows = []
        for r in rows_all:
            try:
                if dt_parse(r["start_dt"]).date() == d0:
                    rows.append(r)
            except Exception:
                pass

        if not rows:
            self.main.toast("На эту дату нет дежурств для копирования", 3000)
            return

        duty_ids = [int(r["id"]) for r in rows]
        breaks_map = self.db.breaks_for_duty_ids(duty_ids)

        duties: list[dict[str, Any]] = []
        for r in rows:
            did = int(r["id"])
            duties.append(
                {
                    "start": dt_parse(r["start_dt"]),
                    "end": dt_parse(r["end_dt"]),
                    "comment": r["comment"] or "",
                    "breaks": breaks_map.get(did, []),
                }
            )

        self._duty_clipboard = {"src_date": d0, "duties": duties}
        self.main.toast(f"Скопировано дежурств: {len(duties)}", 3000)

    def paste_duties_to_day(self, target_date: date) -> None:
        if not self.employee:
            self.main.toast("Выберите сотрудника", 2500)
            return

        if not self._duty_clipboard:
            self.main.toast("Буфер пуст (нажмите Ctrl+C на дате с дежурствами)", 4000)
            return

        src_date = self._duty_clipboard.get("src_date")
        duties = self._duty_clipboard.get("duties") or []
        if not isinstance(src_date, date) or not duties:
            self.main.toast("Буфер пуст (нажмите Ctrl+C на дате с дежурствами)", 4000)
            return

        delta_days = (target_date - src_date).days
        delta = timedelta(days=int(delta_days))

        eid = int(self.employee["id"])

        self.db.conn.execute("BEGIN;")
        try:
            created = 0

            for d in duties:
                start0: datetime = d["start"]
                end0: datetime = d["end"]
                comment0: str = d.get("comment") or ""
                breaks0: list[tuple[datetime, datetime]] = d.get("breaks") or []

                start = start0 + delta
                end = end0 + delta
                breaks = [(bs + delta, be + delta) for bs, be in breaks0]

                # проверка периода сотрудника (и начало, и конец)
                msg = validate_event_in_employee_period(self.employee, start.date())
                if msg:
                    raise DBError(f"Дата {fmt_date_iso(d_iso(start.date()))}: {msg}")

                msg2 = validate_event_in_employee_period(self.employee, (end - timedelta(minutes=1)).date())
                if msg2:
                    raise DBError(f"Дата {fmt_date_iso(d_iso(end.date()))}: {msg2}")

                new_id = self.db.add_duty(eid, start, end, comment0)
                self.db.replace_duty_breaks(new_id, breaks)
                created += 1

            ok, msg3 = self.main.validate_employee_all_years(eid)
            if not ok:
                raise DBError(msg3)

            self.db.conn.execute("COMMIT;")
            self.main.toast(f"Вставлено дежурств: {created}", 4000)
            self.main.refresh_context()

        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.warning(self, "Вставка дежурств", str(e))

    def delete_duty_on_day(self, d0: date) -> None:
        if not self.employee:
            self.main.toast("Выберите сотрудника", 2500)
            return

        eid = int(self.employee["id"])
        s_day = datetime.combine(d0, time(0, 0))
        e_day = s_day + timedelta(days=1)

        rows = self.db.list_duties_for_period(eid, s_day, e_day)
        if not rows:
            self.main.toast("Дежурств на эту дату нет", 2500)
            return

        self.db.conn.execute("BEGIN;")
        try:
            for r in rows:
                self.db.delete_duty(int(r["id"]))

            ok, msg2 = self.main.validate_employee_all_years(eid)
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                self.main.toast(msg2, 7000)
                return

            self.db.conn.execute("COMMIT;")
            self.main.toast(f"Удалено дежурств за день: {len(rows)}", 4000)
            self.main.refresh_context()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    def delete_comp_on_day(self, d0: date) -> None:
        if not self.employee:
            self.main.toast("Выберите сотрудника", 2500)
            return

        eid = int(self.employee["id"])
        d_iso0 = d_iso(d0)

        rows = self.db.conn.execute(
            """
            SELECT * FROM compensation
            WHERE employee_id=?
              AND (
                    (event_date IS NOT NULL AND event_date = ?)
                    OR
                    (unit='days' AND method='day_off' AND id IN (
                        SELECT DISTINCT compensation_id FROM comp_day_off_date
                        WHERE employee_id=? AND day_off_date = ?
                    ))
              )
            ORDER BY id
            """,
            (eid, d_iso0, eid, d_iso0),
        ).fetchall()

        if not rows:
            self.main.toast("Компенсаций на эту дату нет", 2500)
            return

        removed = 0

        self.db.conn.execute("BEGIN;")
        try:
            for r in rows:
                cid = int(r["id"])
                unit = r["unit"]
                method = r["method"]

                if unit == "days" and method == "day_off":
                    dates = self.db.get_comp_dates(cid)
                    if d_iso0 in dates:
                        if len(dates) > 1:
                            new_dates = [x for x in dates if x != d_iso0]
                            self.db.replace_comp_dayoff_dates(cid, eid, [d_parse(x) for x in new_dates])
                            removed += 1
                        else:
                            self.db.delete_compensation(cid)
                            removed += 1
                else:
                    self.db.delete_compensation(cid)
                    removed += 1

            ok, msg2 = self.main.validate_employee_all_years(eid)
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                self.main.toast(msg2, 7000)
                return

            self.db.conn.execute("COMMIT;")
            self.main.toast(f"Удалено компенсаций за день: {removed}", 4000)
            self.main.refresh_context()
        except sqlite3.IntegrityError as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.warning(self, "Ошибка", f"Конфликт ограничений в базе:\n{e}")
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    def _add_colored_menu_action(self, menu: QMenu, text: str, color_hex: str, callback: Callable[[], None]) -> None:
        wa = QWidgetAction(menu)
        lbl = QLabel(text)
        lbl.setStyleSheet(f"color:{color_hex}; padding:4px 18px 4px 6px;")
        lbl.setCursor(Qt.PointingHandCursor)
        wa.setDefaultWidget(lbl)
        menu.addAction(wa)
    
        def _fire(_ev=None):
            menu.close()
            callback()
    
        lbl.mousePressEvent = lambda e: _fire() if e.button() == Qt.LeftButton else None 

    def toggle_working_day(self, d0: date) -> None:
        self.db.conn.execute("BEGIN;")
        try:
            self.db.toggle_calendar_day(d0)
    
            # Важно: смена рабочего/нерабочего может повлиять на баланс дней
            ok, msg = self.main.validate_all_employees_year(d0.year)
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                self.main.toast(msg, 7000)
                return
    
            self.db.conn.execute("COMMIT;")
            self.main.toast("День переключен", 2500)
            self.main.refresh_context()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))        
        
# -----------------------------
# Main window
# -----------------------------
class MainWindow(QMainWindow):
    def workday_minutes(self) -> int:
        ui = self.cfg.ui or {}
        v = safe_int(ui.get("workday_minutes"), 8 * 60)
        return max(1, int(v))

    def default_breaks_for_duty(self, duty_start: datetime, duty_end: datetime) -> list[tuple[datetime, datetime]]:
        specs = self.defaults_widget.default_break_times()

        out: list[tuple[datetime, datetime]] = []
        day0 = duty_start.date()

        for st, en in specs:
            s = datetime.combine(day0, st)

            # если начало получилось раньше начала дежурства — переносим на следующие сутки (и дальше, если нужно)
            while s < duty_start:
                s += timedelta(days=1)

            e = datetime.combine(s.date(), en)

            # если конец меньше/равен началу — значит конец на следующие сутки
            while e <= s:
                e += timedelta(days=1)

            # берем только перерывы, которые полностью внутри дежурства
            if s >= duty_start and e <= duty_end and s < e:
                out.append((s, e))

        out = merge_intervals(sorted(out, key=lambda x: x[0]))
        return out

    def open_workday_settings(self) -> None:
        dlg = QDialog(self)
        dlg.setWindowTitle("Длина рабочего дня")
        lay = QVBoxLayout(dlg)

        lay.addWidget(QLabel("Укажите длительность рабочего дня (максимум списания часов в одну дату)."))

        picker = DurationPicker(max_hours=24)
        picker.set_minutes_total(self.workday_minutes())
        lay.addWidget(picker)

        bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        lay.addWidget(bb)
        bb.accepted.connect(dlg.accept)
        bb.rejected.connect(dlg.reject)

        if dlg.exec() != QDialog.Accepted:
            return

        minutes = int(picker.minutes_total())
        if minutes <= 0:
            QMessageBox.warning(self, "Проверка", "Длина рабочего дня должна быть больше 0.")
            return

        self.cfg.ui = self.cfg.ui or {}
        self.cfg.ui["workday_minutes"] = minutes
        self.cfg.save()

        self.toast(f"Длина рабочего дня: {fmt_minutes_ru_words(minutes)}", 4000)    
    
    def add_money_compensation(self) -> None:
        emp = self.current_employee()
        if not emp:
            self.toast("Выберите сотрудника", 2500)
            return
        y, _ = self.current_year_month()
        dlg = MoneyCompensationDialog(self.db, emp, default_year=y, comp_id=None, parent=self)
        if dlg.exec() == QDialog.Accepted:
            self.toast("Сохранено", 4000)
            self.refresh_context()

    def export_template(self) -> None:
        y, m = self.current_year_month()
    
        # Шаблон: всегда рядом с запускаемым файлом, имя Template.xlsx
        base_dir = Path(sys.argv[0]).resolve().parent
        tpl_path = base_dir / "Template.xlsx"
    
        if not tpl_path.exists():
            QMessageBox.critical(
                self,
                "Шаблон не найден",
                f"Не найден файл шаблона:\n{tpl_path}\n\n"
                f"Положите Template.xlsx в папку с программой.",
            )
            return
    
        out, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт по шаблону",
            f"табель_{y:04d}-{m:02d}.xlsx",
            "Excel (*.xlsx)",
        )
        if not out:
            return
        if not out.endswith(".xlsx"):
            out += ".xlsx"
    
        try:
            TemplateExporter.export(self.db, y, m, str(tpl_path), out, sheet_name="Лист1")
            self.toast("Экспорт по шаблону готов", 5000)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def preview_month(self) -> None:
        y, m = self.current_year_month()

        base_dir = Path(sys.argv[0]).resolve().parent
        tpl_path = base_dir / "Template.xlsx"
        if not tpl_path.exists():
            QMessageBox.critical(self, "Шаблон не найден", f"Не найден файл:\n{tpl_path}")
            return

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = str(Path(td) / f"табель_{y:04d}-{m:02d}.xlsx")
            pdf_path = str(Path(td) / f"табель_{y:04d}-{m:02d}.pdf")

            try:
                TemplateExporter.export(self.db, y, m, str(tpl_path), xlsx_path, sheet_name="Лист1")
                convert_xlsx_to_pdf_auto(xlsx_path, pdf_path)
                PdfPreviewDialog(pdf_path, parent=self).exec()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def _print_pdf_dialog(self, pdf_path: str) -> None:
        printer = QPrinter(QPrinter.HighResolution)
        dlg = QPrintDialog(printer, self)
        if dlg.exec() != QDialog.Accepted:
            return
    
        doc = QPdfDocument(self)
        err = doc.load(pdf_path)
        if err != QPdfDocument.Error.None_:
            raise DBError(f"Не удалось открыть PDF для печати: {err}")
    
        paint_pdf_to_printer(doc, printer)

    def print_month(self) -> None:
        y, m = self.current_year_month()

        base_dir = Path(sys.argv[0]).resolve().parent
        tpl_path = base_dir / "Template.xlsx"
        if not tpl_path.exists():
            QMessageBox.critical(self, "Шаблон не найден", f"Не найден файл:\n{tpl_path}")
            return

        with tempfile.TemporaryDirectory() as td:
            xlsx_path = str(Path(td) / f"табель_{y:04d}-{m:02d}.xlsx")
            pdf_path = str(Path(td) / f"табель_{y:04d}-{m:02d}.pdf")

            try:
                TemplateExporter.export(self.db, y, m, str(tpl_path), xlsx_path, sheet_name="Лист1")
                convert_xlsx_to_pdf_auto(xlsx_path, pdf_path)
                self._print_pdf_dialog(pdf_path)
            except Exception as e:
                QMessageBox.critical(self, "Ошибка печати", str(e))
        
    def __init__(self, cfg: AppConfig):
        super().__init__()
        # Указываем всей программе, что главные кнопки должны краситься в синий
        QApplication.instance().setStyleSheet(QApplication.instance().styleSheet() + 
            "\nQDialogButtonBox QPushButton[default=\"true\"] { background-color: #0A84FF; color: white; font-weight: bold; }"
            "\nQDialogButtonBox QPushButton[default=\"true\"]:hover { background-color: #0070E0; }"
        )
        
        self.cfg = cfg
        self.db: Optional[DB] = None
        self._work_ui_ready = False
        self._declined_calendar_years: set[int] = set()

        self.setWindowTitle("OvertimeTab")
        self.resize(900, 600)

        self.snackbar = Snackbar(self)

        self._init_welcome_ui()

        QTimer.singleShot(3000, self._boot_after_welcome)
    
    def _init_welcome_ui(self) -> None:
        menubar = QMenuBar(self)
        self.setMenuBar(menubar)

        m_file = menubar.addMenu("Файл")
        m_file.addAction("Выход").triggered.connect(self.close)

        w = WelcomeWidget(self)
        self.setCentralWidget(w)

    def _boot_after_welcome(self) -> None:
        # подчистим несуществующие пути
        self.cfg.db_paths = [p for p in (self.cfg.db_paths or []) if p and os.path.exists(p)]
        if self.cfg.last_db_path and (not os.path.exists(self.cfg.last_db_path)):
            self.cfg.last_db_path = None
        self.cfg.save()

        # 1) пробуем открыть последнюю базу
        if self.cfg.last_db_path:
            try:
                self.open_database(self.cfg.last_db_path)
                return
            except Exception:
                pass

        # 2) если баз нет вообще -> создаём первую
        if not self.cfg.db_paths:
            self._create_first_database()
            return

        # 3) если базы есть, но последняя не открылась -> пусть выберут (один раз на старте)
        self._open_base_manager_startup()

    def _open_base_manager_startup(self) -> None:
        dlg = BaseSelectorDialog(self.cfg, parent=self)
        if dlg.exec() != QDialog.Accepted or not dlg.selected_path:
            self.close()
            return
        try:
            self.open_database(dlg.selected_path)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))
            self.close()

    def _create_first_database(self) -> None:
        dlg = CreateSubdivisionDialog(parent=self)
        if dlg.exec() != QDialog.Accepted:
            self.close()
            return

        v = dlg.values()
        name = v["department_name"]

        # 1) пытаемся создать рядом с программой (в подпапке databases)
        preferred_dir = program_base_dir() / "databases"
        fallback_used = False

        if not can_write_dir(preferred_dir):
            fallback_used = True
            preferred_dir = app_db_store_dir()

        db_path = make_unique_db_path(preferred_dir, name)

        # создаём базу
        db = DB(str(db_path))
        try:
            db.update_department_settings(
                department_name=name,
                resp_position=v["resp_position"],
                resp_rank=v["resp_rank"],
                resp_last_name=v["resp_last_name"],
                resp_first_name=v["resp_first_name"],
                resp_middle_name=v["resp_middle_name"],
            )
            db.conn.commit()
        finally:
            db.close()

        # сохраняем в конфиг
        p = str(db_path)
        if p not in self.cfg.db_paths:
            self.cfg.db_paths.append(p)
        self.cfg.last_db_path = p
        self.cfg.save()

        if fallback_used:
            QMessageBox.information(
                self,
                "База создана",
                "Нет прав на запись в папку программы.\n"
                "База создана в папке данных пользователя:\n\n"
                f"{p}",
            )

        self.open_database(p)

    def open_database(self, path: str) -> None:
        # закрываем старую базу
        if self.db is not None:
            try:
                self.db.close()
            except Exception:
                pass
            self.db = None

        self.db = DB(path)

        # автозаполнение календаря на ±10 лет (не затирает существующее)
        y0 = date.today().year
        try:
            self.db.conn.execute("BEGIN;")
            for y in range(y0 - 10, y0 + 11):
                self.db.ensure_calendar_year_default(y)
            self.db.conn.execute("COMMIT;")
        except Exception:
            try:
                self.db.conn.execute("ROLLBACK;")
            except Exception:
                pass

        # обновляем конфиг
        if path not in self.cfg.db_paths:
            self.cfg.db_paths.append(path)
        self.cfg.last_db_path = path
        self.cfg.save()

        # если рабочий UI уже был — пересоздадим его целиком (проще и надежнее)
        if self._work_ui_ready:
            old = self.centralWidget()
            if old is not None:
                old.deleteLater()
            self._work_ui_ready = False

        self._init_work_ui()
        self.refresh_all()

    def _init_work_ui(self) -> None:
        if self.db is None:
            raise DBError("База не открыта.")

        self._work_ui_ready = True

        self.setWindowTitle(f"{self.db.get_department_name()} — учет времени")
        self.resize(1200, 800)

        menubar = QMenuBar(self)
        self.setMenuBar(menubar)

        # ---- menu: File, View, Service (Оставляем как было) ----
        m_file = menubar.addMenu("Файл")
        m_file.addAction("Печать…").triggered.connect(self.print_month)
        m_file.addAction("Предпросмотр печати…").triggered.connect(self.preview_month)
        m_file.addAction("Экспорт (Excel, по шаблону)…").triggered.connect(self.export_template)

        m_view = menubar.addMenu("Вид")
        m_theme = m_view.addMenu("Тема")
        self.act_theme_system = QAction("Системная", self, checkable=True)
        self.act_theme_light = QAction("Светлая", self, checkable=True)
        self.act_theme_dark = QAction("Тёмная", self, checkable=True)
        self.theme_group = [self.act_theme_system, self.act_theme_light, self.act_theme_dark]
        for a in self.theme_group:
            m_theme.addAction(a)
            a.triggered.connect(self.on_theme_changed)

        m_service = menubar.addMenu("Сервис")
        m_service.addAction("Базы подразделений…").triggered.connect(self.open_bases_manager)
        m_service.addSeparator()
        m_service.addAction("Стандарты быстрых действий…").triggered.connect(self.open_quick_defaults)
        m_service.addAction("Настройки подразделения").triggered.connect(self.open_settings)
        m_service.addAction("Длина рабочего дня…").triggered.connect(self.open_workday_settings)

        splitter = QSplitter()
        # Прячем ползунок сплиттера
        splitter.setStyleSheet("QSplitter::handle { background-color: #2C2C2E; width: 1px; }")
        self.splitter = splitter
        self.setCentralWidget(splitter)

        # ==========================================
        # ЛЕВАЯ ПАНЕЛЬ (САЙДБАР)
        # ==========================================
        left_frame = QFrame()
        left_frame.setObjectName("Sidebar")
        # Фон сайдбара темнее основного окна (классика macOS)
        left_frame.setStyleSheet("QFrame#Sidebar { background-color: #161618; border: none; }")
        left_lay = QVBoxLayout(left_frame)
        left_lay.setContentsMargins(15, 20, 15, 20)
        left_lay.setSpacing(15)

        # Скрытые комбобоксы
        self.cb_year = QComboBox(); self.cb_year.setVisible(False)
        for y in range(2000, 2100): self.cb_year.addItem(str(y), y)
        self.cb_month = QComboBox(); self.cb_month.setVisible(False)
        for m in range(1, 13): self.cb_month.addItem(f"{m:02d}", m)

        # Название подразделения
        self.lbl_subdivision = QLabel(self.db.get_department_name())
        self.lbl_subdivision.setWordWrap(True)
        self.lbl_subdivision.setStyleSheet("font-size: 18px; font-weight: 800; color: #FFFFFF;")
        left_lay.addWidget(self.lbl_subdivision)

        # Поиск и фильтр
        search_row = QHBoxLayout()
        self.ed_search = QLineEdit()
        self.ed_search.setPlaceholderText("🔍 Поиск...")
        self.ed_search.setClearButtonEnabled(True)        
        self.cb_active = QComboBox()
        self.cb_active.addItem("Активные", True)
        self.cb_active.addItem("Все", False)
        search_row.addWidget(self.ed_search, 1)
        search_row.addWidget(self.cb_active, 0)
        left_lay.addLayout(search_row)

        # Список сотрудников
        self.list_emp = QListWidget()
        self.list_emp.setStyleSheet("""
            QListWidget { background: transparent; border: none; outline: none; }
            QListWidget::item { padding: 8px; border-radius: 6px; margin-bottom: 2px; }
            QListWidget::item:selected { background-color: #2C2C2E; color: #FFFFFF; }
        """)
        self.list_emp.setContextMenuPolicy(Qt.CustomContextMenu)
        self.list_emp.customContextMenuRequested.connect(self.on_emp_context_menu)
        left_lay.addWidget(self.list_emp, 1)

        # Единственная кнопка внизу (Стиль: Dash border)
        self.btn_add = QPushButton("+ Добавить сотрудника")
        self.btn_add.setStyleSheet("""
            QPushButton {
                background: transparent; color: #0A84FF; font-weight: 600;
                border: 1px dashed #3A3A3C; border-radius: 6px; padding: 10px;
            }
            QPushButton:hover { background: #2C2C2E; border: 1px solid #0A84FF; }
        """)
        self.btn_add.setCursor(Qt.PointingHandCursor)
        left_lay.addWidget(self.btn_add)

        splitter.addWidget(left_frame)

        # ==========================================
        # ЦЕНТРАЛЬНАЯ ПАНЕЛЬ
        # ==========================================
        # ==========================================
        # ЦЕНТРАЛЬНАЯ ПАНЕЛЬ (С ЭКРАНОМ ПУСТОТЫ)
        # ==========================================
        self.defaults_widget = QuickDefaultsWidget(self.cfg)

        # Переключатель: Экран Пустоты ИЛИ Календарь
        self.right_stack = QStackedWidget()

        # 1. Страница: Никто не выбран (Empty State)
        self.empty_page = QWidget()
        empty_lay = QVBoxLayout(self.empty_page)
        empty_icon = QLabel("👤")
        empty_icon.setStyleSheet("font-size: 80px; color: #3A3A3C;")
        empty_text = QLabel("Выберите сотрудника в списке слева\nили добавьте нового")
        empty_text.setStyleSheet("font-size: 16px; color: #8E8E93;")
        empty_text.setAlignment(Qt.AlignCenter)
        empty_icon.setAlignment(Qt.AlignCenter)
        empty_lay.addStretch()
        empty_lay.addWidget(empty_icon)
        empty_lay.addWidget(empty_text)
        empty_lay.addStretch()
        self.right_stack.addWidget(self.empty_page)

        # 2. Страница: Рабочая зона (Календарь и Итоги)
        self.work_page = QWidget()
        center_lay = QVBoxLayout(self.work_page)
        center_lay.setContentsMargins(30, 20, 30, 20)
        center_lay.setSpacing(20)

        self.calendar_panel = CalendarPanel(self.db, self)
        self.calendar_panel.header_layout.addStretch(1)
        
        self.btn_print_top = QPushButton("🖨 Печать")
        self.btn_preview_top = QPushButton("Предпросмотр")
        self.btn_export_top = QPushButton("📤 Экспорт")
        
        for b in (self.btn_print_top, self.btn_preview_top, self.btn_export_top):
            b.setStyleSheet("background: transparent; color: #8E8E93; font-weight: 600; font-size: 13px;")
            b.setCursor(Qt.PointingHandCursor)
            self.calendar_panel.header_layout.addWidget(b)

        self.month_summary = MonthSummaryPanel(self.db, self)

        center_lay.addWidget(self.calendar_panel, 3)
        center_lay.addWidget(self.month_summary, 1)
        self.right_stack.addWidget(self.work_page)

        splitter.addWidget(self.right_stack)

        # Пропорции (Сайдбар уже, контент шире)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 4) 

        # Сигналы (Кнопки "Редактировать", "Удалить" и тд удалены из UI, 
        # но их логика осталась в методах и вызывается через контекстное меню)
        self.cb_year.currentIndexChanged.connect(self.on_period_changed)
        self.cb_month.currentIndexChanged.connect(self.on_period_changed)
        self.cb_active.currentIndexChanged.connect(self.refresh_employees)
        self.ed_search.textChanged.connect(self.refresh_employees)
        self.list_emp.currentItemChanged.connect(self.on_emp_selected)
        
        self.btn_print_top.clicked.connect(self.print_month)
        self.btn_preview_top.clicked.connect(self.preview_month)
        self.btn_export_top.clicked.connect(self.export_template)
        self.btn_add.clicked.connect(self.add_employee)

        self.restore_ui_state()
        self.on_period_changed()

    def open_bases_manager(self) -> None:
        dlg = BaseSelectorDialog(self.cfg, parent=self)
        if dlg.exec() != QDialog.Accepted or not dlg.selected_path:
            return
        self.open_database(dlg.selected_path)
        self.toast("Подразделение переключено", 4000)        
    
    def undo_end_reason(self) -> None:
        eid = self.selected_employee_id()
        if eid is None:
            return

        emp = self.db.get_employee(eid)
        if (not emp["end_date"]) or (emp["end_reason"] not in ("transfer", "dismissal")):
            return

        action_word = "перевод" if emp["end_reason"] == "transfer" else "увольнение"
        if QMessageBox.question(
            self,
            "Отмена статуса",
            f"Отменить {action_word} (снять дату {fmt_date_iso(emp['end_date'])})?",
        ) != QMessageBox.Yes:
            return

        self.db.conn.execute("BEGIN;")
        try:
            self.db.update_employee(eid, end_date=None, end_reason=None)
            self.db.conn.execute("COMMIT;")
            self.toast("Статус снят", 4000)
            self.refresh_all()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))
    
    # ---- snackbar helpers ----
    def toast(self, text: str, duration_ms: int = 4000) -> None:
        self.snackbar.show_message(text, duration_ms=duration_ms)
        self._reposition_snackbar()

    def resizeEvent(self, event) -> None:  # type: ignore[override]
        super().resizeEvent(event)
        self._reposition_snackbar()

    def _reposition_snackbar(self) -> None:
        if not self.snackbar.isVisible():
            return
        w = min(720, self.width() - 40)
        self.snackbar.resize(w, self.snackbar.height())
        self.snackbar.move((self.width() - w) // 2, self.height() - self.snackbar.height() - 20)

    # ---- UI state ----
    def restore_ui_state(self) -> None:
        ui = self.cfg.ui or {}

        theme = (ui.get("theme") or "system").lower()
        self.set_theme_checked(theme)
        apply_theme(QApplication.instance(), theme)

        g = ui.get("main_geometry")
        if isinstance(g, str) and g:
            try:
                ba = QByteArray.fromBase64(g.encode("ascii"))
                self.restoreGeometry(ba)
            except Exception:
                pass

        sizes = ui.get("splitter_sizes")
        if isinstance(sizes, list) and len(sizes) >= 2:
            try:
                self.splitter.setSizes([int(sizes[0]), int(sizes[1])])
            except Exception:
                pass

        y = safe_int(ui.get("year"), date.today().year)
        m = safe_int(ui.get("month"), date.today().month)
        yi = self.cb_year.findData(y)
        if yi >= 0:
            self.cb_year.setCurrentIndex(yi)
        self.cb_month.setCurrentIndex(max(0, min(11, m - 1)))

        active_only = ui.get("active_only", True)
        self.cb_active.setCurrentIndex(0 if active_only else 1)
        self.ed_search.setText(ui.get("search", "") or "")

    def save_ui_state(self) -> None:
        ui = self.cfg.ui or {}
        ui["theme"] = self.current_theme()
        ui["main_geometry"] = bytes(self.saveGeometry().toBase64()).decode("ascii")
        ui["splitter_sizes"] = self.splitter.sizes()

        y, m = self.current_year_month()
        ui["year"] = y
        ui["month"] = m
        ui["active_only"] = bool(self.cb_active.currentData())
        ui["search"] = self.ed_search.text()

        eid = self.selected_employee_id()
        ui["selected_employee_id"] = eid

        self.cfg.ui = ui
        self.cfg.save()

    def closeEvent(self, event) -> None:  # type: ignore[override]
        if self._work_ui_ready:
            self.save_ui_state()
        super().closeEvent(event)

    # ---- theme menu ----
    def set_theme_checked(self, theme: str) -> None:
        theme = (theme or "system").lower()
        for a in self.theme_group:
            a.setChecked(False)
        if theme == "dark":
            self.act_theme_dark.setChecked(True)
        elif theme == "light":
            self.act_theme_light.setChecked(True)
        else:
            self.act_theme_system.setChecked(True)

    def current_theme(self) -> str:
        if self.act_theme_dark.isChecked():
            return "dark"
        if self.act_theme_light.isChecked():
            return "light"
        return "system"

    def on_theme_changed(self) -> None:
        sender = self.sender()
        if sender in self.theme_group:
            for a in self.theme_group:
                if a is not sender:
                    a.setChecked(False)
            if not sender.isChecked():
                sender.setChecked(True)

        theme = self.current_theme()
        apply_theme(QApplication.instance(), theme)
        self.cfg.ui = self.cfg.ui or {}
        self.cfg.ui["theme"] = theme
        self.cfg.save()

    # ---- period helpers ----
    def current_year_month(self) -> tuple[int, int]:
        return int(self.cb_year.currentData()), int(self.cb_month.currentData())

    def set_period(self, year: int, month: int) -> None:
        yi = self.cb_year.findData(year)
        if yi >= 0:
            self.cb_year.setCurrentIndex(yi)
        self.cb_month.setCurrentIndex(max(0, min(11, month - 1)))

    def shift_month(self, delta: int) -> None:
        y, m = self.current_year_month()
        if delta < 0:
            if m == 1:
                y -= 1
                m = 12
            else:
                m -= 1
        else:
            if m == 12:
                y += 1
                m = 1
            else:
                m += 1
        self.set_period(y, m)

    # ---- selection ----
    def selected_employee_id(self) -> Optional[int]:
        it = self.list_emp.currentItem()
        return int(it.data(Qt.UserRole)) if it else None

    def current_employee(self) -> Optional[sqlite3.Row]:
        eid = self.selected_employee_id()
        if eid is None:
            return None
        return self.db.get_employee(eid)

    def on_period_changed(self) -> None:
        y, _ = self.current_year_month()
        if (not self.db.calendar_year_complete(y)) and (y not in self._declined_calendar_years):
            r = QMessageBox.question(
                self,
                "Календарь",
                f"Календарь на {y} год не создан (или неполный).\n"
                f"Создать календарь по умолчанию (Пн–Пт рабочие, Сб–Вс выходные)?\n\n"
                f"Если выбрать 'Нет', операции, зависящие от календаря (дни/выходные), будут недоступны.",
                QMessageBox.Yes | QMessageBox.No,
            )
            if r == QMessageBox.No:
                self._declined_calendar_years.add(y)
            else:
                self.db.conn.execute("BEGIN;")
                try:
                    self.db.create_calendar_year_default(y)
                    ok_all = self.validate_all_employees_year(y)
                    if not ok_all[0]:
                        self.db.conn.execute("ROLLBACK;")
                        QMessageBox.warning(self, "Недостаточно", ok_all[1])
                    else:
                        self.db.conn.execute("COMMIT;")
                        self.toast("Календарь создан. Проверьте праздники/переносы в 'Сервис → Календарь'.", 6000)
                except Exception as e:
                    self.db.conn.execute("ROLLBACK;")
                    QMessageBox.critical(self, "Ошибка", str(e))

        self.refresh_employees()

    def refresh_all(self) -> None:
        if self.db is None:
            return

        name = self.db.get_department_name()
        self.setWindowTitle(f"{name} — учет времени")

        if hasattr(self, "lbl_subdivision") and self.lbl_subdivision is not None:
            self.lbl_subdivision.setText(name)

        self.refresh_employees()
        self.refresh_context()

    def refresh_context(self) -> None:
        y, m = self.current_year_month()
        emp = self.current_employee()
        
        # Если сотрудник не выбран — показываем красивую заглушку
        if emp is None:
            self.right_stack.setCurrentIndex(0)
        else:
            # Если выбран — показываем рабочую зону
            self.right_stack.setCurrentIndex(1)
            self.calendar_panel.set_context(emp, y, m)
            self.month_summary.set_context(emp, y, m)

    def refresh_employees(self) -> None:
        y, m = self.current_year_month()
        active_only = bool(self.cb_active.currentData())

        # ИЩЕМ по тому тексту, который реально показывается в списке (ФИО/звание/должность/статус/дата),
        # чтобы работало “в любой части текста” и нормально работало с русскими буквами.
        query = (self.ed_search.text() or "").strip()
        needle = " ".join(query.split()).casefold()

        # ВАЖНО: из БД берем без SQL-поиска (SQLite lower() плохо работает с кириллицей),
        # фильтруем уже в Python по отображаемому тексту.
        emps = self.db.list_employees_for_month(y, m, active_only=active_only, search="")

        current_id = self.selected_employee_id()

        self.list_emp.blockSignals(True)
        self.list_emp.clear()

        for e in emps:
            fio = f"{e['last_name']} {e['first_name']} {e['middle_name'] or ''}".strip()
            rank = (e["rank"] or "").strip()
            pos = (e["position"] or "").strip()
            sub = " — ".join([x for x in [rank, pos] if x])

            active_in_month = is_emp_active_in_month(e, y, m)
            status = ""
            if (not active_in_month) and e["end_date"]:
                if e["end_reason"] == "transfer":
                    status = f"(переведен {fmt_date_iso(e['end_date'])})"
                elif e["end_reason"] == "dismissal":
                    status = f"(уволен {fmt_date_iso(e['end_date'])})"
                else:
                    status = f"(неактивен с {fmt_date_iso(e['end_date'])})"

            sub2 = (sub + " " + status).strip() if status else sub

            display_text = f"{fio}\n{sub2}".strip()
            hay = " ".join(display_text.split()).casefold()

            if needle and (needle not in hay):
                continue

            it = QListWidgetItem(display_text)
            it.setData(Qt.UserRole, int(e["id"]))
            
            # --- ВКЛЮЧАЕМ АВАТАРКУ ---
            it.setIcon(create_avatar_icon(fio))

            if (not active_only) and (not active_in_month):
                it.setForeground(QBrush(QColor("#808080")))

            self.list_emp.addItem(it)

        self.list_emp.blockSignals(False)

        preferred = current_id
        if preferred is None:
            preferred = (self.cfg.ui or {}).get("selected_employee_id")

        if preferred is not None:
            for i in range(self.list_emp.count()):
                if self.list_emp.item(i).data(Qt.UserRole) == preferred:
                    self.list_emp.setCurrentRow(i)
                    break

        if self.list_emp.currentItem() is None and self.list_emp.count() > 0:
            self.list_emp.setCurrentRow(0)
        else:
            self.on_emp_selected(self.list_emp.currentItem(), None)

    def on_emp_selected(self, cur: QListWidgetItem, prev: QListWidgetItem) -> None:
        self.refresh_context()

    # ---- context menu employees ----
    def on_emp_context_menu(self, pos: QPoint) -> None:
        menu = QMenu(self)
        eid = self.selected_employee_id()

        if eid is not None:
            menu.addAction("Редактировать", self.edit_employee)
            menu.addSeparator()
            menu.addAction("Переведен…", lambda: self.set_end_reason("transfer"))
            menu.addAction("Уволен…", lambda: self.set_end_reason("dismissal"))

            emp = self.db.get_employee(eid)
            if emp["end_date"] and emp["end_reason"] in ("transfer", "dismissal"):
                menu.addSeparator()
                if emp["end_reason"] == "transfer":
                    menu.addAction("Отменить перевод", self.undo_end_reason)
                else:
                    menu.addAction("Отменить увольнение", self.undo_end_reason)

            menu.addSeparator()
            menu.addAction("Удалить", self.delete_employee)

        menu.exec(self.list_emp.mapToGlobal(pos))

    # ---- service actions ----
    def open_calendar(self) -> None:
        CalendarDialog(self.db, self, self).exec()

    def open_settings(self) -> None:
        if DepartmentSettingsDialog(self.db, self).exec() == QDialog.Accepted:
            self.refresh_all()

    def open_quick_defaults(self) -> None:
        dlg = QDialog(self)
        dlg.setWindowTitle("Стандарты быстрых действий")
        lay = QVBoxLayout(dlg)
        w = QuickDefaultsWidget(self.cfg)
        lay.addWidget(w)
        bb = QDialogButtonBox(QDialogButtonBox.Close)
        bb.rejected.connect(dlg.reject)
        bb.accepted.connect(dlg.accept)
        lay.addWidget(bb)
        dlg.exec()
        # обновим "внутренний" defaults_widget (на случай изменения)
        try:
            self.defaults_widget._load()  # type: ignore[attr-defined]
        except Exception:
            pass

    def open_day_dialog(self, d0: date) -> None:
        emp = self.current_employee()
        if not emp:
            self.toast("Выберите сотрудника", 2500)
            return
        y, m = self.current_year_month()
        if d0.year != y or d0.month != m:
            self.toast("Дата вне выбранного месяца", 2500)
            return
        DayDialog(self.db, self, emp, y, d0, parent=self).exec()
        self.refresh_context()

    # ---- validation ----
    def validate_employee_all_years(self, employee_id: int) -> tuple[bool, str]:
        years = self.db.employee_data_years(employee_id)
        if not years:
            y, _ = self.current_year_month()
            years = {y}
        for yr in sorted(years):
            ok, msg = validate_non_negative_over_year(self.db, employee_id, yr)
            if not ok:
                return False, msg
        return True, ""

    def validate_all_employees_year(self, year: int) -> tuple[bool, str]:
        rows = self.db.conn.execute("SELECT id FROM employee").fetchall()
        for r in rows:
            ok, msg = validate_non_negative_over_year(self.db, int(r["id"]), year)
            if not ok:
                return False, msg
        return True, ""

    # ---- employee actions ----
    def add_employee(self) -> None:
        y, m = self.current_year_month()
        dlg = EmployeeDialog("Добавить сотрудника", default_year=y, default_month=m, parent=self)
        if dlg.exec() != QDialog.Accepted:
            return
        v = dlg.get_values()
        self.db.conn.execute("BEGIN;")
        try:
            self.db.add_employee(
                v["last_name"],
                v["first_name"],
                v["middle_name"],
                v["rank"],
                v["position"],
                v["start_month"],
                v["opening_minutes"],
                v["opening_days"],
            )
            self.db.conn.execute("COMMIT;")
            self.toast("Сотрудник добавлен")
            self.refresh_employees()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    def edit_employee(self) -> None:
        eid = self.selected_employee_id()
        if eid is None:
            return
        emp = self.db.get_employee(eid)
        y, m = self.current_year_month()
        dlg = EmployeeDialog("Редактировать сотрудника", default_year=y, default_month=m, parent=self)
        dlg.set_from_employee(emp)
        if dlg.exec() != QDialog.Accepted:
            return
        v = dlg.get_values()

        self.db.conn.execute("BEGIN;")
        try:
            self.db.update_employee(
                eid,
                last_name=v["last_name"],
                first_name=v["first_name"],
                middle_name=v["middle_name"] or None,
                rank=v["rank"] or None,
                position=v["position"] or None,
                start_month=v["start_month"],
                opening_minutes=v["opening_minutes"],
                opening_days=v["opening_days"],
            )
            ok, msg = self.validate_employee_all_years(eid)
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg)
                return
            self.db.conn.execute("COMMIT;")
            self.toast("Сохранено")
            self.refresh_all()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    def delete_employee(self) -> None:
        eid = self.selected_employee_id()
        if eid is None:
            return
        emp = self.db.get_employee(eid)
        fio = f"{emp['last_name']} {emp['first_name']} {emp['middle_name'] or ''}".strip()
        if QMessageBox.question(self, "Удаление", f"Удалить сотрудника '{fio}'? Будут удалены дежурства и компенсации.") != QMessageBox.Yes:
            return
        self.db.conn.execute("BEGIN;")
        try:
            self.db.delete_employee(eid)
            self.db.conn.execute("COMMIT;")
            self.toast("Сотрудник удален")
            self.refresh_all()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    def set_end_reason(self, reason: str) -> None:
        eid = self.selected_employee_id()
        if eid is None:
            return
        title = "Переведен" if reason == "transfer" else "Уволен"
        dlg = EndDateDialog(title, self)
        if dlg.exec() != QDialog.Accepted:
            return
        d0 = dlg.get_date()
        emp = self.db.get_employee(eid)
        start, _ = employee_period_bounds(emp)
        if d0 < start:
            QMessageBox.warning(self, "Проверка", "Дата раньше начала работы в отделе.")
            return
        if self.db.has_records_after_date(eid, d0):
            QMessageBox.warning(
                self,
                "Проверка",
                "Нельзя установить дату перевода/увольнения: у сотрудника есть дежурства или компенсации после этой даты.\n"
                "Сначала исправьте записи после указанной даты.",
            )
            return

        self.db.conn.execute("BEGIN;")
        try:
            self.db.update_employee(eid, end_date=d_iso(d0), end_reason=reason)
            ok, msg = self.validate_employee_all_years(eid)
            if not ok:
                self.db.conn.execute("ROLLBACK;")
                QMessageBox.warning(self, "Недостаточно", msg)
                return
            self.db.conn.execute("COMMIT;")
            self.toast("Сохранено")
            self.refresh_all()
        except Exception as e:
            self.db.conn.execute("ROLLBACK;")
            QMessageBox.critical(self, "Ошибка", str(e))

    # ---- export ----
    def export_simple(self) -> None:
        eid = self.selected_employee_id()
        if eid is None:
            QMessageBox.information(self, "Экспорт", "Выберите сотрудника.")
            return
        y, m = self.current_year_month()
        out, _ = QFileDialog.getSaveFileName(self, "Экспорт (Excel)", f"export_{y:04d}-{m:02d}.xlsx", "Excel (*.xlsx)")
        if not out:
            return
        if not out.endswith(".xlsx"):
            out += ".xlsx"
        try:
            export_simple_xlsx(self.db, eid, y, m, out)
            self.toast("Экспорт готов")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))
            
    def add_money_compensation_from_summary(self) -> None:
        emp = self.current_employee()
        if not emp:
            self.toast("Выберите сотрудника", 2500)
            return
        y, _ = self.current_year_month()
        dlg = MoneyCompensationDialog(self.db, emp, default_year=y, comp_id=None, parent=self)
        if dlg.exec() == QDialog.Accepted:
            self.toast("Сохранено", 4000)
            self.refresh_context()

    def open_money_comp_list(self) -> None:
        emp = self.current_employee()
        if not emp:
            self.toast("Выберите сотрудника", 2500)
            return
        MoneyCompListDialog(self.db, self, emp, parent=self).exec()
        self.refresh_context()


# -----------------------------
# App entry
# -----------------------------
def main() -> None:
    app = QApplication(sys.argv)

    cfg = AppConfig.load()
    theme = (cfg.ui or {}).get("theme", "system")
    apply_theme(app, theme)

    win = MainWindow(cfg)
    win.show()

    rc = app.exec()

    try:
        if win.db is not None:
            win.db.close()
    except Exception:
        pass

    sys.exit(rc)


if __name__ == "__main__":
    main()