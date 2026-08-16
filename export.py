import sqlite3
from datetime import date, datetime, time, timedelta
from typing import Optional
import io
from PySide6.QtCore import QFile, QIODevice

from utils import next_month, d_iso, d_parse, fmt_date_iso, intersect, merge_intervals, subtract_intervals
from logic import compute_month_summary, is_employee_shift

def ensure_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except Exception as e:
        raise Exception("Для экспорта нужен пакет openpyxl. Установите: pip install openpyxl") from e

def fio_initials_last(last_name: str | None, first_name: str | None, middle_name: str | None) -> str:
    last = (last_name or "").strip()
    first = (first_name or "").strip()
    middle = (middle_name or "").strip()
    initials = []
    if first: initials.append(first[0].upper() + ".")
    if middle: initials.append(middle[0].upper() + ".")
    init = "".join(initials).strip()
    if init and last: return f"{init} {last}".strip()
    if last: return last
    return init

def compute_day_intervals_in_month(db, employee_id: int, year: int, month: int):
    start_dt, end_dt = datetime(year, month, 1, 0, 0), datetime(*next_month(year, month), 1, 0, 0)
    duties = db.list_duties_for_period(employee_id, start_dt, end_dt)
    breaks_map = db.breaks_for_duty_ids([int(r["id"]) for r in duties])
    per_day = {}
    
    def get_shift(d): return int(d["is_shift"]) if "is_shift" in d.keys() and d["is_shift"] is not None else 0

    for r in duties:
        did = int(r["id"])
        is_shift = bool(get_shift(r))
        s0 = max(datetime.fromisoformat(r["start_dt"]), start_dt)
        e0 = min(datetime.fromisoformat(r["end_dt"]), end_dt)
        if s0 >= e0: continue

        parts = subtract_intervals((s0, e0), breaks_map.get(did, []))
        for s, e in parts:
            cur = s.date()
            last = (e - timedelta(seconds=1)).date() if e > s else s.date()
            while cur <= last:
                d0 = datetime.combine(cur, time(0, 0))
                d1 = d0 + timedelta(days=1)
                inter = intersect(s, e, d0, d1)
                if inter:
                    per_day.setdefault(cur, []).append((inter[0], inter[1], is_shift))
                cur += timedelta(days=1)

    merged_by_day = {}
    for d, intervals in per_day.items():
        shift_ints = [(s, e) for s, e, sh in intervals if sh]
        non_shift_ints = [(s, e) for s, e, sh in intervals if not sh]
        merged_shift = merge_intervals(shift_ints)
        merged_non_shift = merge_intervals(non_shift_ints)
        merged = [(s, e, True) for s, e in merged_shift] + [(s, e, False) for s, e in merged_non_shift]
        merged.sort(key=lambda x: x[0])
        merged_by_day[d] = merged

    return merged_by_day

class TemplateExporter:
    @staticmethod
    def _hide_daytime_intervals_in_workdays(db, intervals_by_day):
        out = {}
        for d0, intervals in (intervals_by_day or {}).items():
            
            # 1. Узнаем, рабочий ли это день по факту
            try:
                # Достаем из базы и галочку "рабочий", и "праздник"
                row = db.conn.execute("SELECT is_working, is_holiday FROM calendar_day WHERE date=?", (d0.isoformat(),)).fetchone()
                if row:
                    is_working = bool(int(row["is_working"]))
                    # Защита: вдруг колонка is_holiday пустая или ее нет
                    is_holiday = bool(int(row["is_holiday"] if "is_holiday" in row.keys() and row["is_holiday"] is not None else 0))
                    
                    # Настоящий рабочий день: галочка "рабочий" стоит, и это НЕ праздник
                    is_real_workday = is_working and not is_holiday
                else:
                    # Если дня в базе вообще нет, по умолчанию Пн-Пт — рабочие
                    is_real_workday = d0.weekday() < 5
            except Exception:
                # Если произошла ошибка БД, тоже считаем Пн-Пт рабочими
                is_real_workday = d0.weekday() < 5

            # 2. Если день НЕРАБОЧИЙ (Сб, Вс или Праздник) - выводим всё дежурство целиком!
            if not is_real_workday:
                out[d0] = intervals
                continue

            # 3. ДЕНЬ РАБОЧИЙ. Отрезаем дневные куски (с 06:00 до 22:00)
            from datetime import datetime, time, timedelta # На всякий случай импортируем внутри
            
            w0 = datetime.combine(d0, time(0, 0))
            w1 = datetime.combine(d0, time(6, 0))
            w2 = datetime.combine(d0, time(22, 0))
            w3 = datetime.combine(d0 + timedelta(days=1), time(0, 0))

            parts_shift, parts_non = [], []
            
            for a, b, is_shift in intervals:
                # Пересечение с утренним ночным куском 00:00 - 06:00
                inter1 = intersect(a, b, w0, w1)
                if inter1: 
                    if is_shift: parts_shift.append(inter1)
                    else: parts_non.append(inter1)
                    
                # Пересечение с вечерним ночным куском 22:00 - 24:00
                inter2 = intersect(a, b, w2, w3)
                if inter2: 
                    if is_shift: parts_shift.append(inter2)
                    else: parts_non.append(inter2)

            merged_shift = merge_intervals(parts_shift)
            merged_non = merge_intervals(parts_non)
            
            merged = [(s, e, True) for s, e in merged_shift] + [(s, e, False) for s, e in merged_non]
            merged.sort(key=lambda x: x[0])
            out[d0] = merged

        return out

    @staticmethod
    def _safe_cell(ws, row, col):
        from openpyxl.cell.cell import MergedCell
        cell = ws.cell(row=row, column=col)
        if not isinstance(cell, MergedCell): return cell
        for r in ws.merged_cells.ranges:
            if int(r.min_row) <= row <= int(r.max_row) and int(r.min_col) <= col <= int(r.max_col):
                return ws.cell(row=int(r.min_row), column=int(r.min_col))
        return None
        
    @staticmethod
    def month_name_ru(month: int) -> str:
        names = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
        return names[month - 1] if 1 <= month <= 12 else str(month)

    @staticmethod
    def _fmt_minutes_ru(minutes: int) -> str:
        sign = "-" if minutes < 0 else ""
        m = abs(int(minutes))
        h = m // 60
        mm = m % 60
        if mm == 0: return f"{sign}{h} ч."
        return f"{sign}{h} ч. {mm} мин."

    @staticmethod
    def _fmt_days_ru(days: int) -> str:
        sign = "-" if days < 0 else ""
        return f"{sign}{abs(int(days))} д."

    @staticmethod
    def _find_first_cell_with_value(ws, value: str):
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() == value:
                    return cell
        return None

    @staticmethod
    def _replace_markers_everywhere(ws, mapping: dict):
        for row in ws.iter_rows():
            for cell0 in row:
                if not isinstance(cell0.value, str): continue
                s = cell0.value
                s2 = s
                for k, v in mapping.items():
                    if k in s2: s2 = s2.replace(k, v)
                if s2 != s:
                    cell = TemplateExporter._safe_cell(ws, int(cell0.row), int(cell0.column))
                    if cell is not None: cell.value = s2

    @staticmethod
    def _collect_marker_columns_in_row(ws, row: int, markers: set):
        out = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, str):
                s = v.strip()
                if s in markers: out.setdefault(s, []).append(col)
        return out

    @staticmethod
    def _copy_row_style(ws, src_row: int, dst_row: int) -> None:
        from copy import copy
        from openpyxl.cell.cell import MergedCell
        if ws.row_dimensions.get(src_row) and ws.row_dimensions[src_row].height:
            ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
        for col in range(1, ws.max_column + 1):
            s = ws.cell(row=src_row, column=col)
            d = ws.cell(row=dst_row, column=col)
            if isinstance(d, MergedCell): continue
            d._style = copy(s._style)
            d.font = copy(s.font)
            d.border = copy(s.border)
            d.fill = copy(s.fill)
            d.number_format = s.number_format
            d.protection = copy(s.protection)
            d.alignment = copy(s.alignment)

    @staticmethod
    def _row_merge_patterns(ws, row: int) -> list[tuple[int, int]]:
        out = []
        try: ranges = list(ws.merged_cells.ranges)
        except Exception: ranges = []
        for r in ranges:
            if int(r.min_row) == row and int(r.max_row) == row:
                out.append((int(r.min_col), int(r.max_col)))
        return out

    @staticmethod
    def _apply_row_merge_patterns(ws, row: int, patterns: list[tuple[int, int]]) -> None:
        for min_c, max_c in patterns:
            try: ws.merge_cells(start_row=row, end_row=row, start_column=min_c, end_column=max_c)
            except Exception: pass

    # --- Функции сбора итогов по компенсациям ---
    @staticmethod
    def _fmt_dual_export(real, prev, unit_type="hours"):
        """Вспомогательная функция для форматирования строк экспорта"""
        if real == 0 and prev == 0: return "—"
        
        # Переводим числа в красивые строки (8 ч. или 1 д.)
        real_str = TemplateExporter._fmt_minutes_ru(real) if unit_type == "hours" else TemplateExporter._fmt_days_ru(real)
        prev_str = TemplateExporter._fmt_minutes_ru(prev) if unit_type == "hours" else TemplateExporter._fmt_days_ru(prev)
        
        if real > 0 and prev > 0:
            return f"{real_str} ({prev_str} за пред. год)"
        if real > 0:
            return real_str
        return f"{prev_str} за пред. год"

    @staticmethod
    def _sum_comp_rest_hours_minutes(db, employee_id: int, year: int, month: int) -> str:
        m0, m1 = date(year, month, 1), (date(year, month, 1) + timedelta(days=32)).replace(day=1)
        
        # 1. Считаем текущие часы (event_date != 1900)
        real = db.conn.execute("""
            SELECT COALESCE(SUM(amount_minutes),0) FROM compensation 
            WHERE employee_id=? AND unit IN ('hours', 'overtime') AND method='day_off' 
            AND event_date >= ? AND event_date < ? AND event_date != '1900-01-01'
        """, (employee_id, d_iso(m0), d_iso(m1))).fetchone()[0]
        
        # 2. Считаем часы прошлого года (event_date == 1900, смотрим на order_date)
        prev = db.conn.execute("""
            SELECT COALESCE(SUM(amount_minutes),0) FROM compensation 
            WHERE employee_id=? AND unit IN ('hours', 'overtime') AND method='day_off' 
            AND order_date >= ? AND order_date < ? AND event_date = '1900-01-01'
        """, (employee_id, d_iso(m0), d_iso(m1))).fetchone()[0]
        
        return TemplateExporter._fmt_dual_export(real, prev, "hours")

    @staticmethod
    def _count_comp_rest_days(db, employee_id: int, year: int, month: int) -> str:
        m0, m1 = date(year, month, 1), (date(year, month, 1) + timedelta(days=32)).replace(day=1)
        
        # 1. Считаем текущие дни (event_date != 1900)
        real = db.conn.execute("""
            SELECT COUNT(*) FROM comp_day_off_date 
            WHERE employee_id=? AND day_off_date >= ? AND day_off_date < ?
            AND compensation_id IN (SELECT id FROM compensation WHERE event_date != '1900-01-01' AND unit='days')
        """, (employee_id, d_iso(m0), d_iso(m1))).fetchone()[0]
        
        # 2. Считаем дни прошлого года (event_date == 1900)
        prev = db.conn.execute("""
            SELECT COUNT(*) FROM comp_day_off_date 
            WHERE employee_id=? AND day_off_date >= ? AND day_off_date < ?
            AND compensation_id IN (SELECT id FROM compensation WHERE event_date = '1900-01-01' AND unit='days')
        """, (employee_id, d_iso(m0), d_iso(m1))).fetchone()[0]
        
        return TemplateExporter._fmt_dual_export(real, prev, "days")

    @staticmethod
    def _money_comp_text_for_month(db, employee_id: int, year: int, month: int) -> str:
        m0 = date(year, month, 1)
        ny, nm = next_month(year, month)
        m1 = date(ny, nm, 1)
        
        # МАГИЯ: Мы добавили в запрос event_date, чтобы понять, "прошлый" это год или нет
        rows = db.conn.execute(
            """
            SELECT unit, amount_minutes, amount_days, order_no, order_date, event_date 
            FROM compensation 
            WHERE employee_id=? AND method='money' 
              AND (
                  (substr(order_date, 1, 7) = ?) -- Поиск по дате приказа
                  OR (event_date = '1900-01-01' AND substr(order_date, 1, 4) = ?) -- Или по спец. метке
              )
            ORDER BY order_date, order_no, id
            """,
            (employee_id, f"{year:04d}-{month:02d}", str(year)),
        ).fetchall()

        # Группируем, добавляя в ключ признак "прошлого года" (is_prev)
        grouped = {}
        for r in rows:
            order_no = (r["order_no"] or "").strip()
            order_date = (r["order_date"] or "").strip()
            is_prev = (r["event_date"] == "1900-01-01")
            
            # Ключ теперь включает флаг прошлого года, чтобы разделить строки
            key = (order_date, order_no, is_prev)
            g = grouped.setdefault(key, {"minutes": 0, "days": 0})
            
            if r["unit"] in ("hours", "overtime"): 
                g["minutes"] += int(r["amount_minutes"] or 0)
            else: 
                g["days"] += int(r["amount_days"] or 0)

        out = []
        # Сортируем по дате и номеру
        for (order_date, order_no, is_prev), g in sorted(grouped.items(), key=lambda x: (x[0][0], x[0][1])):
            parts = []
            if g["minutes"] > 0: parts.append(TemplateExporter._fmt_minutes_ru(g["minutes"]))
            if g["days"] > 0: parts.append(TemplateExporter._fmt_days_ru(g["days"]))
            
            amount_text = " и ".join(parts).strip()
            if not amount_text: continue
            
            # Добавляем твою приписку
            suffix = " (за пред. год)" if is_prev else ""
            od = fmt_date_iso(order_date)
            
            if order_no:
                out.append(f"{amount_text}{suffix} Приказом от {od} №{order_no}")
            else:
                out.append(f"{amount_text}{suffix} Приказом от {od}")
            
        res = "\n".join(out)
        return res if res else "—"

    @staticmethod
    def export(db, year: int, month: int, template_path: str, out_path: str, sheet_name: str = "Лист1"):
        openpyxl = ensure_openpyxl()
        from openpyxl.styles import Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.cell_range import CellRange

        # УМНОЕ ЧТЕНИЕ: Из памяти (PROD) или с диска (DEV)
        if template_path.startswith(":/"):
            qfile = QFile(template_path)
            if not qfile.open(QIODevice.ReadOnly):
                raise Exception("Не удалось загрузить шаблон из оперативной памяти!")
            template_stream = io.BytesIO(qfile.readAll().data())
            qfile.close()
            wb = openpyxl.load_workbook(template_stream)
        else:
            # Обычное чтение файла с диска для режима тестирования
            wb = openpyxl.load_workbook(template_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        start_cell = TemplateExporter._find_first_cell_with_value(ws, "{{EMP_START}}")
        if not start_cell: raise Exception("В шаблоне не найдена метка {{EMP_START}}.")

        start_row = int(start_cell.row)
        start_cell.value = None

        dept = db.get_department_settings()
        resp_fio = fio_initials_last(dept["resp_last_name"], dept["resp_first_name"], dept["resp_middle_name"])

        base_markers = {"{{EMP_NO}}", "{{EMP_FIO_RANK_POS}}", "{{OPEN_HOURS}}", "{{OPEN_OVERTIME}}", "{{OPEN_DAYS}}", "{{MONTH_HOURS}}", "{{MONTH_OVERTIME}}", "{{MONTH_DAYS}}", "{{COMP_REST_HOURS}}", "{{COMP_REST_DAYS}}", "{{COMP_MONEY}}", "{{BAL_END_HOURS}}", "{{BAL_END_OVERTIME}}", "{{BAL_END_DAYS}}"}
        day_markers = {f"{{{{DAY_{d:02d}}}}}" for d in range(1, 32)}
        all_markers = set(base_markers) | set(day_markers)

        marker_cols = TemplateExporter._collect_marker_columns_in_row(ws, start_row, all_markers)

        emps = db.list_employees_for_month(year, month, active_only=True, search="")
        n = len(emps)

        if n > 1:
            try: ws.insert_rows(start_row + 1, amount=n - 1)
            except TypeError:
                for _ in range(n - 1): ws.insert_rows(start_row + 1)

        merge_patterns = TemplateExporter._row_merge_patterns(ws, start_row)
        for r in range(start_row + 1, start_row + n):
            TemplateExporter._copy_row_style(ws, start_row, r)
            TemplateExporter._apply_row_merge_patterns(ws, r, merge_patterns)

        TemplateExporter._replace_markers_everywhere(ws, {
            "{{HDR_DEPARTMENT}}": dept["department_name"],
            "{{HDR_MONTH_NAME}}": TemplateExporter.month_name_ru(month),
            "{{HDR_YEAR}}": str(year),
        })

        ny, nm = next_month(year, month)
        last_day = (date(ny, nm, 1) - timedelta(days=1)).day
        m0 = date(year, month, 1)
        m1 = date(ny, nm, 1)
        work_map = db.get_calendar_month(d_iso(m0), d_iso(m1 - timedelta(days=1)))

        # Желтая заливка для выходных дней (Solid fill, цвет FFFF00)
        yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

        # Перебираем все возможные дни от 1 до 31
        for d in range(1, 32):
            marker = f"{{{{DAY_{d:02d}}}}}"
            cols = marker_cols.get(marker, [])
            
            # 1. Если день больше, чем дней в месяце (например 31-е февраля) -> прячем колонку
            if d > last_day:
                for col in cols: 
                    ws.column_dimensions[get_column_letter(col)].hidden = True
            
            # 2. Если день существует, проверяем, выходной ли он
            else:
                day_date = date(year, month, d)
                
                # Получаем статус дня из базы. Если в базе пусто, смотрим на календарь (Сб и Вс - выходные)
                is_working = work_map.get(day_date, day_date.weekday() < 5)
                
                # Если день нерабочий (выходной) -> красим всю колонку сплошняком
                if not is_working:
                    for col in cols:
                        # Начинаем красить за 3 строки до сотрудников и идем вниз до конца списка.
                        # max(1, ...) защищает от ошибки, если шапка случайно окажется на самом верху листа.
                        for r in range(max(1, start_row - 3), start_row + n):
                            c = TemplateExporter._safe_cell(ws, r, col)
                            if c: 
                                c.fill = yellow_fill

        for i, emp in enumerate(emps):
            row = start_row + i
            eid = int(emp["id"])
            fio_rank_pos = "\n".join(x for x in [f"{emp['last_name']} {emp['first_name']} {emp['middle_name'] or ''}".strip(), (emp["rank"] or "").strip(), "", (emp["position"] or "").strip()] if x).rstrip()

            # ... здесь код выше (fio_rank_pos = ...)
            
            # Эти строки должны быть ВНУТРИ цикла for i, emp in enumerate(emps):
            summ = compute_month_summary(db, eid, year, month)
            comp_h = TemplateExporter._sum_comp_rest_hours_minutes(db, eid, year, month)
            comp_d = TemplateExporter._count_comp_rest_days(db, eid, year, month)
            
            # --- ФОРМИРУЕМ ДОПОЛНИТЕЛЬНЫЕ СТРОКИ ДЛЯ СМЕН ---
            # Ночные часы внутри графика — добавляем к MONTH_HOURS
            night_shift_str = ""
            shift_night_val = summ.get("shift_night", 0) or 0
            if is_employee_shift(db, eid) and shift_night_val > 0:
                night_shift_str = f"\n({TemplateExporter._fmt_minutes_ru(shift_night_val)})"

            # Праздничные часы внутри графика — добавляем к MONTH_DAYS
            holiday_shift_str = ""
            shift_holiday_val = summ.get("shift_holiday", 0) or 0
            if is_employee_shift(db, eid) and shift_holiday_val > 0:
                holiday_shift_str = f"\n({TemplateExporter._fmt_minutes_ru(shift_holiday_val)})"
            
            # --- Вспомогательная функция для остатков в Excel (как в панели) ---
            def fmt_bal_excel(val, prev_val, is_days=False):
                main = (TemplateExporter._fmt_days_ru(val) if is_days else TemplateExporter._fmt_minutes_ru(val))
                if prev_val == 0: return main
                prev = (TemplateExporter._fmt_days_ru(prev_val) if is_days else TemplateExporter._fmt_minutes_ru(prev_val))
                return f"{main} ({prev})"

            values = {
                "{{EMP_NO}}": str(i + 1),
                "{{EMP_FIO_RANK_POS}}": fio_rank_pos,
                
                # Остатки на начало (с Эталоном в скобках)
                "{{OPEN_HOURS}}": fmt_bal_excel(summ["start_hours"], summ["prev_h_start"]),
                "{{OPEN_OVERTIME}}": fmt_bal_excel(summ["start_overtime"], summ["prev_o_start"]),
                "{{OPEN_DAYS}}": fmt_bal_excel(summ["start_days"], summ["prev_d_start"], True),
                
                "{{MONTH_HOURS}}": TemplateExporter._fmt_minutes_ru(summ["acc_hours"]) + night_shift_str,
                "{{MONTH_OVERTIME}}": TemplateExporter._fmt_minutes_ru(summ["acc_overtime"]),
                "{{MONTH_DAYS}}": TemplateExporter._fmt_days_ru(summ["acc_days"]) + holiday_shift_str,
                
                # Компенсации (теперь со строкой "за пред. год")
                "{{COMP_REST_HOURS}}": comp_h, 
                "{{COMP_REST_DAYS}}": comp_d,
                "{{COMP_MONEY}}": TemplateExporter._money_comp_text_for_month(db, eid, year, month),
                
                # Остатки на конец (с Эталоном в скобках)
                "{{BAL_END_HOURS}}": fmt_bal_excel(summ["end_hours"], summ["prev_h_end"]),
                "{{BAL_END_OVERTIME}}": fmt_bal_excel(summ["end_overtime"], summ["prev_o_end"]),
                "{{BAL_END_DAYS}}": fmt_bal_excel(summ["end_days"], summ["prev_d_end"], True),
            }

            intervals_by_day = compute_day_intervals_in_month(db, eid, year, month)
            if not is_employee_shift(db, eid):
                intervals_by_day = TemplateExporter._hide_daytime_intervals_in_workdays(db, intervals_by_day)

            statuses = db.get_statuses_for_period(eid, f"{year:04d}-{month:02d}-01", f"{ny:04d}-{nm:02d}-01")

            # --- СБОР ВЫХОДНЫХ (КОМПЕНСАЦИЙ) ПО ДНЯМ ДЛЯ КАЛЕНДАРЯ ---
            comp_by_day = {}
            rows_h = db.conn.execute("SELECT event_date, amount_minutes FROM compensation WHERE employee_id=? AND unit IN ('hours', 'overtime') AND method='day_off' AND event_date IS NOT NULL AND event_date >= ? AND event_date < ?", (eid, d_iso(m0), d_iso(m1))).fetchall()
            for rr in rows_h:
                try: 
                    d0 = d_parse(rr["event_date"])
                    if int(rr["amount_minutes"] or 0) > 0: comp_by_day.setdefault(d0, []).append("В")
                except Exception: pass

            rows_d = db.conn.execute("SELECT cd.day_off_date FROM comp_day_off_date cd JOIN compensation c ON c.id = cd.compensation_id WHERE cd.employee_id=? AND cd.day_off_date >= ? AND cd.day_off_date < ?", (eid, d_iso(m0), d_iso(m1))).fetchall()
            for rr in rows_d:
                try: 
                    d0 = d_parse(rr["day_off_date"])
                    comp_by_day.setdefault(d0, []).append("В")
                except Exception: pass

            # Заполняем ячейки
            for marker, cols in marker_cols.items():
                if marker in values:
                    for col in cols:
                        c = TemplateExporter._safe_cell(ws, row, col)
                        if c:
                            c.value = values[marker]
                            if marker == "{{EMP_FIO_RANK_POS}}": c.alignment = Alignment(wrap_text=True, vertical="top")

            # Заполняем дни
            for d in range(1, 32):
                marker = f"{{{{DAY_{d:02d}}}}}"
                cols = marker_cols.get(marker, [])
                if not cols: continue
                
                txt = ""
                if d <= last_day:
                    day_date = date(year, month, d)
                    
                    # 1. Статус (К, Б, О)
                    day_status = statuses.get(day_date, "")
                    if day_status: txt = day_status
                    
                    # 2. Дежурства
                    intervals = intervals_by_day.get(day_date, [])
                    if intervals:
                        parts = [f"{a.strftime('%H:%M')}-{b.strftime('%H:%M')}" for a, b, _ in intervals]
                        duty_txt = "\n".join(parts)
                        txt = f"{txt}\n{duty_txt}" if txt else duty_txt
                        
                    # 3. Компенсации (Буква "В")
                    comps = comp_by_day.get(day_date, [])
                    if comps:
                        comp_txt = "\n".join(comps)
                        txt = f"{txt}\n{comp_txt}" if txt else comp_txt

                for col in cols:
                    c = TemplateExporter._safe_cell(ws, row, col)
                    if c:
                        c.value = txt
                        if txt: c.alignment = c.alignment.copy(wrap_text=True)

        # Подписи и область печати
        sign_rows = []
        for marker in ("{{SIGN_POS}}", "{{SIGN_RANK}}", "{{SIGN_FIO}}"):
            for row0 in ws.iter_rows():
                for cell0 in row0:
                    if isinstance(cell0.value, str) and cell0.value.strip() == marker:
                        sign_rows.append(int(cell0.row))

        bottom_row = max(sign_rows) + 2 if sign_rows else ws.max_row + 2
        
        TemplateExporter._replace_markers_everywhere(ws, {
            "{{SIGN_POS}}": dept["resp_position"] or "",
            "{{SIGN_RANK}}": dept["resp_rank"] or "",
            "{{SIGN_FIO}}": resp_fio,
        })

        # Увеличиваем область печати вниз
        if ws.print_area:
            pa_str = ws.print_area if isinstance(ws.print_area, str) else ws.print_area[0]
            pa_str = pa_str.split(",")[0].split("!")[-1].replace("$", "")
            try:
                cr = CellRange(pa_str)
                ws.print_area = f"{get_column_letter(int(cr.min_col))}{int(cr.min_row)}:{get_column_letter(int(cr.max_col))}{bottom_row}"
            except Exception:
                ws.print_area = f"A1:{get_column_letter(ws.max_column)}{bottom_row}"
        else:
            ws.print_area = f"A1:{get_column_letter(ws.max_column)}{bottom_row}"

        # Удаляем оставшиеся метки из первой строки
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=start_row, column=col)
            if isinstance(cell.value, str) and cell.value.strip() in all_markers:
                c = TemplateExporter._safe_cell(ws, start_row, col)
                if c: c.value = None

        wb.save(out_path)